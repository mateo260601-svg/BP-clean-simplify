from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from historical_accounts_engine import normalize_historical_accounts
from institutional_debt_engine_v12 import add_v12_debt_engine, default_debt_stack_for_saas


COLORS = {
    "navy": "0D1B3E",
    "blue": "0000FF",
    "input": "EBF5FF",
    "output": "EAF2FF",
    "white": "FFFFFF",
    "line": "D0D5DD",
    "grey": "64748B",
}

FIRST_COL = 8


def _fill(c):
    return PatternFill("solid", fgColor=COLORS[c])


def _font(c="000000", bold=False, size=9, italic=False):
    return Font(name="Calibri", color=COLORS.get(c, c), bold=bold, size=size, italic=italic)


def _border():
    s = Side(style="thin", color=COLORS["line"])
    return Border(left=s, right=s, top=s, bottom=s)


def fmt_money():
    return '#,##0.0;[Red](#,##0.0);"-"'


def fmt_pct():
    return '0.0%;[Red](0.0%);"-"'


def fmt_mult():
    return '0.0x;[Red](0.0x);"-"'


def setup(ws, freeze="H6", tab="0D1B3E"):
    ws.sheet_properties.tabColor = tab
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze
    ws.sheet_view.zoomScale = 85
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 2
    for c in range(FIRST_COL, 60):
        ws.column_dimensions[get_column_letter(c)].width = 13


def title(ws, text, subtitle=""):
    ws["B2"] = text
    ws["B2"].font = _font("navy", True, 16)
    if subtitle:
        ws["B3"] = subtitle
        ws["B3"].font = _font("grey", False, 9, True)


def header(cell, value):
    cell.value = value
    cell.fill = _fill("navy")
    cell.font = _font("white", True, 9)
    cell.border = _border()
    cell.alignment = Alignment(horizontal="center")


def label(ws, row, text, unit="", source="", total=False):
    ws.cell(row, 2).value = text
    ws.cell(row, 3).value = unit
    ws.cell(row, 4).value = source
    for c in [2, 3, 4, 5, 6]:
        ws.cell(row, c).border = _border()
        ws.cell(row, c).font = _font("navy", bold=total)
        if total:
            ws.cell(row, c).fill = _fill("output")


def input_cell(cell, value, number_format=None):
    cell.value = value
    cell.fill = _fill("input")
    cell.font = _font("blue")
    cell.border = _border()
    cell.alignment = Alignment(horizontal="right")
    if number_format:
        cell.number_format = number_format


def formula_cell(cell, formula, number_format=None, output=False):
    if not str(formula).startswith("="):
        formula = "=" + str(formula)
    cell.value = formula
    cell.fill = _fill("output" if output else "white")
    cell.font = _font("000000", bold=output)
    cell.border = _border()
    cell.alignment = Alignment(horizontal="right")
    if number_format:
        cell.number_format = number_format


def _periods_from_historical_and_forecast(historical, start_year, forecast_years):
    periods = []
    for p in historical.get("periods", []) or []:
        periods.append(str(p.get("period", "HIST")))
    for i in range(int(forecast_years)):
        periods.append(f"FY {int(start_year)+i}")
    return periods


def build_v12_formula_model(intake: Dict[str, Any], historical: Dict[str, Any], output_path: str) -> str:
    """
    Complete V12 model builder:
    - historical accounts
    - assumptions
    - revenue build
    - working capital
    - financial statements
    - institutional debt engine V12
    """
    intake = dict(intake or {})
    historical = historical or {"periods": [], "evidence": [], "warnings": []}

    start_year = int(float(intake.get("start_year", 2026) or 2026))
    forecast_years = int(float(intake.get("forecast_years", 5) or 5))
    hist_periods = historical.get("periods", []) or []
    periods = _periods_from_historical_and_forecast(historical, start_year, forecast_years)
    hist_n = len(hist_periods)

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    # Cover
    ws = wb.active
    ws.title = "Cover"
    setup(ws)
    title(ws, "MG Advisory Institutional BP V12", "Historical accounts + formula model + institutional debt engine")
    cover = [
        ("Company", intake.get("company_name", "Target Company")),
        ("Currency", intake.get("currency", "EUR")),
        ("Sector", intake.get("sector", "Industrial / Manufacturing")),
        ("Deal Type", intake.get("deal_type", "M&A")),
        ("Historical Periods Loaded", hist_n),
        ("Debt Engine", "V12 Institutional"),
    ]
    r = 6
    for k, v in cover:
        label(ws, r, k, "", "Input/System")
        input_cell(ws.cell(r, FIRST_COL), v)
        r += 1

    # Historical Inputs
    ws = wb.create_sheet("Historical Inputs")
    setup(ws, tab="1F5C8B")
    title(ws, "Historical Inputs", "Extracted historical accounts. Blue cells may be corrected manually.")
    for i, p in enumerate(hist_periods):
        c = FIRST_COL + i
        header(ws.cell(4, c), p.get("period", f"HIST {i+1}"))
        header(ws.cell(5, c), "Historical")
    hist_metrics = [
        ("Revenue", "revenue"),
        ("Gross Profit", "gross_profit"),
        ("EBITDA", "ebitda"),
        ("EBIT", "ebit"),
        ("Net Income", "net_income"),
        ("Cash", "cash"),
        ("Debt", "debt"),
        ("Net Debt", "net_debt"),
        ("Capex", "capex"),
        ("Receivables", "receivables"),
        ("Inventory", "inventory"),
        ("Payables", "payables"),
        ("Working Capital", "working_capital"),
    ]
    for r, (name, key) in enumerate(hist_metrics, start=7):
        label(ws, r, name, "000s", "Extracted", total=name in ["Revenue", "EBITDA", "Net Debt"])
        for i, p in enumerate(hist_periods):
            input_cell(ws.cell(r, FIRST_COL+i), p.get(key, ""), fmt_money())

    # Source Evidence
    ws = wb.create_sheet("Source Evidence")
    setup(ws, freeze="A6", tab="64748B")
    title(ws, "Source Evidence", "Audit trail for extracted metrics.")
    headers = ["File", "Period", "Metric", "Value", "Source", "Confidence", "Raw Label"]
    for c, h in enumerate(headers, start=1):
        header(ws.cell(5, c), h)
    for r, e in enumerate(historical.get("evidence", []) or [], start=6):
        vals = [e.get("file"), e.get("period"), e.get("metric"), e.get("value"), e.get("source"), e.get("confidence"), e.get("raw_label")]
        for c, v in enumerate(vals, start=1):
            ws.cell(r, c).value = v
            ws.cell(r, c).border = _border()
            if c == 4:
                ws.cell(r, c).number_format = fmt_money()
            if c == 6:
                ws.cell(r, c).number_format = fmt_pct()

    # Assumptions
    ws = wb.create_sheet("Assumptions")
    setup(ws, tab="0D1B3E")
    title(ws, "Assumptions", "Only blue cells are hardcoded assumptions.")
    assumptions = [
        ("Company", "Text", intake.get("company_name", "Target Company")),
        ("Currency", "Text", intake.get("currency", "EUR")),
        ("Start Year", "Year", start_year),
        ("Forecast Years", "Years", forecast_years),
        ("Revenue Growth", "%", intake.get("revenue_growth", 0.073)),
        ("Gross Margin", "%", intake.get("gross_margin", 0.391)),
        ("EBITDA Margin", "%", intake.get("ebitda_margin", 0.229)),
        ("Tax Rate", "%", intake.get("tax_rate", 0.25)),
        ("DSO", "Days", intake.get("dso", 55)),
        ("DIO", "Days", intake.get("dio", 40)),
        ("DPO", "Days", intake.get("dpo", 55)),
        ("Maintenance Capex % Revenue", "%", intake.get("maintenance_capex_pct_revenue", 0.055)),
        ("Opening Cash", "000s", intake.get("cash", historical.get("latest", {}).get("cash", 18700))),
        ("Opening Debt", "000s", intake.get("debt", historical.get("latest", {}).get("debt", 201000))),
    ]
    arow = {}
    for r, (name, unit, value) in enumerate(assumptions, start=6):
        label(ws, r, name, unit, "Input")
        input_cell(ws.cell(r, FIRST_COL), value, fmt_pct() if unit == "%" else fmt_money() if unit == "000s" else None)
        arow[name] = r

    # Revenue Build
    ws = wb.create_sheet("Revenue Build")
    setup(ws, tab="1F5C8B")
    title(ws, "Revenue Build", "Historical values are linked; forecasts are formulas.")
    for i, p in enumerate(periods):
        c = FIRST_COL + i
        header(ws.cell(4, c), p)
        header(ws.cell(5, c), "Historical" if i < hist_n else "Forecast")
    rows = {"Revenue":7, "Growth":8, "Gross Profit":9, "Gross Margin":10, "EBITDA":11, "EBITDA Margin":12}
    for name, r in rows.items():
        label(ws, r, name, "%" if "Margin" in name or name=="Growth" else "000s", "Formula", total=name in ["Revenue", "EBITDA"])
    for i in range(len(periods)):
        c = FIRST_COL+i
        col = get_column_letter(c)
        if i < hist_n:
            hcol = get_column_letter(FIRST_COL+i)
            formula_cell(ws.cell(7, c), f"='Historical Inputs'!{hcol}7", fmt_money(), True)
            formula_cell(ws.cell(8, c), "=0" if i==0 else f"=IFERROR({col}7/{get_column_letter(c-1)}7-1,0)", fmt_pct())
            formula_cell(ws.cell(9, c), f"='Historical Inputs'!{hcol}8", fmt_money())
            formula_cell(ws.cell(10, c), f"=IFERROR({col}9/{col}7,0)", fmt_pct())
            formula_cell(ws.cell(11, c), f"='Historical Inputs'!{hcol}9", fmt_money(), True)
            formula_cell(ws.cell(12, c), f"=IFERROR({col}11/{col}7,0)", fmt_pct())
        else:
            prev = get_column_letter(c-1)
            formula_cell(ws.cell(7, c), f"={prev}7*(1+'Assumptions'!$H${arow['Revenue Growth']})", fmt_money(), True)
            formula_cell(ws.cell(8, c), f"='Assumptions'!$H${arow['Revenue Growth']}", fmt_pct())
            formula_cell(ws.cell(9, c), f"={col}7*'Assumptions'!$H${arow['Gross Margin']}", fmt_money())
            formula_cell(ws.cell(10, c), f"=IFERROR({col}9/{col}7,0)", fmt_pct())
            formula_cell(ws.cell(11, c), f"={col}7*'Assumptions'!$H${arow['EBITDA Margin']}", fmt_money(), True)
            formula_cell(ws.cell(12, c), f"=IFERROR({col}11/{col}7,0)", fmt_pct())

    # Working Capital
    ws = wb.create_sheet("Working Capital")
    setup(ws, tab="2C5282")
    title(ws, "Working Capital", "Historical / forecast working capital.")
    for i, p in enumerate(periods):
        c = FIRST_COL + i
        header(ws.cell(4, c), p)
        header(ws.cell(5, c), "Historical" if i < hist_n else "Forecast")
    wc_rows = {"Receivables":7, "Inventory":8, "Payables":9, "NWC":10, "Change in NWC":11}
    for name, r in wc_rows.items():
        label(ws, r, name, "000s", "Formula", total=name in ["NWC", "Change in NWC"])
    for i in range(len(periods)):
        c = FIRST_COL+i
        col = get_column_letter(c)
        if i < hist_n:
            hcol = get_column_letter(FIRST_COL+i)
            formula_cell(ws.cell(7, c), f"='Historical Inputs'!{hcol}11", fmt_money())
            formula_cell(ws.cell(8, c), f"='Historical Inputs'!{hcol}12", fmt_money())
            formula_cell(ws.cell(9, c), f"='Historical Inputs'!{hcol}13", fmt_money())
            formula_cell(ws.cell(10, c), f"=IF('Historical Inputs'!{hcol}14<>0,'Historical Inputs'!{hcol}14,{col}7+{col}8-{col}9)", fmt_money(), True)
        else:
            formula_cell(ws.cell(7, c), f"='Revenue Build'!{col}7*'Assumptions'!$H${arow['DSO']}/365", fmt_money())
            formula_cell(ws.cell(8, c), f"='Revenue Build'!{col}7*'Assumptions'!$H${arow['DIO']}/365", fmt_money())
            formula_cell(ws.cell(9, c), f"='Revenue Build'!{col}7*'Assumptions'!$H${arow['DPO']}/365", fmt_money())
            formula_cell(ws.cell(10, c), f"={col}7+{col}8-{col}9", fmt_money(), True)
        formula_cell(ws.cell(11, c), f"={col}10" if i==0 else f"={col}10-{get_column_letter(c-1)}10", fmt_money())

    # Financial Statements - before debt engine, so debt engine can link to it.
    ws = wb.create_sheet("Financial Statements")
    setup(ws, tab="006400")
    title(ws, "Financial Statements", "Integrated core statements before full debt module links.")
    for i, p in enumerate(periods):
        c = FIRST_COL+i
        header(ws.cell(4, c), p)
        header(ws.cell(5, c), "Historical" if i < hist_n else "Forecast")
    fs_rows = {
        "Revenue": 7, "Gross Profit": 8, "EBITDA": 9, "EBITDA Margin": 10,
        "Cash Interest": 11, "PBT": 12, "Tax": 13, "Net Income": 14,
        "FCF Before Debt Paydown": 16, "Closing Cash": 17, "Closing Debt": 18,
        "Net Debt": 19, "Net Debt / EBITDA": 20, "Interest Cover": 21
    }
    for name, r in fs_rows.items():
        unit = "%" if "Margin" in name else "x" if "Cover" in name or "/" in name else "000s"
        label(ws, r, name, unit, "Formula", total=name in ["EBITDA", "Net Debt", "Net Debt / EBITDA"])
    for i in range(len(periods)):
        c = FIRST_COL+i
        col = get_column_letter(c)
        formula_cell(ws.cell(7, c), f"='Revenue Build'!{col}7", fmt_money())
        formula_cell(ws.cell(8, c), f"='Revenue Build'!{col}9", fmt_money())
        formula_cell(ws.cell(9, c), f"='Revenue Build'!{col}11", fmt_money(), True)
        formula_cell(ws.cell(10, c), f"='Revenue Build'!{col}12", fmt_pct())
        # After debt engine exists, these could be linked manually or by post-processing. Use formulas that become valid once debt schedule added.
        formula_cell(ws.cell(11, c), f"=IFERROR('Debt Outputs'!{col}17,0)", fmt_money())
        formula_cell(ws.cell(12, c), f"={col}9-{col}11", fmt_money())
        formula_cell(ws.cell(13, c), f"=MAX(0,{col}12*'Assumptions'!$H${arow['Tax Rate']})", fmt_money())
        formula_cell(ws.cell(14, c), f"={col}12-{col}13", fmt_money())
        formula_cell(ws.cell(16, c), f"={col}9-{col}13-'Working Capital'!{col}11-('Revenue Build'!{col}7*'Assumptions'!$H${arow['Maintenance Capex % Revenue']})", fmt_money())
        if i == 0:
            formula_cell(ws.cell(17, c), f"='Assumptions'!$H${arow['Opening Cash']}", fmt_money())
            formula_cell(ws.cell(18, c), f"='Assumptions'!$H${arow['Opening Debt']}", fmt_money())
        else:
            formula_cell(ws.cell(17, c), f"=IFERROR('Debt Waterfall'!{col}19,{get_column_letter(c-1)}17+{col}16-{col}11)", fmt_money(), True)
            formula_cell(ws.cell(18, c), f"=IFERROR('Debt Outputs'!{col}7,{get_column_letter(c-1)}18)", fmt_money(), True)
        formula_cell(ws.cell(19, c), f"={col}18-{col}17", fmt_money(), True)
        formula_cell(ws.cell(20, c), f"=IFERROR({col}19/{col}9,0)", fmt_mult(), True)
        formula_cell(ws.cell(21, c), f"=IFERROR({col}9/{col}11,99)", fmt_mult())

    # Add debt engine
    default_stack = intake.get("debt_stack") or default_debt_stack_for_saas()
    add_v12_debt_engine(
        wb,
        periods=periods,
        historical_period_count=hist_n,
        financials_sheet_name="Financial Statements",
        revenue_sheet_name="Revenue Build",
        working_capital_sheet_name="Working Capital",
        start_year=start_year,
        forecast_years=forecast_years,
        default_debt_stack=default_stack,
    )

    wb.save(output_path)
    return output_path
