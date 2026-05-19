"""
extractor_llm.py  —  LLM-powered financial statement extractor
Uses Claude API to understand ANY format, ANY language, ANY structure.

Flow:
  1. Raw extraction   — pull text/tables from file (openpyxl / pdfplumber / csv)
  2. Context chunking — split into digestible chunks if large
  3. Claude call      — structured JSON extraction via API
  4. Validation       — sanity-check and scale detection
  5. Assumptions      — derive wizard pre-fills from extracted data
"""

import io, json, re, os
from typing import Dict, List, Optional, Tuple

ANTHROPIC_API_URL = "https://api.anthropic.com/v1/messages"

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — RAW TEXT/TABLE EXTRACTION  (format-agnostic)
# ══════════════════════════════════════════════════════════════════════════════

def _extract_raw_text_excel(file_bytes: bytes) -> str:
    """Extract all text from Excel — preserve row/column structure as TSV."""
    import openpyxl
    wb    = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 3 or ws.max_column < 2:
            continue
        parts.append(f"\n=== Sheet: {sheet_name} ===")
        rows = []
        for row in ws.iter_rows(values_only=True, max_row=min(ws.max_row, 120)):
            cells = [str(c).strip() if c is not None else "" for c in row]
            # Skip fully empty rows
            if all(c == "" for c in cells):
                continue
            rows.append("\t".join(cells))
        parts.append("\n".join(rows[:100]))  # max 100 rows per sheet

    return "\n".join(parts)[:12000]  # cap at 12k chars for LLM context


def _extract_raw_text_csv(file_bytes: bytes) -> str:
    import csv
    text   = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows   = ["\t".join(r) for r in reader if any(c.strip() for c in r)]
    return "\n".join(rows[:150])[:8000]


def _extract_raw_text_pdf(file_bytes: bytes) -> str:
    try:
        import pdfplumber
        parts = []
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages[:12]:
                # Try tables first (more structured)
                tables = page.extract_tables()
                for table in tables:
                    if not table: continue
                    rows = ["\t".join(c if c else "" for c in row) for row in table]
                    parts.append("\n".join(rows))
                # Also grab raw text for context
                text = page.extract_text()
                if text:
                    parts.append(text[:800])
        return "\n".join(parts)[:12000]
    except ImportError:
        return "PDF parser (pdfplumber) not installed."


def extract_raw(file_bytes: bytes, filename: str) -> str:
    ext = filename.lower().rsplit(".", 1)[-1]
    if ext in ("xlsx", "xls", "xlsm"):
        return _extract_raw_text_excel(file_bytes)
    elif ext == "csv":
        return _extract_raw_text_csv(file_bytes)
    elif ext == "pdf":
        return _extract_raw_text_pdf(file_bytes)
    return ""


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — CLAUDE STRUCTURED EXTRACTION PROMPT
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = """You are an expert financial analyst specialising in extracting structured financial data from raw financial statements.
You receive raw text extracted from Excel, CSV, or PDF files that may contain P&L, Balance Sheet, or Cash Flow statements.
The format may be irregular, multi-language (EN/FR/DE/AR/IT), with merged cells, subtotals, notes, or unusual layouts.

Your job: identify and extract key financial metrics, return ONLY valid JSON, no explanation.

Rules:
- Map line items intelligently regardless of language or naming convention
- Detect the reporting unit automatically (thousands, millions, etc.) and normalise to thousands
- If a value appears negative in the source (e.g. cost shown as negative), preserve sign
- Report annual figures only (aggregate monthly if needed)
- If a field is not found, omit it from the JSON (do not return null/0 for missing data)
- Return ONLY the JSON object, no markdown fences, no explanation
"""

EXTRACTION_PROMPT = """Extract financial data from this raw financial statement text.

RAW DATA:
{raw_text}

Return a JSON object with this exact schema (omit fields you cannot find with confidence):

{{
  "detected_unit": "thousands | millions | actual",
  "detected_currency": "USD | EUR | GBP | AED | other",
  "detected_language": "EN | FR | DE | AR | IT | other",
  "periods": ["2022", "2023", "2024"],
  "income_statement": {{
    "gross_revenue":     {{"2022": 0, "2023": 0, "2024": 0}},
    "freight_deductions":{{"2022": 0, "2023": 0, "2024": 0}},
    "net_revenue":       {{"2022": 0, "2023": 0, "2024": 0}},
    "direct_materials":  {{"2022": 0, "2023": 0, "2024": 0}},
    "utilities":         {{"2022": 0, "2023": 0, "2024": 0}},
    "gross_profit":      {{"2022": 0, "2023": 0, "2024": 0}},
    "salary_wages":      {{"2022": 0, "2023": 0, "2024": 0}},
    "other_opex":        {{"2022": 0, "2023": 0, "2024": 0}},
    "ebitda_adj":        {{"2022": 0, "2023": 0, "2024": 0}},
    "ebitda_reported":   {{"2022": 0, "2023": 0, "2024": 0}},
    "depreciation":      {{"2022": 0, "2023": 0, "2024": 0}},
    "ebit":              {{"2022": 0, "2023": 0, "2024": 0}},
    "interest_expense":  {{"2022": 0, "2023": 0, "2024": 0}},
    "pbt":               {{"2022": 0, "2023": 0, "2024": 0}},
    "tax":               {{"2022": 0, "2023": 0, "2024": 0}},
    "net_income":        {{"2022": 0, "2023": 0, "2024": 0}}
  }},
  "balance_sheet": {{
    "total_assets":       {{"2022": 0, "2023": 0, "2024": 0}},
    "ppe_net":            {{"2022": 0, "2023": 0, "2024": 0}},
    "inventories":        {{"2022": 0, "2023": 0, "2024": 0}},
    "trade_receivables":  {{"2022": 0, "2023": 0, "2024": 0}},
    "cash":               {{"2022": 0, "2023": 0, "2024": 0}},
    "total_equity":       {{"2022": 0, "2023": 0, "2024": 0}},
    "financial_debt":     {{"2022": 0, "2023": 0, "2024": 0}},
    "trade_payables":     {{"2022": 0, "2023": 0, "2024": 0}}
  }},
  "cash_flow": {{
    "operating_cf":  {{"2022": 0, "2023": 0, "2024": 0}},
    "capex":         {{"2022": 0, "2023": 0, "2024": 0}},
    "free_cash_flow":{{"2022": 0, "2023": 0, "2024": 0}}
  }},
  "operational_kpis": {{
    "sales_volume_mt":    {{"2022": 0, "2023": 0, "2024": 0}},
    "avg_price_per_mt":   {{"2022": 0, "2023": 0, "2024": 0}},
    "spread_per_mt":      {{"2022": 0, "2023": 0, "2024": 0}},
    "capacity_utilisation":{{"2022": 0, "2023": 0, "2024": 0}}
  }},
  "notes": "Any important observations about the data quality, anomalies, or structure"
}}"""


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — CLAUDE API CALL
# ══════════════════════════════════════════════════════════════════════════════

async def call_claude_extract(raw_text: str) -> Dict:
    """Call Claude API asynchronously to extract structured financial data."""
    import httpx

    prompt = EXTRACTION_PROMPT.format(raw_text=raw_text[:10000])

    payload = {
        "model":      "claude-sonnet-4-20250514",
        "max_tokens": 4096,
        "system":     SYSTEM_PROMPT,
        "messages":   [{"role": "user", "content": prompt}],
    }

    async with httpx.AsyncClient(timeout=60.0) as client:
        r = await client.post(
            ANTHROPIC_API_URL,
            headers={"Content-Type": "application/json"},
            json=payload,
        )
        r.raise_for_status()
        data = r.json()

    # Extract text from response
    content = data.get("content", [])
    raw_json = ""
    for block in content:
        if block.get("type") == "text":
            raw_json += block["text"]

    # Parse JSON — strip any accidental markdown fences
    raw_json = raw_json.strip()
    if raw_json.startswith("```"):
        raw_json = re.sub(r"^```[a-z]*\n?", "", raw_json)
        raw_json = re.sub(r"\n?```$", "", raw_json)

    return json.loads(raw_json)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — VALIDATION & SCALE NORMALISATION
# ══════════════════════════════════════════════════════════════════════════════

def _to_k(value: float, unit: str) -> float:
    """Normalise to thousands."""
    if unit == "millions":
        return value * 1000
    elif unit == "actual":
        return value / 1000
    return value  # already thousands


def normalise(extracted: Dict) -> Dict:
    """Normalise all values to thousands."""
    unit = extracted.get("detected_unit", "thousands")
    if unit == "thousands":
        return extracted  # no conversion needed

    def recurse(obj):
        if isinstance(obj, dict):
            return {k: recurse(v) for k, v in obj.items()}
        elif isinstance(obj, (int, float)):
            return round(_to_k(obj, unit), 1)
        return obj

    for section in ["income_statement", "balance_sheet", "cash_flow"]:
        if section in extracted:
            extracted[section] = recurse(extracted[section])

    extracted["detected_unit"] = "thousands"
    return extracted


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — DERIVE WIZARD ASSUMPTIONS
# ══════════════════════════════════════════════════════════════════════════════

def derive_assumptions(extracted: Dict) -> Dict:
    """Compute suggested wizard pre-fills from extracted data."""
    sugg  = {}
    is_   = extracted.get("income_statement", {})
    bs_   = extracted.get("balance_sheet", {})
    cf_   = extracted.get("cash_flow", {})
    kpis_ = extracted.get("operational_kpis", {})
    periods = sorted(extracted.get("periods", []))
    if not periods:
        return sugg

    def get_series(section, field):
        d = section.get(field, {})
        return {p: v for p, v in d.items() if v and abs(float(v)) > 0.01}

    rev_s = get_series(is_, "net_revenue") or get_series(is_, "gross_revenue")

    if rev_s:
        yrs = sorted(rev_s.keys())
        last_yr = yrs[-1]
        sugg["base_revenue"] = round(float(rev_s[last_yr]))
        sugg["start_year"]   = int(last_yr) + 1

        if len(yrs) >= 2:
            growths = []
            for i in range(1, len(yrs)):
                prev, curr = float(rev_s[yrs[i-1]]), float(rev_s[yrs[i]])
                if prev > 0:
                    growths.append((curr - prev) / prev)
            if growths:
                sugg["revenue_growth"] = round(sum(growths)/len(growths)*100, 1)

    # Margins
    for field, key in [("ebitda_adj","ebitda_margin"),("gross_profit","gross_margin"),("ebit","ebit_margin"),("net_income","net_margin")]:
        s = get_series(is_, field)
        if s and rev_s:
            common = sorted(set(s.keys()) & set(rev_s.keys()))
            if common:
                margins = [float(s[y])/max(abs(float(rev_s[y])),1) for y in common if float(rev_s[y])>0]
                if margins:
                    sugg[key] = round(sum(margins)/len(margins)*100, 1)

    # NWC days
    if rev_s:
        ar_s  = get_series(bs_, "trade_receivables")
        inv_s = get_series(bs_, "inventories")
        ap_s  = get_series(bs_, "trade_payables")
        for bs_field, sugg_key in [(ar_s,"dso"),(inv_s,"dio"),(ap_s,"dpo")]:
            if bs_field:
                common = sorted(set(bs_field.keys()) & set(rev_s.keys()))
                if common:
                    days = [abs(float(bs_field[y]))/max(abs(float(rev_s[y])),1)*365 for y in common]
                    sugg[sugg_key] = round(sum(days)/len(days))

    # Opening cash
    cash_s = get_series(bs_, "cash")
    if cash_s:
        last = sorted(cash_s.keys())[-1]
        sugg["opening_cash"] = round(abs(float(cash_s[last])))

    # Total debt
    debt_s = get_series(bs_, "financial_debt")
    if debt_s:
        last = sorted(debt_s.keys())[-1]
        sugg["total_debt"] = round(abs(float(debt_s[last])))

    # Tax rate
    tax_s  = get_series(is_, "tax")
    pbt_s  = get_series(is_, "pbt")
    if tax_s and pbt_s:
        common = sorted(set(tax_s.keys()) & set(pbt_s.keys()))
        rates  = [abs(float(tax_s[y]))/max(abs(float(pbt_s[y])),1) for y in common if float(pbt_s.get(y,0))>0]
        if rates:
            sugg["tax_rate"] = round(sum(rates)/len(rates)*100, 1)

    # Industrial KPIs
    vol_s = get_series(kpis_, "sales_volume_mt")
    price_s = get_series(kpis_, "avg_price_per_mt")
    if vol_s:
        last = sorted(vol_s.keys())[-1]
        sugg["capacity_mt"] = round(float(vol_s[last]) * 1.2)  # assume 120% of last actuals = capacity
    if price_s:
        last = sorted(price_s.keys())[-1]
        sugg["price_per_mt"] = round(float(price_s[last]))

    return sugg


# ══════════════════════════════════════════════════════════════════════════════
# STEP 6 — BUILD HISTORICAL TABLE FOR DISPLAY
# ══════════════════════════════════════════════════════════════════════════════

def build_historical_table(extracted: Dict) -> List[Dict]:
    is_    = extracted.get("income_statement", {})
    bs_    = extracted.get("balance_sheet", {})
    cf_    = extracted.get("cash_flow", {})
    periods = sorted(extracted.get("periods", []))

    table = []
    for p in periods:
        row = {"year": p}
        for field, section in [
            ("net_revenue",    is_), ("gross_profit", is_),
            ("ebitda_adj",     is_), ("ebit",         is_),
            ("net_income",     is_), ("cash",         bs_),
            ("financial_debt", bs_), ("capex",        cf_),
        ]:
            val = section.get(field, {}).get(p)
            if val is not None and abs(float(val)) > 0.01:
                row[field] = round(float(val))
        table.append(row)
    return table


# ══════════════════════════════════════════════════════════════════════════════
# STEP 7 — QUALITY SCORING
# ══════════════════════════════════════════════════════════════════════════════

def quality_score(extracted: Dict) -> Dict:
    is_ = extracted.get("income_statement", {})
    bs_ = extracted.get("balance_sheet", {})

    key_fields = ["net_revenue","ebitda_adj","net_income","cash","financial_debt"]
    found = sum(
        1 for f in key_fields
        if any(abs(float(v)) > 0.01 for v in (is_.get(f) or bs_.get(f) or {}).values() if v)
    )
    pct = round(found / len(key_fields) * 100)
    label = "Excellent" if pct >= 80 else "Good" if pct >= 60 else "Partial" if pct >= 40 else "Low"

    all_fields = [f for s in [is_, bs_, extracted.get("cash_flow",{})] for f, d in s.items() if d and any(abs(float(v))>0.01 for v in d.values() if v)]

    return {
        "score": pct, "label": label,
        "found_fields": all_fields,
        "periods_detected": len(extracted.get("periods", [])),
        "currency": extracted.get("detected_currency", "—"),
        "language": extracted.get("detected_language", "—"),
        "notes": extracted.get("notes", ""),
    }


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT  (async — called from FastAPI)
# ══════════════════════════════════════════════════════════════════════════════

async def extract_financials_llm(file_bytes: bytes, filename: str) -> Dict:
    """
    Full LLM-powered extraction pipeline.
    Returns structured dict ready for frontend rendering.
    """
    # 1. Raw extraction
    raw_text = extract_raw(file_bytes, filename)
    if not raw_text.strip():
        return {
            "error": "Could not extract any text from this file. Please check the file is not empty or image-only.",
            "quality": {"score": 0, "label": "Error"},
        }

    # 2. Claude call
    try:
        extracted = await call_claude_extract(raw_text)
    except json.JSONDecodeError as e:
        return {"error": f"Claude returned invalid JSON: {e}", "quality": {"score": 0, "label": "Error"}}
    except Exception as e:
        return {"error": f"API call failed: {e}", "quality": {"score": 0, "label": "Error"}}

    # 3. Normalise
    extracted = normalise(extracted)

    # 4. Build outputs
    suggestions = derive_assumptions(extracted)
    hist_table  = build_historical_table(extracted)
    quality     = quality_score(extracted)

    return {
        "extracted":       extracted,
        "suggestions":     suggestions,
        "historical_table":hist_table,
        "quality":         quality,
        "hist_years":      extracted.get("periods", []),
        "filename":        filename,
    }
