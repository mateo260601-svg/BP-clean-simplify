import json
import uuid
import importlib
import traceback
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional

from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel


app = FastAPI(title="MG Advisory Institutional BP - Safe V12.1", version="12.1")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

BASE_DIR = Path(__file__).parent.resolve()
DATA_DIR = BASE_DIR / "data"
PROJECTS_DIR = DATA_DIR / "projects"
OUTPUTS_DIR = BASE_DIR / "outputs"
UPLOADS_DIR = BASE_DIR / "uploads"
for d in [DATA_DIR, PROJECTS_DIR, OUTPUTS_DIR, UPLOADS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

VALID_LICENSES = {
    "JRC-MATEO-2025": {"user": "Mateo Girard", "expires": "2027-12-31"},
    "JRC-JAUFRE-2025": {"user": "Jaufre Rouanet", "expires": "2027-12-31"},
    "JRC-OZAN-2025": {"user": "Ozan OK", "expires": "2027-12-31"},
    "JRC-DEV-LOCAL": {"user": "Developer", "expires": "2099-12-31"},
    "MG-ACCESS-2025": {"user": "MG Advisory", "expires": "2027-12-31"},
    "BP-ELITE-2025": {"user": "MG Advisory", "expires": "2027-12-31"},
}

IMPORT_STATUS: Dict[str, Dict[str, Any]] = {}


def safe_import(module_name: str):
    """
    Import helper that never crashes server startup.
    The previous server imported V12 modules directly at boot;
    one missing/broken module was enough to create Railway 502.
    """
    if module_name in IMPORT_STATUS and IMPORT_STATUS[module_name].get("module") is not None:
        return IMPORT_STATUS[module_name]["module"]
    try:
        module = importlib.import_module(module_name)
        IMPORT_STATUS[module_name] = {"ok": True, "error": None, "module": module}
        return module
    except Exception as e:
        IMPORT_STATUS[module_name] = {
            "ok": False,
            "error": f"{type(e).__name__}: {str(e)}",
            "traceback": traceback.format_exc()[-3000:],
            "module": None,
        }
        return None


def module_status(module_name: str):
    mod = safe_import(module_name)
    data = IMPORT_STATUS.get(module_name, {})
    return {
        "ok": bool(mod),
        "error": data.get("error"),
    }


class LoginRequest(BaseModel):
    license_key: str


class ProjectCreate(BaseModel):
    license_key: str
    name: str = "New Investment Project"
    company_name: str = "Target Company"
    sector: str = "Industrial / Manufacturing"
    currency: str = "EUR"
    deal_type: str = "M&A"


class IntakeRequest(BaseModel):
    license_key: str
    project_id: str
    intake: Dict[str, Any] = {}


class GenerateRequest(BaseModel):
    license_key: str
    project_id: Optional[str] = None
    intake: Optional[Dict[str, Any]] = None
    company_name: str = "Target Company"


def check_license(key: str):
    k = (key or "").strip().upper()
    if k not in VALID_LICENSES and not k.startswith("JRC-"):
        raise HTTPException(status_code=401, detail="Invalid license key")
    return VALID_LICENSES.get(k, {"user": "JRC User"})


def fallback_debt_stack():
    return [
        {
            "instrument": "Super Senior RCF",
            "debt_type": "Super Senior RCF",
            "currency": "EUR",
            "commitment": 30000,
            "opening_balance": 0,
            "rate": 0.045,
            "margin": 0.018,
            "cash_pay": True,
            "pik": False,
            "amortization": "Revolver",
            "tenor_years": 5,
            "grace_periods": 0,
            "ranking": 0,
            "secured": True,
            "sweep_eligible": False,
            "bullet": False,
            "min_cash": 10000,
            "commitment_fee": 0.010,
            "oid": 0.000,
            "exit_fee": 0.000,
        },
        {
            "instrument": "Senior Term Loan B",
            "debt_type": "Senior Term Loan B",
            "currency": "EUR",
            "commitment": 130000,
            "opening_balance": 130000,
            "rate": 0.065,
            "margin": 0.030,
            "cash_pay": True,
            "pik": False,
            "amortization": "Bullet",
            "tenor_years": 5,
            "grace_periods": 0,
            "ranking": 1,
            "secured": True,
            "sweep_eligible": True,
            "bullet": True,
            "min_cash": 10000,
            "commitment_fee": 0.000,
            "oid": 0.010,
            "exit_fee": 0.010,
        },
        {
            "instrument": "Mezzanine PIK",
            "debt_type": "Mezzanine PIK",
            "currency": "EUR",
            "commitment": 40000,
            "opening_balance": 40000,
            "rate": 0.120,
            "margin": 0.000,
            "cash_pay": False,
            "pik": True,
            "amortization": "PIK",
            "tenor_years": 5,
            "grace_periods": 0,
            "ranking": 3,
            "secured": False,
            "sweep_eligible": False,
            "bullet": True,
            "min_cash": 10000,
            "commitment_fee": 0.000,
            "oid": 0.020,
            "exit_fee": 0.020,
        },
    ]


def get_default_debt_stack():
    mod = safe_import("institutional_debt_engine_v12")
    if mod and hasattr(mod, "default_debt_stack_for_saas"):
        try:
            return mod.default_debt_stack_for_saas()
        except Exception:
            pass
    return fallback_debt_stack()


def get_debt_library():
    mod = safe_import("institutional_debt_engine_v12")
    if mod and hasattr(mod, "debt_library_payload"):
        try:
            return mod.debt_library_payload()
        except Exception as e:
            IMPORT_STATUS["institutional_debt_engine_v12_runtime"] = {"ok": False, "error": str(e)}
    debt_types = [
        "Super Senior RCF", "RCF", "Senior Term Loan A", "Senior Term Loan B", "Unitranche",
        "Second Lien", "Mezzanine Cash Pay", "Mezzanine PIK", "Mezzanine Toggle",
        "Shareholder Loan Cash Pay", "Shareholder Loan PIK", "Vendor Loan", "Seller Note",
        "Bridge Loan", "DIP Financing", "Asset Based Lending", "Factoring Recourse",
        "Factoring Non-Recourse", "Securitisation", "Finance Lease", "Operating Lease IFRS16",
        "High Yield Bond", "Convertible Note", "Preferred Equity", "Tax Debt Payment Plan",
    ]
    return {
        "debt_types": debt_types,
        "amortization_types": ["Bullet", "Linear", "Annuity", "Cash Sweep", "Revolver", "Borrowing Base", "PIK", "PIK Toggle"],
        "rate_basis": ["Fixed", "Floating", "EURIBOR", "SOFR", "Base + Margin"],
        "currencies": ["EUR", "USD", "GBP", "CHF"],
        "case_list": ["Base", "Downside", "Upside", "Bank Case", "IC Case"],
        "library": {},
    }


def default_intake(company_name="Target Company", currency="EUR", sector="Industrial / Manufacturing", deal_type="M&A"):
    return {
        "company_name": company_name,
        "currency": currency,
        "sector": sector,
        "deal_type": deal_type,
        "start_year": 2026,
        "forecast_years": 5,
        "revenue": 202000,
        "revenue_growth": 0.073,
        "gross_margin": 0.391,
        "ebitda_margin": 0.229,
        "tax_rate": 0.25,
        "cash": 18700,
        "debt": 201000,
        "dso": 55,
        "dio": 40,
        "dpo": 55,
        "maintenance_capex_pct_revenue": 0.055,
        "debt_stack": get_default_debt_stack(),
    }


def project_file(project_id):
    return PROJECTS_DIR / project_id / "project.json"


def save_project(project):
    project["updated_at"] = datetime.utcnow().isoformat()
    folder = PROJECTS_DIR / project["project_id"]
    folder.mkdir(parents=True, exist_ok=True)
    project_file(project["project_id"]).write_text(json.dumps(project, indent=2), encoding="utf-8")
    return project


def load_project(project_id, license_key=None):
    path = project_file(project_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Project not found")
    project = json.loads(path.read_text(encoding="utf-8"))
    if license_key and project.get("license_key") != license_key:
        raise HTTPException(status_code=403, detail="Access denied")
    return project


def normalize_historical(extractions):
    mod = safe_import("historical_accounts_engine")
    if mod and hasattr(mod, "normalize_historical_accounts"):
        return mod.normalize_historical_accounts(extractions)
    return {"periods": [], "evidence": [], "warnings": ["historical_accounts_engine unavailable"], "latest": {}, "assumptions": {}}


def merge_historical(intake, historical):
    mod = safe_import("historical_accounts_engine")
    if mod and hasattr(mod, "merge_historical_into_intake"):
        return mod.merge_historical_into_intake(intake, historical)
    return dict(intake or {})


def fallback_excel_model(intake: Dict[str, Any], historical: Dict[str, Any], output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    navy = "0D1B3E"
    input_fill = "EBF5FF"
    output_fill = "EAF2FF"
    line = "D0D5DD"
    side = Side(style="thin", color=line)

    def setup(ws):
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "F6"
        for c in range(1, 30):
            ws.column_dimensions[get_column_letter(c)].width = 16
        ws.column_dimensions["B"].width = 34

    def header(cell, val):
        cell.value = val
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=side, right=side, top=side, bottom=side)

    def inp(cell, val, fmt=None):
        cell.value = val
        cell.fill = PatternFill("solid", fgColor=input_fill)
        cell.font = Font(color="0000CD")
        cell.border = Border(left=side, right=side, top=side, bottom=side)
        if fmt:
            cell.number_format = fmt

    def formula(cell, val, fmt=None, out=False):
        if not str(val).startswith("="):
            val = "=" + str(val)
        cell.value = val
        cell.fill = PatternFill("solid", fgColor=output_fill if out else "FFFFFF")
        cell.border = Border(left=side, right=side, top=side, bottom=side)
        if fmt:
            cell.number_format = fmt

    money = '#,##0.0;[Red](#,##0.0);"-"'
    pct = '0.0%;[Red](0.0%);"-"'
    mult = '0.0x;[Red](0.0x);"-"'

    ws = wb.active
    ws.title = "Assumptions"
    setup(ws)
    ws["B2"] = "MG Advisory BP - Safe Fallback Model"
    ws["B2"].font = Font(size=16, bold=True, color=navy)
    assumptions = [
        ("Company", intake.get("company_name", "Target Company")),
        ("Currency", intake.get("currency", "EUR")),
        ("Start Year", int(float(intake.get("start_year", 2026) or 2026))),
        ("Forecast Years", int(float(intake.get("forecast_years", 5) or 5))),
        ("Revenue", float(intake.get("revenue", 202000) or 202000)),
        ("Revenue Growth", float(intake.get("revenue_growth", 0.073) or 0.073)),
        ("Gross Margin", float(intake.get("gross_margin", 0.391) or 0.391)),
        ("EBITDA Margin", float(intake.get("ebitda_margin", 0.229) or 0.229)),
        ("Opening Cash", float(intake.get("cash", 18700) or 18700)),
        ("Opening Debt", float(intake.get("debt", 201000) or 201000)),
        ("Cash Sweep %", 0.50),
        ("Interest Rate", 0.075),
    ]
    for r, (k, v) in enumerate(assumptions, start=6):
        ws.cell(r, 2).value = k
        ws.cell(r, 2).font = Font(bold=True, color=navy)
        fmt = pct if "%" in k or "Growth" in k or "Margin" in k or "Rate" in k else money if isinstance(v, (int, float)) and k not in ["Start Year", "Forecast Years"] else None
        inp(ws.cell(r, 6), v, fmt)

    ws2 = wb.create_sheet("Financial Statements")
    setup(ws2)
    years = int(float(intake.get("forecast_years", 5) or 5))
    start_year = int(float(intake.get("start_year", 2026) or 2026))
    for i in range(years):
        col = 6 + i
        header(ws2.cell(4, col), f"FY {start_year+i}")
        header(ws2.cell(5, col), "Forecast")
    rows = {
        "Revenue": 7, "Gross Profit": 8, "EBITDA": 9, "EBITDA Margin": 10,
        "Cash Interest": 11, "FCF": 12, "Closing Cash": 13, "Closing Debt": 14,
        "Net Debt": 15, "Net Debt / EBITDA": 16,
    }
    for name, r in rows.items():
        ws2.cell(r, 2).value = name
        ws2.cell(r, 2).font = Font(bold=True, color=navy)
    for i in range(years):
        c = 6 + i
        col = get_column_letter(c)
        prev = get_column_letter(c - 1)
        if i == 0:
            formula(ws2.cell(7, c), "='Assumptions'!$F$10", money, True)
            formula(ws2.cell(13, c), "='Assumptions'!$F$14", money, True)
            formula(ws2.cell(14, c), "='Assumptions'!$F$15", money, True)
        else:
            formula(ws2.cell(7, c), f"={prev}7*(1+'Assumptions'!$F$11)", money, True)
            formula(ws2.cell(13, c), f"={prev}13+{col}12", money, True)
            formula(ws2.cell(14, c), f"=MAX(0,{prev}14-{col}12*'Assumptions'!$F$16)", money, True)
        formula(ws2.cell(8, c), f"={col}7*'Assumptions'!$F$12", money)
        formula(ws2.cell(9, c), f"={col}7*'Assumptions'!$F$13", money, True)
        formula(ws2.cell(10, c), f"=IFERROR({col}9/{col}7,0)", pct)
        formula(ws2.cell(11, c), f"={col}14*'Assumptions'!$F$17", money)
        formula(ws2.cell(12, c), f"={col}9-{col}11", money, True)
        formula(ws2.cell(15, c), f"={col}14-{col}13", money, True)
        formula(ws2.cell(16, c), f"=IFERROR({col}15/{col}9,0)", mult, True)

    ws3 = wb.create_sheet("Debt Config")
    setup(ws3)
    ws3["B2"] = "Debt Config - Safe Fallback"
    ws3["B2"].font = Font(size=16, bold=True, color=navy)
    headers = ["Instrument", "Type", "Opening Balance", "Rate", "Amortization", "PIK?"]
    for c, h in enumerate(headers, start=2):
        header(ws3.cell(5, c), h)
    for r, tranche in enumerate(get_default_debt_stack(), start=6):
        vals = [tranche.get("instrument"), tranche.get("debt_type"), tranche.get("opening_balance"), tranche.get("rate"), tranche.get("amortization"), tranche.get("pik")]
        for c, v in enumerate(vals, start=2):
            inp(ws3.cell(r, c), v, pct if c == 5 else money if c == 4 else None)

    ws4 = wb.create_sheet("Checks")
    setup(ws4)
    header(ws4["B5"], "Check")
    header(ws4["C5"], "Status")
    ws4["B6"] = "Model generated"
    ws4["C6"] = "OK"
    ws4["B7"] = "Engine status"
    ws4["C7"] = "Safe fallback active if V12 module failed"

    wb.save(output_path)


def generate_excel(intake, historical, output_path):
    mod = safe_import("institutional_formula_model_builder_v12")
    if mod and hasattr(mod, "build_v12_formula_model"):
        try:
            return mod.build_v12_formula_model(intake, historical, str(output_path))
        except Exception as e:
            IMPORT_STATUS["institutional_formula_model_builder_v12_runtime"] = {
                "ok": False,
                "error": f"{type(e).__name__}: {str(e)}",
                "traceback": traceback.format_exc()[-3000:],
            }
    fallback_excel_model(intake, historical, str(output_path))
    return str(output_path)


@app.get("/", response_class=HTMLResponse)
def index():
    path = BASE_DIR / "index.html"
    if path.exists():
        return path.read_text(encoding="utf-8")
    return "<h1>MG Advisory Safe V12.1 is online</h1>"


@app.get("/app", response_class=HTMLResponse)
def app_page():
    return index()


@app.get("/health")
def health():
    debt_payload = get_debt_library()
    # Probe core modules, but never crash.
    probes = {
        "historical_accounts_engine": module_status("historical_accounts_engine"),
        "institutional_debt_engine_v12": module_status("institutional_debt_engine_v12"),
        "institutional_formula_model_builder_v12": module_status("institutional_formula_model_builder_v12"),
        "extraction_core": module_status("extraction_core"),
        "claude_client": module_status("claude_client"),
        "ai_financial_extractor": module_status("ai_financial_extractor"),
    }
    claude = {"configured": False}
    claude_mod = safe_import("claude_client")
    if claude_mod and hasattr(claude_mod, "get_claude_status"):
        try:
            claude = claude_mod.get_claude_status()
        except Exception as e:
            claude = {"configured": False, "error": str(e)}
    return {
        "ok": True,
        "version": "12.1-safe",
        "railway_should_not_502": True,
        "institutional_debt_engine_v12": probes["institutional_debt_engine_v12"]["ok"],
        "debt_types": len(debt_payload.get("debt_types", [])),
        "max_tranches": 60,
        "modules": probes,
        "claude": claude,
    }


@app.get("/api/debt/library")
def debt_library():
    return {"ok": True, **get_debt_library()}


@app.post("/api/auth/login")
def login(payload: LoginRequest):
    user = check_license(payload.license_key)
    return {"ok": True, "license": payload.license_key, "user": user["user"]}


@app.get("/api/wizards")
def wizards():
    return {"ok": True, "default_intake": default_intake(), "debt": get_debt_library()}


@app.post("/api/projects")
def create_project(payload: ProjectCreate):
    check_license(payload.license_key)
    project = {
        "project_id": str(uuid.uuid4()),
        "license_key": payload.license_key,
        "name": payload.name,
        "company_name": payload.company_name,
        "sector": payload.sector,
        "currency": payload.currency,
        "deal_type": payload.deal_type,
        "created_at": datetime.utcnow().isoformat(),
        "updated_at": datetime.utcnow().isoformat(),
        "documents": [],
        "extractions": [],
        "historical": {"periods": [], "evidence": [], "warnings": []},
        "intake": default_intake(payload.company_name, payload.currency, payload.sector, payload.deal_type),
        "outputs": {},
    }
    save_project(project)
    return {"ok": True, "project": project}


@app.get("/api/projects")
def list_projects(license_key: str):
    check_license(license_key)
    projects = []
    for fp in PROJECTS_DIR.glob("*/project.json"):
        try:
            p = json.loads(fp.read_text(encoding="utf-8"))
            if p.get("license_key") == license_key:
                projects.append(p)
        except Exception:
            pass
    return {"ok": True, "projects": sorted(projects, key=lambda x: x.get("updated_at", ""), reverse=True)}


@app.get("/api/projects/{project_id}")
def get_project(project_id: str, license_key: str):
    check_license(license_key)
    return {"ok": True, "project": load_project(project_id, license_key)}


@app.post("/api/projects/{project_id}/upload")
async def upload_project_file(
    project_id: str,
    license_key: str = Form(...),
    category: str = Form("financials"),
    file: UploadFile = File(...),
    use_ai: bool = Form(True),
):
    check_license(license_key)
    project = load_project(project_id, license_key)

    file_id = str(uuid.uuid4())
    safe = file.filename.replace("/", "_").replace("\\", "_")
    folder = PROJECTS_DIR / project_id / "documents"
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{file_id}_{safe}"
    path.write_bytes(await file.read())

    extraction = {"file_name": safe, "normalized": {"periods": []}, "metrics": [], "audit": {"warnings": []}}
    extraction_mod = safe_import("extraction_core")
    if extraction_mod and hasattr(extraction_mod, "extract_document"):
        try:
            extraction = extraction_mod.extract_document(str(path))
        except Exception as e:
            extraction["audit"] = {"warnings": [f"Extraction failed: {type(e).__name__}: {str(e)}"]}
    else:
        extraction["audit"] = {"warnings": ["extraction_core.py is missing; file stored but not parsed"]}

    if use_ai:
        ai_mod = safe_import("ai_financial_extractor")
        if ai_mod and hasattr(ai_mod, "enhance_extraction_with_claude"):
            try:
                extraction = ai_mod.enhance_extraction_with_claude(extraction)
            except Exception as e:
                extraction.setdefault("ai", {"used": False, "status": "error", "reason": str(e)})

    project["documents"].append({
        "file_id": file_id,
        "filename": safe,
        "path": str(path),
        "category": category,
        "uploaded_at": datetime.utcnow().isoformat(),
    })
    project["extractions"].append(extraction)
    historical = normalize_historical(project["extractions"])
    project["historical"] = historical
    project["intake"] = merge_historical(project.get("intake") or {}, historical)
    project["intake"].setdefault("debt_stack", get_default_debt_stack())
    save_project(project)

    return {
        "ok": True,
        "file_id": file_id,
        "project": project,
        "historical": historical,
        "suggested_intake": project["intake"],
        "warnings": historical.get("warnings", []),
    }


@app.post("/api/import")
async def import_file(license_key: str = Form(...), file: UploadFile = File(...), use_ai: bool = Form(True)):
    check_license(license_key)
    file_id = str(uuid.uuid4())
    safe = file.filename.replace("/", "_").replace("\\", "_")
    path = UPLOADS_DIR / f"{file_id}_{safe}"
    path.write_bytes(await file.read())

    extraction = {"file_name": safe, "normalized": {"periods": []}, "metrics": [], "audit": {"warnings": []}}
    extraction_mod = safe_import("extraction_core")
    if extraction_mod and hasattr(extraction_mod, "extract_document"):
        try:
            extraction = extraction_mod.extract_document(str(path))
        except Exception as e:
            extraction["audit"] = {"warnings": [f"Extraction failed: {type(e).__name__}: {str(e)}"]}
    else:
        extraction["audit"] = {"warnings": ["extraction_core.py is missing; file stored but not parsed"]}

    if use_ai:
        ai_mod = safe_import("ai_financial_extractor")
        if ai_mod and hasattr(ai_mod, "enhance_extraction_with_claude"):
            try:
                extraction = ai_mod.enhance_extraction_with_claude(extraction)
            except Exception as e:
                extraction.setdefault("ai", {"used": False, "status": "error", "reason": str(e)})

    historical = normalize_historical([extraction])
    intake = merge_historical(default_intake(safe.rsplit(".", 1)[0]), historical)
    return {
        "ok": True,
        "file_id": file_id,
        "filename": safe,
        "historical": historical,
        "financials": historical,
        "suggested_intake": intake,
        "warnings": historical.get("warnings", []),
    }


@app.post("/api/projects/{project_id}/intake")
def save_intake(project_id: str, payload: IntakeRequest):
    check_license(payload.license_key)
    project = load_project(project_id, payload.license_key)
    project["intake"].update(payload.intake or {})
    project["intake"].setdefault("debt_stack", get_default_debt_stack())
    save_project(project)
    return {"ok": True, "project": project, "intake": project["intake"]}


@app.post("/api/projects/{project_id}/generate-bp")
def generate_project_bp(project_id: str, payload: GenerateRequest):
    check_license(payload.license_key)
    project = load_project(project_id, payload.license_key)

    intake = dict(project.get("intake") or {})
    if payload.intake:
        intake.update(payload.intake)
    intake.setdefault("debt_stack", get_default_debt_stack())

    historical = project.get("historical") or {"periods": [], "evidence": [], "warnings": []}
    file_id = str(uuid.uuid4())
    safe = project.get("company_name", "Company").replace(" ", "_").replace("/", "_")
    output = OUTPUTS_DIR / f"{file_id}_{safe}_Institutional_BP_Debt_Engine_V12.xlsx"

    generate_excel(intake, historical, output)

    project["intake"] = intake
    project["outputs"]["bp_model"] = str(output)
    save_project(project)
    return {"ok": True, "file_id": file_id, "filename": output.name, "project": project}


@app.post("/api/generate")
def legacy_generate(payload: GenerateRequest):
    check_license(payload.license_key)
    intake = default_intake(payload.company_name)
    if payload.intake:
        intake.update(payload.intake)
    historical = {"periods": [], "evidence": [], "warnings": ["No historical accounts uploaded in legacy mode"]}
    file_id = str(uuid.uuid4())
    safe = intake.get("company_name", "Company").replace(" ", "_").replace("/", "_")
    output = OUTPUTS_DIR / f"{file_id}_{safe}_Institutional_BP_Debt_Engine_V12.xlsx"
    generate_excel(intake, historical, output)
    return {"ok": True, "file_id": file_id, "filename": output.name}


@app.get("/api/projects/{project_id}/download/{output_type}")
def download_project(project_id: str, output_type: str, license_key: str):
    check_license(license_key)
    project = load_project(project_id, license_key)
    path = project.get("outputs", {}).get(output_type)
    if not path or not Path(path).exists():
        raise HTTPException(status_code=404, detail="Output not found")
    return FileResponse(path, filename=Path(path).name)


@app.get("/api/download/{file_id}")
def download_legacy(file_id: str):
    for p in OUTPUTS_DIR.glob(f"{file_id}_*"):
        return FileResponse(str(p), filename=p.name)
    raise HTTPException(status_code=404, detail="File not found")
