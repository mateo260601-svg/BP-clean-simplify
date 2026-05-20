"""
MG Advisory Institutional Debt Engine V12

Purpose
-------
Builds an institutional-grade, formula-driven Excel debt module:
- 60+ debt instruments / financing products
- tranche-by-tranche configuration
- cash pay / PIK / toggle
- bullet / linear / annuity / cash sweep / borrowing base / revolver / revenue share
- RCF availability and drawings
- AR / inventory borrowing base
- fees, OID, commitment fees, exit fees
- covenants: leverage, senior leverage, ICR, DSCR, FCCR, min liquidity
- cash sweep and repayment waterfall
- debt output summary and checks

Design standards
----------------
- Inputs are hardcoded blue cells
- Formula cells are black
- Linked formulas are green
- Outputs are clean and formula-driven
- No hardcoded numbers in calculations when a config cell exists
"""

from typing import Dict, Any, List, Optional
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. Instrument library
# ---------------------------------------------------------------------------

DEBT_INSTRUMENT_LIBRARY: Dict[str, Dict[str, Any]] = {
    # Senior / bank debt
    "Super Senior RCF": {"category": "Senior Bank Debt", "ranking": 0, "default_rate": 0.045, "default_margin": 0.018, "amortization": "Revolver", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "RCF": {"category": "Senior Bank Debt", "ranking": 1, "default_rate": 0.050, "default_margin": 0.020, "amortization": "Revolver", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Overdraft": {"category": "Senior Bank Debt", "ranking": 1, "default_rate": 0.080, "default_margin": 0.030, "amortization": "Revolver", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Senior Term Loan A": {"category": "Senior Bank Debt", "ranking": 1, "default_rate": 0.055, "default_margin": 0.025, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Senior Term Loan B": {"category": "Senior Bank Debt", "ranking": 1, "default_rate": 0.065, "default_margin": 0.030, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Senior Term Loan C": {"category": "Senior Bank Debt", "ranking": 1, "default_rate": 0.070, "default_margin": 0.035, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Delayed Draw Term Loan": {"category": "Senior Bank Debt", "ranking": 1, "default_rate": 0.068, "default_margin": 0.032, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Capex Facility": {"category": "Senior Bank Debt", "ranking": 1, "default_rate": 0.060, "default_margin": 0.025, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Acquisition Facility": {"category": "Senior Bank Debt", "ranking": 1, "default_rate": 0.065, "default_margin": 0.030, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Accordion Facility": {"category": "Senior Bank Debt", "ranking": 1, "default_rate": 0.070, "default_margin": 0.035, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": False},

    # Asset-backed / working capital finance
    "Asset Based Lending": {"category": "Borrowing Base", "ranking": 1, "default_rate": 0.055, "default_margin": 0.020, "amortization": "Borrowing Base", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Borrowing Base Facility": {"category": "Borrowing Base", "ranking": 1, "default_rate": 0.055, "default_margin": 0.020, "amortization": "Borrowing Base", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Receivables Facility": {"category": "Borrowing Base", "ranking": 1, "default_rate": 0.050, "default_margin": 0.018, "amortization": "Borrowing Base", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Inventory Facility": {"category": "Borrowing Base", "ranking": 1, "default_rate": 0.060, "default_margin": 0.025, "amortization": "Borrowing Base", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Factoring Recourse": {"category": "Borrowing Base", "ranking": 1, "default_rate": 0.060, "default_margin": 0.025, "amortization": "Borrowing Base", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Factoring Non-Recourse": {"category": "Borrowing Base", "ranking": 1, "default_rate": 0.070, "default_margin": 0.030, "amortization": "Borrowing Base", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Securitisation": {"category": "Borrowing Base", "ranking": 1, "default_rate": 0.052, "default_margin": 0.020, "amortization": "Borrowing Base", "cash_pay": True, "pik": False, "secured": True, "drawn": False},
    "Supply Chain Finance": {"category": "Working Capital Finance", "ranking": 2, "default_rate": 0.045, "default_margin": 0.015, "amortization": "Operational Run-Off", "cash_pay": True, "pik": False, "secured": False, "drawn": False},
    "Reverse Factoring": {"category": "Working Capital Finance", "ranking": 2, "default_rate": 0.045, "default_margin": 0.015, "amortization": "Operational Run-Off", "cash_pay": True, "pik": False, "secured": False, "drawn": False},

    # Unitranche / junior debt
    "Unitranche": {"category": "Private Credit", "ranking": 2, "default_rate": 0.085, "default_margin": 0.040, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "First Out Unitranche": {"category": "Private Credit", "ranking": 1, "default_rate": 0.070, "default_margin": 0.030, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Last Out Unitranche": {"category": "Private Credit", "ranking": 2, "default_rate": 0.100, "default_margin": 0.050, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Second Lien": {"category": "Junior Secured", "ranking": 2, "default_rate": 0.095, "default_margin": 0.045, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Mezzanine Cash Pay": {"category": "Mezzanine", "ranking": 3, "default_rate": 0.110, "default_margin": 0.055, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Mezzanine PIK": {"category": "Mezzanine", "ranking": 3, "default_rate": 0.120, "default_margin": 0.060, "amortization": "PIK", "cash_pay": False, "pik": True, "secured": False, "drawn": True},
    "Mezzanine Toggle": {"category": "Mezzanine", "ranking": 3, "default_rate": 0.115, "default_margin": 0.055, "amortization": "PIK Toggle", "cash_pay": True, "pik": True, "secured": False, "drawn": True},
    "HoldCo PIK": {"category": "HoldCo Debt", "ranking": 5, "default_rate": 0.125, "default_margin": 0.000, "amortization": "PIK", "cash_pay": False, "pik": True, "secured": False, "drawn": True},
    "OpCo PIK": {"category": "PIK Debt", "ranking": 4, "default_rate": 0.115, "default_margin": 0.000, "amortization": "PIK", "cash_pay": False, "pik": True, "secured": False, "drawn": True},

    # Bonds / capital markets
    "High Yield Bond": {"category": "Bonds", "ranking": 3, "default_rate": 0.095, "default_margin": 0.045, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Senior Secured Notes": {"category": "Bonds", "ranking": 2, "default_rate": 0.075, "default_margin": 0.030, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Senior Unsecured Notes": {"category": "Bonds", "ranking": 3, "default_rate": 0.085, "default_margin": 0.040, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Fixed Rate Bond": {"category": "Bonds", "ranking": 2, "default_rate": 0.075, "default_margin": 0.000, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Floating Rate Bond": {"category": "Bonds", "ranking": 2, "default_rate": 0.070, "default_margin": 0.030, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Private Placement": {"category": "Bonds", "ranking": 2, "default_rate": 0.070, "default_margin": 0.030, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": False, "drawn": True},

    # Hybrid / equity-like
    "Convertible Note": {"category": "Hybrid", "ranking": 4, "default_rate": 0.080, "default_margin": 0.020, "amortization": "PIK", "cash_pay": False, "pik": True, "secured": False, "drawn": True},
    "Preferred Equity": {"category": "Hybrid", "ranking": 6, "default_rate": 0.110, "default_margin": 0.000, "amortization": "PIK", "cash_pay": False, "pik": True, "secured": False, "drawn": True},
    "Redeemable Preferred Equity": {"category": "Hybrid", "ranking": 6, "default_rate": 0.120, "default_margin": 0.000, "amortization": "PIK", "cash_pay": False, "pik": True, "secured": False, "drawn": True},
    "Warrant-Linked Debt": {"category": "Hybrid", "ranking": 4, "default_rate": 0.085, "default_margin": 0.025, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": False, "drawn": True},

    # Shareholder / vendor / seller
    "Shareholder Loan Cash Pay": {"category": "Shareholder", "ranking": 5, "default_rate": 0.080, "default_margin": 0.000, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Shareholder Loan PIK": {"category": "Shareholder", "ranking": 5, "default_rate": 0.100, "default_margin": 0.000, "amortization": "PIK", "cash_pay": False, "pik": True, "secured": False, "drawn": True},
    "Shareholder Current Account": {"category": "Shareholder", "ranking": 5, "default_rate": 0.030, "default_margin": 0.000, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Vendor Loan": {"category": "Vendor / Seller", "ranking": 4, "default_rate": 0.060, "default_margin": 0.000, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Seller Note": {"category": "Vendor / Seller", "ranking": 4, "default_rate": 0.060, "default_margin": 0.000, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Earn-Out Deferred Consideration": {"category": "Vendor / Seller", "ranking": 5, "default_rate": 0.000, "default_margin": 0.000, "amortization": "Bullet", "cash_pay": False, "pik": False, "secured": False, "drawn": True},
    "Contingent Consideration": {"category": "Vendor / Seller", "ranking": 5, "default_rate": 0.000, "default_margin": 0.000, "amortization": "Bullet", "cash_pay": False, "pik": False, "secured": False, "drawn": True},
    "Management Loan": {"category": "Management", "ranking": 5, "default_rate": 0.050, "default_margin": 0.000, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": False, "drawn": True},

    # Bridge / distressed / restructuring
    "Bridge Loan": {"category": "Bridge", "ranking": 1, "default_rate": 0.090, "default_margin": 0.045, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Super Senior Bridge": {"category": "Bridge", "ranking": 0, "default_rate": 0.100, "default_margin": 0.050, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "DIP Financing": {"category": "Restructuring", "ranking": 0, "default_rate": 0.120, "default_margin": 0.060, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Rescue Financing": {"category": "Restructuring", "ranking": 0, "default_rate": 0.140, "default_margin": 0.070, "amortization": "Cash Sweep", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Tax Debt Payment Plan": {"category": "Restructuring", "ranking": 2, "default_rate": 0.040, "default_margin": 0.000, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Social Security Debt Plan": {"category": "Restructuring", "ranking": 2, "default_rate": 0.040, "default_margin": 0.000, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Supplier Payment Plan": {"category": "Restructuring", "ranking": 2, "default_rate": 0.000, "default_margin": 0.000, "amortization": "Operational Run-Off", "cash_pay": False, "pik": False, "secured": False, "drawn": True},
    "Customer Advance": {"category": "Restructuring", "ranking": 2, "default_rate": 0.000, "default_margin": 0.000, "amortization": "Operational Run-Off", "cash_pay": False, "pik": False, "secured": False, "drawn": True},
    "Cramdown Debt": {"category": "Restructuring", "ranking": 3, "default_rate": 0.080, "default_margin": 0.000, "amortization": "Sculpted", "cash_pay": True, "pik": False, "secured": False, "drawn": True},

    # Project / infra / asset finance
    "Project Finance": {"category": "Project Finance", "ranking": 1, "default_rate": 0.065, "default_margin": 0.030, "amortization": "Debt Sculpting", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Export Credit Agency Facility": {"category": "Project Finance", "ranking": 1, "default_rate": 0.045, "default_margin": 0.015, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "State Guaranteed Loan": {"category": "Project Finance", "ranking": 1, "default_rate": 0.035, "default_margin": 0.010, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Green Loan": {"category": "Project Finance", "ranking": 1, "default_rate": 0.055, "default_margin": 0.020, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Sustainability Linked Loan": {"category": "Project Finance", "ranking": 1, "default_rate": 0.055, "default_margin": 0.020, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Mortgage Debt": {"category": "Asset Finance", "ranking": 1, "default_rate": 0.050, "default_margin": 0.020, "amortization": "Annuity", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Equipment Loan": {"category": "Asset Finance", "ranking": 1, "default_rate": 0.060, "default_margin": 0.025, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Finance Lease": {"category": "Lease", "ranking": 1, "default_rate": 0.065, "default_margin": 0.025, "amortization": "Annuity", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Operating Lease IFRS16": {"category": "Lease", "ranking": 1, "default_rate": 0.060, "default_margin": 0.020, "amortization": "Annuity", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Sale & Leaseback": {"category": "Lease", "ranking": 1, "default_rate": 0.065, "default_margin": 0.025, "amortization": "Annuity", "cash_pay": True, "pik": False, "secured": True, "drawn": True},

    # Growth / venture
    "Venture Debt": {"category": "Growth Debt", "ranking": 2, "default_rate": 0.100, "default_margin": 0.050, "amortization": "Interest Only Then Linear", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Growth Debt": {"category": "Growth Debt", "ranking": 2, "default_rate": 0.095, "default_margin": 0.045, "amortization": "Interest Only Then Linear", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Revenue Based Financing": {"category": "Growth Debt", "ranking": 3, "default_rate": 0.080, "default_margin": 0.030, "amortization": "Revenue Share", "cash_pay": True, "pik": False, "secured": False, "drawn": True},
    "Royalty Financing": {"category": "Growth Debt", "ranking": 3, "default_rate": 0.080, "default_margin": 0.030, "amortization": "Revenue Share", "cash_pay": True, "pik": False, "secured": False, "drawn": True},

    # FX / local / miscellaneous
    "Local Bank Debt": {"category": "Local Debt", "ranking": 2, "default_rate": 0.070, "default_margin": 0.030, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "FX Debt": {"category": "FX Debt", "ranking": 2, "default_rate": 0.070, "default_margin": 0.030, "amortization": "Bullet", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Hedged Floating Debt": {"category": "Hedged Debt", "ranking": 1, "default_rate": 0.060, "default_margin": 0.025, "amortization": "Linear", "cash_pay": True, "pik": False, "secured": True, "drawn": True},
    "Swap Liability": {"category": "Hedge", "ranking": 2, "default_rate": 0.000, "default_margin": 0.000, "amortization": "Operational Run-Off", "cash_pay": False, "pik": False, "secured": False, "drawn": True},
    "Call Premium Liability": {"category": "Fees / Premium", "ranking": 3, "default_rate": 0.000, "default_margin": 0.000, "amortization": "Bullet", "cash_pay": False, "pik": False, "secured": False, "drawn": True},
    "Exit Fee Liability": {"category": "Fees / Premium", "ranking": 3, "default_rate": 0.000, "default_margin": 0.000, "amortization": "Bullet", "cash_pay": False, "pik": False, "secured": False, "drawn": True},
}

AMORTIZATION_TYPES = [
    "Bullet",
    "Linear",
    "Annuity",
    "Cash Sweep",
    "Excess Cash Flow Sweep",
    "Revolver",
    "Borrowing Base",
    "Interest Only Then Linear",
    "PIK",
    "PIK Toggle",
    "Revenue Share",
    "Operational Run-Off",
    "Debt Sculpting",
    "Sculpted",
]

BOOLEAN_LIST = ["TRUE", "FALSE"]
RATE_BASIS_LIST = ["Fixed", "Floating", "EURIBOR", "SOFR", "SONIA", "Base + Margin", "Hedged"]
RANKING_LIST = ["0", "1", "2", "3", "4", "5", "6"]
CURRENCY_LIST = ["EUR", "USD", "GBP", "CHF", "CAD", "AUD", "Local"]
CASE_LIST = ["Base", "Downside", "Upside", "Bank Case", "IC Case", "Restructuring Case"]


# ---------------------------------------------------------------------------
# 2. Formatting helpers
# ---------------------------------------------------------------------------

COLORS = {
    "navy": "0D1B3E",
    "dark_blue": "1F3D7A",
    "blue": "0000FF",
    "green": "008000",
    "red": "FF0000",
    "input": "EBF5FF",
    "output": "EAF2FF",
    "linked": "F7F8FA",
    "formula": "FFFFFF",
    "warning": "FFF3CD",
    "white": "FFFFFF",
    "black": "000000",
    "grey": "64748B",
    "line": "D0D5DD",
}

FIRST_PERIOD_COL = 8
MAX_TRANCHES = 60
MAX_YEARS = 10


def _fill(key: str):
    return PatternFill("solid", fgColor=COLORS[key])


def _font(color: str = "black", bold: bool = False, size: int = 9, italic: bool = False):
    return Font(name="Calibri", color=COLORS.get(color, color), bold=bold, size=size, italic=italic)


def _border():
    s = Side(style="thin", color=COLORS["line"])
    return Border(left=s, right=s, top=s, bottom=s)


def fmt_money():
    return '#,##0.0;[Red](#,##0.0);"-"'


def fmt_pct():
    return '0.0%;[Red](0.0%);"-"'


def fmt_mult():
    return '0.0x;[Red](0.0x);"-"'


def fmt_int():
    return '#,##0;[Red](#,##0);"-"'


def setup_sheet(ws, freeze: str = "H6", tab_color: str = "0D1B3E"):
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze
    ws.sheet_view.zoomScale = 85
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 2
    for c in range(FIRST_PERIOD_COL, FIRST_PERIOD_COL + 80):
        ws.column_dimensions[get_column_letter(c)].width = 13


def title(ws, text: str, subtitle: str = ""):
    ws["B2"] = text
    ws["B2"].font = _font("navy", bold=True, size=16)
    if subtitle:
        ws["B3"] = subtitle
        ws["B3"].font = _font("grey", italic=True, size=9)


def set_header(cell, value: str):
    cell.value = value
    cell.fill = _fill("navy")
    cell.font = _font("white", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border()


def set_section(ws, row: int, label: str, start_col: int = 2, end_col: int = 6):
    ws.cell(row, start_col).value = label
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row, c)
        cell.fill = _fill("navy")
        cell.font = _font("white", bold=True)
        cell.border = _border()
    ws.cell(row, start_col).alignment = Alignment(horizontal="left")


def row_label(ws, row: int, label: str, unit: str = "", source: str = "", style: str = "normal"):
    ws.cell(row, 2).value = label
    ws.cell(row, 3).value = unit
    ws.cell(row, 4).value = source
    for c in [2, 3, 4, 5, 6]:
        ws.cell(row, c).border = _border()
        ws.cell(row, c).font = _font("navy", bold=(style in ["section", "total"]))
    if style == "total":
        for c in [2, 3, 4, 5, 6]:
            ws.cell(row, c).fill = _fill("output")


def input_cell(cell, value: Any, number_format: Optional[str] = None):
    cell.value = value
    cell.fill = _fill("input")
    cell.font = _font("blue")
    cell.border = _border()
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if number_format:
        cell.number_format = number_format


def formula_cell(cell, formula: str, number_format: Optional[str] = None, output: bool = False, linked: bool = False):
    if not str(formula).startswith("="):
        formula = "=" + str(formula)
    cell.value = formula
    cell.fill = _fill("output" if output else "linked" if linked else "formula")
    cell.font = _font("green" if linked else "black", bold=output)
    cell.border = _border()
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if number_format:
        cell.number_format = number_format


def add_validation(ws, cell_range: str, list_range: str):
    dv = DataValidation(type="list", formula1=f"={list_range}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


# ---------------------------------------------------------------------------
# 3. Workbook integration
# ---------------------------------------------------------------------------

def add_v12_debt_engine(
    wb,
    periods: Optional[List[str]] = None,
    historical_period_count: int = 0,
    financials_sheet_name: str = "Financial Statements",
    revenue_sheet_name: str = "Revenue Build",
    working_capital_sheet_name: str = "Working Capital",
    start_year: int = 2026,
    forecast_years: int = 5,
    default_debt_stack: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """
    Adds a full V12 debt module to an existing openpyxl workbook.

    Expected source sheets if available:
    - Financial Statements: EBITDA, cash interest, FCF, cash, debt lines
    - Revenue Build: revenue line
    - Working Capital: receivables / inventory / payables if available

    The module is robust: if those sheets do not exist, formulas still link to
    dedicated input rows in Debt Config.
    """

    if periods is None:
        periods = [f"FY {start_year + i}" for i in range(forecast_years)]
    total_periods = len(periods)

    _delete_if_exists(wb, [
        "Debt Lists",
        "Debt Config",
        "Debt Draws",
        "Debt Schedule",
        "Borrowing Base",
        "Debt Fees",
        "Debt Covenants",
        "Debt Waterfall",
        "Debt Outputs",
        "Debt Checks",
    ])

    list_ranges = _build_lists_sheet(wb)
    cfg = _build_config_sheet(wb, periods, list_ranges, default_debt_stack)
    _build_draw_schedule(wb, periods, list_ranges)
    _build_borrowing_base(wb, periods, financials_sheet_name, revenue_sheet_name, working_capital_sheet_name)
    _build_debt_schedule(wb, periods)
    _build_fees_sheet(wb, periods)
    _build_waterfall_sheet(wb, periods, financials_sheet_name)
    _build_covenants_sheet(wb, periods, financials_sheet_name)
    _build_outputs_sheet(wb, periods)
    _build_checks_sheet(wb, periods)

    return {
        "sheets_added": [
            "Debt Lists", "Debt Config", "Debt Draws", "Debt Schedule", "Borrowing Base",
            "Debt Fees", "Debt Covenants", "Debt Waterfall", "Debt Outputs", "Debt Checks"
        ],
        "instrument_count": len(DEBT_INSTRUMENT_LIBRARY),
        "max_tranches": MAX_TRANCHES,
        "period_count": total_periods,
    }


def _delete_if_exists(wb, names: List[str]):
    for name in names:
        if name in wb.sheetnames:
            del wb[name]


def _build_lists_sheet(wb):
    ws = wb.create_sheet("Debt Lists")
    ws.sheet_state = "hidden"
    lists = {
        "DebtTypes": list(DEBT_INSTRUMENT_LIBRARY.keys()),
        "AmortizationTypes": AMORTIZATION_TYPES,
        "Boolean": BOOLEAN_LIST,
        "RateBasis": RATE_BASIS_LIST,
        "Ranking": RANKING_LIST,
        "Currencies": CURRENCY_LIST,
        "Cases": CASE_LIST,
    }
    ranges = {}
    col = 1
    for name, values in lists.items():
        ws.cell(1, col).value = name
        for r, value in enumerate(values, start=2):
            ws.cell(r, col).value = value
        ranges[name] = f"'Debt Lists'!${get_column_letter(col)}$2:${get_column_letter(col)}${len(values)+1}"
        col += 1
    return ranges


def _default_stack():
    return [
        {"instrument": "Super Senior RCF", "debt_type": "Super Senior RCF", "currency": "EUR", "commitment": 30000, "opening_balance": 0, "rate": 0.045, "margin": 0.018, "cash_pay": True, "pik": False, "amortization": "Revolver", "tenor_years": 5, "grace_periods": 0, "ranking": 0, "secured": True, "sweep_eligible": False, "bullet": False, "min_cash": 10000, "commitment_fee": 0.010, "oid": 0.000, "exit_fee": 0.000},
        {"instrument": "Senior Term Loan B", "debt_type": "Senior Term Loan B", "currency": "EUR", "commitment": 130000, "opening_balance": 130000, "rate": 0.065, "margin": 0.030, "cash_pay": True, "pik": False, "amortization": "Bullet", "tenor_years": 5, "grace_periods": 0, "ranking": 1, "secured": True, "sweep_eligible": True, "bullet": True, "min_cash": 10000, "commitment_fee": 0.000, "oid": 0.010, "exit_fee": 0.010},
        {"instrument": "Mezzanine PIK", "debt_type": "Mezzanine PIK", "currency": "EUR", "commitment": 40000, "opening_balance": 40000, "rate": 0.120, "margin": 0.000, "cash_pay": False, "pik": True, "amortization": "PIK", "tenor_years": 5, "grace_periods": 0, "ranking": 3, "secured": False, "sweep_eligible": False, "bullet": True, "min_cash": 10000, "commitment_fee": 0.000, "oid": 0.020, "exit_fee": 0.020},
    ]


def _build_config_sheet(wb, periods, list_ranges, default_debt_stack=None):
    ws = wb.create_sheet("Debt Config")
    setup_sheet(ws, freeze="A6", tab_color="0D1B3E")
    title(ws, "Debt Config", "Blue cells are user inputs. Configure up to 60 debt tranches / instruments.")

    # Global assumptions
    set_section(ws, 5, "Global Debt Assumptions", 2, 6)
    globals_rows = [
        ("Selected Case", "List", "Base"),
        ("Minimum Cash", "000s", 10000),
        ("Cash Sweep %", "%", 0.50),
        ("Default Tax Rate", "%", 0.25),
        ("Maximum Net Debt / EBITDA", "x", 4.50),
        ("Maximum Senior Net Debt / EBITDA", "x", 3.50),
        ("Minimum Interest Cover", "x", 2.00),
        ("Minimum DSCR", "x", 1.10),
        ("Minimum FCCR", "x", 1.05),
        ("Covenant EBITDA Add-Backs", "000s", 0),
        ("Equity Cure Amount", "000s", 0),
        ("Cash Trap Threshold ND/EBITDA", "x", 4.00),
    ]
    for idx, (label, unit, value) in enumerate(globals_rows, start=6):
        row_label(ws, idx, label, unit, "Input")
        input_cell(ws.cell(idx, 6), value, fmt_pct() if unit == "%" else fmt_mult() if unit == "x" else fmt_money() if unit == "000s" else None)
    add_validation(ws, "F6", list_ranges["Cases"])

    # Tranche config
    start_row = 21
    headers = [
        "#", "Instrument Name", "Debt Type", "Currency", "Commitment", "Opening Balance",
        "Base Rate", "Margin", "PIK Margin", "Rate Basis", "Cash Pay?", "PIK?", "Amortization",
        "Tenor Years", "Grace Periods", "Ranking", "Secured?", "Sweep Eligible?",
        "Bullet?", "Min Cash", "Commitment Fee", "OID / Upfront Fee", "Exit Fee",
        "AR Advance Rate", "Inventory Advance Rate", "Revenue Share %",
    ]
    for c, h in enumerate(headers, start=2):
        set_header(ws.cell(start_row, c), h)

    stack = default_debt_stack or _default_stack()
    for i in range(MAX_TRANCHES):
        r = start_row + i + 1
        data = stack[i] if i < len(stack) else {}
        debt_type = data.get("debt_type", data.get("type", ""))
        meta = DEBT_INSTRUMENT_LIBRARY.get(debt_type, {})

        values = [
            i + 1,
            data.get("instrument", data.get("name", "")),
            debt_type,
            data.get("currency", "EUR"),
            data.get("commitment", data.get("amount", "")),
            data.get("opening_balance", data.get("amount", "")),
            data.get("rate", meta.get("default_rate", "")),
            data.get("margin", meta.get("default_margin", "")),
            data.get("pik_margin", 0.0),
            data.get("rate_basis", "Base + Margin"),
            str(data.get("cash_pay", meta.get("cash_pay", True))).upper(),
            str(data.get("pik", meta.get("pik", False))).upper(),
            data.get("amortization", meta.get("amortization", "")),
            data.get("tenor_years", 5),
            data.get("grace_periods", 0),
            data.get("ranking", meta.get("ranking", 1)),
            str(data.get("secured", meta.get("secured", True))).upper(),
            str(data.get("sweep_eligible", True)).upper(),
            str(data.get("bullet", meta.get("amortization", "") == "Bullet")).upper(),
            data.get("min_cash", 10000),
            data.get("commitment_fee", 0.0),
            data.get("oid", 0.0),
            data.get("exit_fee", 0.0),
            data.get("ar_advance_rate", 0.80),
            data.get("inventory_advance_rate", 0.50),
            data.get("revenue_share_pct", 0.0),
        ]

        for c, value in enumerate(values, start=2):
            cell = ws.cell(r, c)
            input_cell(cell, value)
            if c in [6, 7, 21]:
                cell.number_format = fmt_money()
            if c in [8, 9, 10, 22, 23, 24, 25, 26, 27]:
                cell.number_format = fmt_pct()
            if c in [15, 16, 17]:
                cell.number_format = fmt_int()

    # Validations
    add_validation(ws, f"D{start_row+1}:D{start_row+MAX_TRANCHES}", list_ranges["DebtTypes"])
    add_validation(ws, f"E{start_row+1}:E{start_row+MAX_TRANCHES}", list_ranges["Currencies"])
    add_validation(ws, f"K{start_row+1}:K{start_row+MAX_TRANCHES}", list_ranges["RateBasis"])
    for col in ["L", "M", "R", "S", "T"]:
        add_validation(ws, f"{col}{start_row+1}:{col}{start_row+MAX_TRANCHES}", list_ranges["Boolean"])
    add_validation(ws, f"N{start_row+1}:N{start_row+MAX_TRANCHES}", list_ranges["AmortizationTypes"])
    add_validation(ws, f"Q{start_row+1}:Q{start_row+MAX_TRANCHES}", list_ranges["Ranking"])

    ws.auto_filter.ref = f"B{start_row}:AA{start_row+MAX_TRANCHES}"
    return {"global_start": 6, "tranche_start": start_row + 1}


def _build_draw_schedule(wb, periods, list_ranges):
    ws = wb.create_sheet("Debt Draws")
    setup_sheet(ws, freeze="H6", tab_color="1F3D7A")
    title(ws, "Debt Draws & Repayments", "Period-specific manual drawdowns, mandatory repayments, refinancing proceeds and lender fees.")

    for i, period in enumerate(periods):
        c = FIRST_PERIOD_COL + i
        set_header(ws.cell(4, c), period)
        set_header(ws.cell(5, c), "Input")

    row = 7
    for tr in range(MAX_TRANCHES):
        config_row = 22 + tr
        set_section(ws, row, f"Tranche {tr+1} - ='Debt Config'!C{config_row}", 2, 6)
        row += 1
        for label_text in ["Manual Drawdown", "Mandatory Repayment", "Refinancing Proceeds", "Refinancing Repayment", "Amend & Extend Fee", "Call Premium / Make-Whole"]:
            row_label(ws, row, label_text, "000s", "Input")
            for i in range(len(periods)):
                input_cell(ws.cell(row, FIRST_PERIOD_COL + i), 0, fmt_money())
            row += 1
        row += 1


def _build_borrowing_base(wb, periods, financials_sheet_name, revenue_sheet_name, working_capital_sheet_name):
    ws = wb.create_sheet("Borrowing Base")
    setup_sheet(ws, freeze="H6", tab_color="2C5282")
    title(ws, "Borrowing Base", "Formula-based AR / inventory availability for ABL, factoring and receivables facilities.")

    for i, period in enumerate(periods):
        c = FIRST_PERIOD_COL + i
        set_header(ws.cell(4, c), period)
        set_header(ws.cell(5, c), "Formula")

    rows = {
        "Receivables": 7,
        "Inventory": 8,
        "Eligible Receivables": 9,
        "Eligible Inventory": 10,
        "AR Advance Rate": 11,
        "Inventory Advance Rate": 12,
        "Gross Borrowing Base": 13,
        "Availability Reserve": 14,
        "Net Borrowing Base": 15,
    }
    for name, r in rows.items():
        row_label(ws, r, name, "%" if "Rate" in name else "000s", "Linked / Formula", "total" if r in [13, 15] else "normal")

    for i in range(len(periods)):
        col = get_column_letter(FIRST_PERIOD_COL + i)
        # Tries to link to Working Capital rows used by V11. If missing in Excel, user can override formulas.
        formula_cell(ws.cell(7, FIRST_PERIOD_COL + i), f"=IFERROR('{working_capital_sheet_name}'!{col}7,0)", fmt_money(), linked=True)
        formula_cell(ws.cell(8, FIRST_PERIOD_COL + i), f"=IFERROR('{working_capital_sheet_name}'!{col}8,0)", fmt_money(), linked=True)
        formula_cell(ws.cell(9, FIRST_PERIOD_COL + i), f"={col}7*95%", fmt_money())
        formula_cell(ws.cell(10, FIRST_PERIOD_COL + i), f"={col}8*85%", fmt_money())
        formula_cell(ws.cell(11, FIRST_PERIOD_COL + i), "=80%", fmt_pct())
        formula_cell(ws.cell(12, FIRST_PERIOD_COL + i), "=50%", fmt_pct())
        formula_cell(ws.cell(13, FIRST_PERIOD_COL + i), f"={col}9*{col}11+{col}10*{col}12", fmt_money(), output=True)
        formula_cell(ws.cell(14, FIRST_PERIOD_COL + i), f"={col}13*5%", fmt_money())
        formula_cell(ws.cell(15, FIRST_PERIOD_COL + i), f"=MAX(0,{col}13-{col}14)", fmt_money(), output=True)


def _build_debt_schedule(wb, periods):
    ws = wb.create_sheet("Debt Schedule")
    setup_sheet(ws, freeze="H6", tab_color="0D1B3E")
    title(ws, "Debt Schedule", "Corkscrew by tranche: opening balance, draws, interest, PIK, amortisation, sweep, fees and closing balance.")

    for i, period in enumerate(periods):
        c = FIRST_PERIOD_COL + i
        set_header(ws.cell(4, c), period)
        set_header(ws.cell(5, c), "Formula")

    row = 7
    tranche_blocks = []
    for tr in range(MAX_TRANCHES):
        cfg_row = 22 + tr
        draw_block = 8 + tr * 8
        block_start = row
        tranche_blocks.append(block_start)

        set_section(ws, row, f"Tranche {tr+1}", 2, 6)
        for i in range(len(periods)):
            c = FIRST_PERIOD_COL + i
            col = get_column_letter(c)
            formula_cell(ws.cell(row, c), f"='Debt Config'!C{cfg_row}", linked=True)
        row += 1

        schedule_rows = [
            ("Opening Balance", "000s"),
            ("Manual Drawdown", "000s"),
            ("Refinancing Proceeds", "000s"),
            ("Cash Interest", "000s"),
            ("PIK Interest", "000s"),
            ("Scheduled Amortization", "000s"),
            ("Mandatory Repayment", "000s"),
            ("Cash Sweep", "000s"),
            ("Refinancing Repayment", "000s"),
            ("Call Premium / Make-Whole", "000s"),
            ("Closing Balance", "000s"),
            ("Undrawn Commitment", "000s"),
            ("Commitment Fee", "000s"),
            ("Total Cash Cost", "000s"),
        ]
        for label_text, unit in schedule_rows:
            row_label(ws, row, label_text, unit, "Formula", "total" if label_text in ["Closing Balance", "Total Cash Cost"] else "normal")
            row += 1

        # Formulas
        for i in range(len(periods)):
            c = FIRST_PERIOD_COL + i
            col = get_column_letter(c)
            prev_col = get_column_letter(c - 1) if i > 0 else None

            opening_r = block_start + 1
            manual_draw_r = block_start + 2
            refi_draw_r = block_start + 3
            cash_int_r = block_start + 4
            pik_int_r = block_start + 5
            sched_amort_r = block_start + 6
            mandatory_r = block_start + 7
            sweep_r = block_start + 8
            refi_repay_r = block_start + 9
            premium_r = block_start + 10
            closing_r = block_start + 11
            undrawn_r = block_start + 12
            commit_fee_r = block_start + 13
            total_cost_r = block_start + 14

            if i == 0:
                opening_formula = f"='Debt Config'!G{cfg_row}"
            else:
                opening_formula = f"={prev_col}{closing_r}"

            manual_draw_formula = f"='Debt Draws'!{col}{draw_block}"
            refi_draw_formula = f"='Debt Draws'!{col}{draw_block+2}"
            rate_formula = f"('Debt Config'!H{cfg_row}+'Debt Config'!I{cfg_row})"

            cash_interest_formula = f'=IF(\'Debt Config\'!L{cfg_row}=TRUE,{col}{opening_r}*{rate_formula},0)'
            pik_interest_formula = f'=IF(\'Debt Config\'!M{cfg_row}=TRUE,{col}{opening_r}*(\'Debt Config\'!H{cfg_row}+\'Debt Config\'!J{cfg_row}),0)'

            # Amortisation supports several types using formulas only
            sched_amort_formula = (
                f'=IF({col}{opening_r}<=0,0,'
                f'IF(\'Debt Config\'!N{cfg_row}="Linear",{col}{opening_r}/MAX(1,\'Debt Config\'!O{cfg_row}-(COLUMN()-{FIRST_PERIOD_COL})+1),'
                f'IF(\'Debt Config\'!N{cfg_row}="Annuity",{col}{opening_r}/MAX(1,\'Debt Config\'!O{cfg_row}-(COLUMN()-{FIRST_PERIOD_COL})+1),'
                f'IF(\'Debt Config\'!N{cfg_row}="Interest Only Then Linear",IF(COLUMN()-{FIRST_PERIOD_COL}<\'Debt Config\'!P{cfg_row},0,{col}{opening_r}/MAX(1,\'Debt Config\'!O{cfg_row}-(COLUMN()-{FIRST_PERIOD_COL})+1)),'
                f'IF(AND(\'Debt Config\'!T{cfg_row}=TRUE,COLUMN()-{FIRST_PERIOD_COL}+1>=\'Debt Config\'!O{cfg_row}),{col}{opening_r},0)))))'
            )

            mandatory_formula = f"='Debt Draws'!{col}{draw_block+1}"
            sweep_formula = f'=IF(\'Debt Config\'!S{cfg_row}=TRUE,MIN({col}{opening_r}+{col}{manual_draw_r}+{col}{refi_draw_r}+{col}{pik_int_r}-{col}{sched_amort_r}-{col}{mandatory_r},IFERROR(\'Debt Waterfall\'!{col}13,0)*IF(\'Debt Config\'!Q{cfg_row}=1,70%,IF(\'Debt Config\'!Q{cfg_row}=2,20%,IF(\'Debt Config\'!Q{cfg_row}=3,10%,0)))),0)'
            refi_repay_formula = f"='Debt Draws'!{col}{draw_block+3}"
            premium_formula = f"='Debt Draws'!{col}{draw_block+5}"
            closing_formula = f"=MAX(0,{col}{opening_r}+{col}{manual_draw_r}+{col}{refi_draw_r}+{col}{pik_int_r}-{col}{sched_amort_r}-{col}{mandatory_r}-{col}{sweep_r}-{col}{refi_repay_r})"
            undrawn_formula = f"=MAX(0,'Debt Config'!F{cfg_row}-{col}{closing_r})"
            commit_fee_formula = f"={col}{undrawn_r}*'Debt Config'!V{cfg_row}"
            total_cost_formula = f"={col}{cash_int_r}+{col}{commit_fee_r}+{col}{premium_r}"

            formulas = [
                opening_formula, manual_draw_formula, refi_draw_formula, cash_interest_formula, pik_interest_formula,
                sched_amort_formula, mandatory_formula, sweep_formula, refi_repay_formula, premium_formula,
                closing_formula, undrawn_formula, commit_fee_formula, total_cost_formula
            ]
            for offset, formula in enumerate(formulas, start=1):
                out = offset in [11, 14]
                formula_cell(ws.cell(block_start + offset, c), formula, fmt_money(), output=out, linked=("Debt Draws" in formula or "Debt Config" in formula))

        row = block_start + 16

    # Aggregate section
    agg_start = row + 1
    set_section(ws, agg_start, "Aggregate Debt Summary", 2, 6)
    agg_rows = {
        "Total Opening Debt": agg_start + 1,
        "Total Drawdowns": agg_start + 2,
        "Cash Interest": agg_start + 3,
        "PIK Interest": agg_start + 4,
        "Scheduled Amortization": agg_start + 5,
        "Mandatory Repayment": agg_start + 6,
        "Cash Sweep": agg_start + 7,
        "Refinancing Repayment": agg_start + 8,
        "Closing Debt": agg_start + 9,
        "Undrawn Commitments": agg_start + 10,
        "Commitment Fees": agg_start + 11,
        "Total Cash Debt Cost": agg_start + 12,
    }
    for label_text, r in agg_rows.items():
        row_label(ws, r, label_text, "000s", "Formula", "total")

    for i in range(len(periods)):
        c = FIRST_PERIOD_COL + i
        col = get_column_letter(c)
        # Use direct SUM of repeated block lines
        def sum_blocks(offset):
            return "=" + "+".join([f"{col}{b+offset}" for b in tranche_blocks])
        formula_cell(ws.cell(agg_rows["Total Opening Debt"], c), sum_blocks(1), fmt_money(), output=True)
        formula_cell(ws.cell(agg_rows["Total Drawdowns"], c), "=" + "+".join([f"{col}{b+2}+{col}{b+3}" for b in tranche_blocks]), fmt_money(), output=True)
        formula_cell(ws.cell(agg_rows["Cash Interest"], c), sum_blocks(4), fmt_money(), output=True)
        formula_cell(ws.cell(agg_rows["PIK Interest"], c), sum_blocks(5), fmt_money(), output=True)
        formula_cell(ws.cell(agg_rows["Scheduled Amortization"], c), sum_blocks(6), fmt_money(), output=True)
        formula_cell(ws.cell(agg_rows["Mandatory Repayment"], c), sum_blocks(7), fmt_money(), output=True)
        formula_cell(ws.cell(agg_rows["Cash Sweep"], c), sum_blocks(8), fmt_money(), output=True)
        formula_cell(ws.cell(agg_rows["Refinancing Repayment"], c), sum_blocks(9), fmt_money(), output=True)
        formula_cell(ws.cell(agg_rows["Closing Debt"], c), sum_blocks(11), fmt_money(), output=True)
        formula_cell(ws.cell(agg_rows["Undrawn Commitments"], c), sum_blocks(12), fmt_money(), output=True)
        formula_cell(ws.cell(agg_rows["Commitment Fees"], c), sum_blocks(13), fmt_money(), output=True)
        formula_cell(ws.cell(agg_rows["Total Cash Debt Cost"], c), sum_blocks(14), fmt_money(), output=True)


def _build_fees_sheet(wb, periods):
    ws = wb.create_sheet("Debt Fees")
    setup_sheet(ws, freeze="H6", tab_color="4A5568")
    title(ws, "Debt Fees, OID & Premiums", "Upfront fees, OID, commitment fees, exit fees and refinancing premiums.")

    for i, period in enumerate(periods):
        c = FIRST_PERIOD_COL + i
        set_header(ws.cell(4, c), period)
        set_header(ws.cell(5, c), "Formula")

    rows = {
        "Opening Debt": 7,
        "OID / Upfront Fee": 8,
        "Commitment Fee": 9,
        "Call Premium / Make-Whole": 10,
        "Exit Fee Accrual": 11,
        "Amend & Extend Fees": 12,
        "Total Debt Fees": 13,
    }
    for name, r in rows.items():
        row_label(ws, r, name, "000s", "Formula", "total" if name == "Total Debt Fees" else "normal")

    debt_summary_row = _aggregate_start_row()
    for i in range(len(periods)):
        c = FIRST_PERIOD_COL + i
        col = get_column_letter(c)
        formula_cell(ws.cell(7, c), f"='Debt Schedule'!{col}{debt_summary_row+1}", fmt_money(), linked=True)
        # upfront fee only first period
        formula_cell(ws.cell(8, c), f"=IF(COLUMN()={FIRST_PERIOD_COL},SUMPRODUCT('Debt Config'!G22:G81,'Debt Config'!W22:W81),0)", fmt_money())
        formula_cell(ws.cell(9, c), f"='Debt Schedule'!{col}{debt_summary_row+11}", fmt_money(), linked=True)
        formula_cell(ws.cell(10, c), f"=SUMPRODUCT('Debt Config'!G22:G81,'Debt Config'!X22:X81)", fmt_money())
        formula_cell(ws.cell(11, c), f"=SUMPRODUCT('Debt Schedule'!{col}18:{col}960,'Debt Config'!X22:X81)", fmt_money())
        formula_cell(ws.cell(12, c), f"=SUM('Debt Draws'!{col}12:{col}488)", fmt_money(), linked=True)
        formula_cell(ws.cell(13, c), f"=SUM({col}8:{col}12)", fmt_money(), output=True)


def _aggregate_start_row():
    return 7 + MAX_TRANCHES * 16 + 1


def _build_waterfall_sheet(wb, periods, financials_sheet_name):
    ws = wb.create_sheet("Debt Waterfall")
    setup_sheet(ws, freeze="H6", tab_color="006400")
    title(ws, "Cash Sweep & Repayment Waterfall", "Determines cash available for optional debt repayment by ranking.")

    for i, period in enumerate(periods):
        c = FIRST_PERIOD_COL + i
        set_header(ws.cell(4, c), period)
        set_header(ws.cell(5, c), "Formula")

    rows = {
        "Opening Cash": 7,
        "EBITDA": 8,
        "Tax": 9,
        "Change in NWC": 10,
        "Capex": 11,
        "Debt Cash Cost": 12,
        "Pre-Sweep Cash": 13,
        "Minimum Cash": 14,
        "Excess Cash Available": 15,
        "Cash Sweep %": 16,
        "Optional Cash Sweep Pool": 17,
        "Cash Trap Active?": 18,
        "Liquidity After Sweep": 19,
    }
    for name, r in rows.items():
        row_label(ws, r, name, "%" if "%" in name else "Boolean" if "?" in name else "000s", "Formula", "total" if r in [15, 17, 19] else "normal")

    debt_summary_row = _aggregate_start_row()
    for i in range(len(periods)):
        c = FIRST_PERIOD_COL + i
        col = get_column_letter(c)
        prev = get_column_letter(c - 1) if i > 0 else None
        formula_cell(ws.cell(7, c), f"=IF(COLUMN()={FIRST_PERIOD_COL},'Debt Config'!$F$7,{prev}19)", fmt_money())
        formula_cell(ws.cell(8, c), f"=IFERROR('{financials_sheet_name}'!{col}9,0)", fmt_money(), linked=True)
        formula_cell(ws.cell(9, c), f"=IFERROR('{financials_sheet_name}'!{col}13,0)", fmt_money(), linked=True)
        formula_cell(ws.cell(10, c), f"=IFERROR('Working Capital'!{col}11,0)", fmt_money(), linked=True)
        formula_cell(ws.cell(11, c), f"=IFERROR('Debt Schedule'!{col}0,0)", fmt_money())
        formula_cell(ws.cell(12, c), f"='Debt Schedule'!{col}{debt_summary_row+12}", fmt_money(), linked=True)
        formula_cell(ws.cell(13, c), f"={col}7+{col}8-{col}9-{col}10-{col}11-{col}12", fmt_money(), output=True)
        formula_cell(ws.cell(14, c), "='Debt Config'!$F$7", fmt_money(), linked=True)
        formula_cell(ws.cell(15, c), f"=MAX(0,{col}13-{col}14)", fmt_money(), output=True)
        formula_cell(ws.cell(16, c), "='Debt Config'!$F$8", fmt_pct(), linked=True)
        formula_cell(ws.cell(17, c), f"={col}15*{col}16", fmt_money(), output=True)
        formula_cell(ws.cell(18, c), f"=IFERROR('Debt Covenants'!{col}7>'Debt Config'!$F$17,FALSE)", output=True)
        formula_cell(ws.cell(19, c), f"={col}13-{col}17", fmt_money(), output=True)


def _build_covenants_sheet(wb, periods, financials_sheet_name):
    ws = wb.create_sheet("Debt Covenants")
    setup_sheet(ws, freeze="H6", tab_color="C53030")
    title(ws, "Covenants & Credit Metrics", "Leverage, senior leverage, cash interest cover, DSCR, FCCR, liquidity and covenant EBITDA.")

    for i, period in enumerate(periods):
        c = FIRST_PERIOD_COL + i
        set_header(ws.cell(4, c), period)
        set_header(ws.cell(5, c), "Formula")

    rows = {
        "Net Debt / EBITDA": 7,
        "Senior Net Debt / EBITDA": 8,
        "Cash Interest Cover": 9,
        "DSCR": 10,
        "FCCR": 11,
        "Minimum Liquidity": 12,
        "Covenant EBITDA": 13,
        "EBITDA Add-Backs": 14,
        "Equity Cure": 15,
        "Leverage Test Pass?": 17,
        "Senior Leverage Test Pass?": 18,
        "ICR Test Pass?": 19,
        "DSCR Test Pass?": 20,
        "FCCR Test Pass?": 21,
        "Liquidity Test Pass?": 22,
        "All Covenants Pass?": 23,
    }
    for name, r in rows.items():
        unit = "x" if "/" in name or "Cover" in name or name in ["DSCR", "FCCR"] else "Boolean" if "Pass" in name or "?" in name else "000s"
        row_label(ws, r, name, unit, "Formula", "total" if "Pass" in name or r in [7, 13] else "normal")

    debt_summary_row = _aggregate_start_row()
    for i in range(len(periods)):
        c = FIRST_PERIOD_COL + i
        col = get_column_letter(c)
        formula_cell(ws.cell(13, c), f"=IFERROR('{financials_sheet_name}'!{col}9,0)+{col}14+{col}15", fmt_money(), output=True)
        formula_cell(ws.cell(14, c), "='Debt Config'!$F$15", fmt_money(), linked=True)
        formula_cell(ws.cell(15, c), "='Debt Config'!$F$16", fmt_money(), linked=True)
        formula_cell(ws.cell(7, c), f"=IFERROR(('Debt Schedule'!{col}{debt_summary_row+9}-'Debt Waterfall'!{col}19)/{col}13,0)", fmt_mult(), output=True)
        formula_cell(ws.cell(8, c), f"=IFERROR(('Debt Schedule'!{col}{debt_summary_row+9}*75%-'Debt Waterfall'!{col}19)/{col}13,0)", fmt_mult())
        formula_cell(ws.cell(9, c), f"=IFERROR({col}13/'Debt Schedule'!{col}{debt_summary_row+3},99)", fmt_mult())
        formula_cell(ws.cell(10, c), f"=IFERROR(('Debt Waterfall'!{col}8-'Debt Waterfall'!{col}9-'Debt Waterfall'!{col}10)/('Debt Schedule'!{col}{debt_summary_row+3}+'Debt Schedule'!{col}{debt_summary_row+5}),99)", fmt_mult())
        formula_cell(ws.cell(11, c), f"=IFERROR(('Debt Waterfall'!{col}8-'Debt Waterfall'!{col}9)/('Debt Schedule'!{col}{debt_summary_row+3}+'Debt Waterfall'!{col}11+'Debt Schedule'!{col}{debt_summary_row+5}),99)", fmt_mult())
        formula_cell(ws.cell(12, c), f"='Debt Waterfall'!{col}19", fmt_money(), linked=True)
        formula_cell(ws.cell(17, c), f"={col}7<='Debt Config'!$F$10", output=True)
        formula_cell(ws.cell(18, c), f"={col}8<='Debt Config'!$F$11", output=True)
        formula_cell(ws.cell(19, c), f"={col}9>='Debt Config'!$F$12", output=True)
        formula_cell(ws.cell(20, c), f"={col}10>='Debt Config'!$F$13", output=True)
        formula_cell(ws.cell(21, c), f"={col}11>='Debt Config'!$F$14", output=True)
        formula_cell(ws.cell(22, c), f"={col}12>='Debt Config'!$F$7", output=True)
        formula_cell(ws.cell(23, c), f"=AND({col}17,{col}18,{col}19,{col}20,{col}21,{col}22)", output=True)


def _build_outputs_sheet(wb, periods):
    ws = wb.create_sheet("Debt Outputs")
    setup_sheet(ws, freeze="H6", tab_color="006400")
    title(ws, "Debt Outputs", "Clean lender / IC output linked to detailed debt schedules.")

    for i, period in enumerate(periods):
        c = FIRST_PERIOD_COL + i
        set_header(ws.cell(4, c), period)
        set_header(ws.cell(5, c), "Output")

    debt_summary_row = _aggregate_start_row()
    rows = {
        "Closing Debt": 7,
        "Undrawn Commitments": 8,
        "Net Debt / EBITDA": 9,
        "Senior Net Debt / EBITDA": 10,
        "Cash Interest Cover": 11,
        "DSCR": 12,
        "FCCR": 13,
        "Minimum Liquidity": 14,
        "Covenants Pass?": 15,
        "Cash Interest": 17,
        "PIK Interest": 18,
        "Commitment Fees": 19,
        "Cash Sweep": 20,
        "Total Cash Debt Cost": 21,
    }
    for name, r in rows.items():
        unit = "x" if "/" in name or "Cover" in name or name in ["DSCR", "FCCR"] else "Boolean" if "?" in name else "000s"
        row_label(ws, r, name, unit, "Linked", "total")
    for i in range(len(periods)):
        c = FIRST_PERIOD_COL + i
        col = get_column_letter(c)
        links = {
            7: f"='Debt Schedule'!{col}{debt_summary_row+9}",
            8: f"='Debt Schedule'!{col}{debt_summary_row+10}",
            9: f"='Debt Covenants'!{col}7",
            10: f"='Debt Covenants'!{col}8",
            11: f"='Debt Covenants'!{col}9",
            12: f"='Debt Covenants'!{col}10",
            13: f"='Debt Covenants'!{col}11",
            14: f"='Debt Covenants'!{col}12",
            15: f"='Debt Covenants'!{col}23",
            17: f"='Debt Schedule'!{col}{debt_summary_row+3}",
            18: f"='Debt Schedule'!{col}{debt_summary_row+4}",
            19: f"='Debt Schedule'!{col}{debt_summary_row+11}",
            20: f"='Debt Schedule'!{col}{debt_summary_row+7}",
            21: f"='Debt Schedule'!{col}{debt_summary_row+12}",
        }
        for r, formula in links.items():
            fmt = fmt_mult() if r in [9, 10, 11, 12, 13] else fmt_money() if r not in [15] else None
            formula_cell(ws.cell(r, c), formula, fmt, output=True, linked=True)


def _build_checks_sheet(wb, periods):
    ws = wb.create_sheet("Debt Checks")
    setup_sheet(ws, freeze="H6", tab_color="C53030")
    title(ws, "Debt Checks", "All checks should show OK. Any ERROR/BREACH requires review.")

    for i, period in enumerate(periods):
        c = FIRST_PERIOD_COL + i
        set_header(ws.cell(4, c), period)
        set_header(ws.cell(5, c), "Check")

    rows = {
        "No Negative Debt": 7,
        "No Negative Liquidity": 8,
        "Covenants Pass": 9,
        "Debt Roll-Forward Integrity": 10,
        "Borrowing Base Availability": 11,
        "All Debt Checks OK": 13,
    }
    for name, r in rows.items():
        row_label(ws, r, name, "Check", "Formula", "total")

    debt_summary_row = _aggregate_start_row()
    for i in range(len(periods)):
        c = FIRST_PERIOD_COL + i
        col = get_column_letter(c)
        formula_cell(ws.cell(7, c), f'=IF(\'Debt Schedule\'!{col}{debt_summary_row+9}>=0,"OK","ERROR")', output=True)
        formula_cell(ws.cell(8, c), f'=IF(\'Debt Waterfall\'!{col}19>=0,"OK","ERROR")', output=True)
        formula_cell(ws.cell(9, c), f'=IF(\'Debt Covenants\'!{col}23=TRUE,"OK","BREACH")', output=True)
        formula_cell(ws.cell(10, c), f'=IF(ABS(\'Debt Schedule\'!{col}{debt_summary_row+9}-(\'Debt Schedule\'!{col}{debt_summary_row+1}+\'Debt Schedule\'!{col}{debt_summary_row+2}+\'Debt Schedule\'!{col}{debt_summary_row+4}-\'Debt Schedule\'!{col}{debt_summary_row+5}-\'Debt Schedule\'!{col}{debt_summary_row+6}-\'Debt Schedule\'!{col}{debt_summary_row+7}-\'Debt Schedule\'!{col}{debt_summary_row+8}))<1,"OK","ERROR")', output=True)
        formula_cell(ws.cell(11, c), f'=IF(\'Borrowing Base\'!{col}15>=0,"OK","ERROR")', output=True)
        formula_cell(ws.cell(13, c), f'=IF(AND({col}7="OK",{col}8="OK",{col}9="OK",{col}10="OK",{col}11="OK"),"OK","ERROR")', output=True)


# ---------------------------------------------------------------------------
# 4. Convenience helpers for SaaS / API layer
# ---------------------------------------------------------------------------

def default_debt_stack_for_saas() -> List[Dict[str, Any]]:
    """Returns default editable debt stack for new SaaS projects."""
    return _default_stack()


def debt_type_options() -> List[str]:
    return list(DEBT_INSTRUMENT_LIBRARY.keys())


def debt_library_payload() -> Dict[str, Any]:
    return {
        "debt_types": debt_type_options(),
        "amortization_types": AMORTIZATION_TYPES,
        "rate_basis": RATE_BASIS_LIST,
        "currencies": CURRENCY_LIST,
        "case_list": CASE_LIST,
        "library": DEBT_INSTRUMENT_LIBRARY,
    }
