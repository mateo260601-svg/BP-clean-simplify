from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, Rule
import math, datetime
try:
    from debt_engine import parse_tranches, DebtEngine, BASE_RATES as _BASE_RATES
    from sheet_debt_v2 import build_debt_sheet
    _DEBT_ENGINE_AVAILABLE = True
except ImportError:
    _DEBT_ENGINE_AVAILABLE = False


# ══════════════════════════════════════════════════════════════════════════════
# DESIGN SYSTEM  —  Goldman / Lazard / Rothschild aesthetic
# ══════════════════════════════════════════════════════════════════════════════

FONT_NAME = "Calibri"

# Color palette
C = {
    # Backgrounds
    "hdr_year":    "0D1B3E",   # deep navy  — year row
    "hdr_qtr":     "1A2F5E",   # mid navy   — quarter row
    "hdr_month":   "243F70",   # lighter    — month row
    "hdr_tag_act": "4A5568",   # grey       — Actual tag
    "hdr_tag_fcst":"1F5C8B",   # blue       — Forecast tag
    "hdr_tag_bp":  "2D6A9F",   # blue       — BP tag
    "sec_1":       "0D1B3E",   # section header level 1
    "sec_2":       "1F3D7A",   # section header level 2
    "sec_3":       "2C5282",   # section header level 3
    "total":       "C5DCF5",   # total rows
    "subtot":      "DDEEFF",   # subtotal rows
    "ebitda_bg":   "0D1B3E",   # EBITDA highlight
    "input_bg":    "EBF5FF",   # input cells
    "grey_row":    "F7F8FA",   # alternate light rows
    "white":       "FFFFFF",
    "check_ok":    "E8F5E9",
    "check_err":   "FFEBEE",
    # Text
    "txt_white":   "FFFFFF",
    "txt_navy":    "0D1B3E",
    "txt_black":   "000000",
    "txt_input":   "0000CD",   # blue  — hardcoded inputs
    "txt_link":    "006400",   # green — cross-sheet links
    "txt_formula": "000000",   # black — formulas
    "txt_grey":    "718096",
    "txt_red":     "C53030",
}

def F(hex_): return PatternFill("solid", fgColor=hex_)
def FK(hex_,bold=False,size=9,italic=False,color=None):
    return Font(name=FONT_NAME, size=size, bold=bold,
                italic=italic, color=color or C["txt_black"])

def AL(h="left",v="center",wrap=False,indent=0):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap,indent=indent)

def _side(style="thin",color="D0D5DD"): return Side(style=style,color=color)
def _side_med(color="0D1B3E"):          return Side(style="medium",color=color)
def _side_none():                        return Side(style=None)

BORDER_THIN  = Border(left=_side(),right=_side(),top=_side(),bottom=_side())
BORDER_BOT   = Border(bottom=_side_med())
BORDER_TOP   = Border(top=_side_med())
BORDER_TB    = Border(top=_side_med(),bottom=_side_med())
BORDER_NONE  = Border()

FMT = {
    "int":   '#,##0_);(#,##0);"-"',
    "dec1":  '#,##0.0_);(#,##0.0);"-"',
    "dec2":  '#,##0.00_);(#,##0.00);"-"',
    "pct1":  '0.0%_);(0.0%);"-"',
    "pct0":  '0%_);(0%);"-"',
    "mult":  '0.0"x"',
    "usdmt": '$#,##0.00_);($#,##0.00);"-"',
    "date":  'MMM-YY',
    "year":  '"FY "0',
    "text":  "@",
    "neg":   '#,##0_);(#,##0)',
}

MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]

def col_l(c): return get_column_letter(c)

# ── Layout constants ──────────────────────────────────────────────────────────
LCOL  = 2   # Label
UCOL  = 3   # Unit
DCOL  = 5   # First data column
HDR_ROWS = 4 # Year / Quarter / Month / Tag  (rows below title)

# ══════════════════════════════════════════════════════════════════════════════
# PERIOD UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def gen_periods(start_year, n_months, actuals_months=0):
    periods = []
    y, m = start_year, 1
    for i in range(n_months):
        q = (m-1)//3 + 1
        tag = "Actual" if i < actuals_months else "BP"
        periods.append(dict(
            year=y, month=m, quarter=q, idx=i,
            label=f"{MONTH_ABBR[m-1]}-{str(y)[2:]}",
            year_label=f"FY {y}",
            qtr_label=f"Q{q} {str(y)[2:]}",
            tag=tag, is_actual=(i < actuals_months),
        ))
        m += 1
        if m > 12: m = 1; y += 1
    return periods

def annual_groups(periods):
    g = {}
    for p in periods:
        g.setdefault(p["year"], []).append(p["idx"])
    return g

def qtr_groups(periods):
    g = {}
    for p in periods:
        k = p["qtr_label"]
        g.setdefault(k, []).append(p["idx"])
    return g

def ann_sum(vals, ag, year):
    return sum(vals[i] for i in ag.get(year,[]) if i < len(vals) and vals[i] is not None)

def ann_avg(vals, ag, year):
    idxs = ag.get(year,[])
    valid = [vals[i] for i in idxs if i < len(vals)]
    return sum(valid)/len(valid) if valid else 0

# ══════════════════════════════════════════════════════════════════════════════
# SHEET SETUP
# ══════════════════════════════════════════════════════════════════════════════

def setup_ws(wb, name, tab_color, zoom=75, freeze="F7"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = freeze
    ws.page_setup.paperSize   = 9   # A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.4
    ws.page_margins.right  = 0.4
    ws.page_margins.top    = 0.6
    ws.page_margins.bottom = 0.6
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    ws.oddHeader.left.text  = f"&\"Calibri,Bold\"&9{name}"
    ws.oddHeader.right.text = "&\"Calibri\"&9Confidential"
    ws.oddFooter.left.text  = "&\"Calibri\"&8Prepared by MG Advisory"
    ws.oddFooter.center.text= "&\"Calibri\"&8Page &P of &N"
    ws.oddFooter.right.text = "&\"Calibri\"&8&D"
    return ws

def set_col_widths(ws, n_months, annual_cols):
    ws.column_dimensions["A"].width            = 1.5
    ws.column_dimensions[col_l(LCOL)].width    = 44
    ws.column_dimensions[col_l(UCOL)].width    = 8
    ws.column_dimensions[col_l(DCOL-1)].width  = 1  # spacer
    for i in range(n_months):
        ws.column_dimensions[col_l(DCOL+i)].width = 10.5
    if annual_cols:
        ws.column_dimensions[col_l(annual_cols[0]["col"]-1)].width = 1.5  # spacer
    for ac in (annual_cols or []):
        ws.column_dimensions[col_l(ac["col"])].width = 13

# ══════════════════════════════════════════════════════════════════════════════
# HEADER BLOCK  (4 rows: Year / Quarter / Month / Tag)
# ══════════════════════════════════════════════════════════════════════════════

def write_title(ws, row, company, sheet_title, unit_lbl, last_col):
    ws.row_dimensions[row].height   = 32
    ws.row_dimensions[row+1].height = 4
    # Navy bar
    for c in range(LCOL, last_col+1):
        ws.cell(row, c).fill = F(C["hdr_year"])
    cell = ws.cell(row, LCOL)
    cell.value     = f"  {company}   ·   {sheet_title}   ·   {unit_lbl}"
    cell.font      = FK(C["txt_white"], bold=True, size=11)
    cell.alignment = AL("left")
    # thin accent line below title
    for c in range(LCOL, last_col+1):
        ws.cell(row+1, c).fill = F("3A6BBF")

def write_period_headers(ws, row_start, periods, annual_cols=None):
    """
    row_start : first of 4 header rows
    Row 0 : Year  (merged spans)
    Row 1 : Quarter (merged spans)
    Row 2 : Month label (individual)
    Row 3 : Actual / BP tag (individual)
    """
    R_YR, R_QT, R_MN, R_TG = row_start, row_start+1, row_start+2, row_start+3
    ws.row_dimensions[R_YR].height = 14
    ws.row_dimensions[R_QT].height = 12
    ws.row_dimensions[R_MN].height = 14
    ws.row_dimensions[R_TG].height = 11

    # ── Year spans ────────────────────────────────────────────────────────────
    spans = []
    cur_yr, span_start = None, None
    for p in periods:
        if p["year"] != cur_yr:
            if cur_yr is not None:
                spans.append((cur_yr, span_start, DCOL+p["idx"]-1))
            cur_yr, span_start = p["year"], DCOL+p["idx"]
    if cur_yr is not None:
        spans.append((cur_yr, span_start, DCOL+len(periods)-1))

    for yr, c1, c2 in spans:
        if c1 == c2:
            cell = ws.cell(R_YR, c1)
        else:
            ws.merge_cells(start_row=R_YR,start_column=c1,end_row=R_YR,end_column=c2)
            cell = ws.cell(R_YR, c1)
        cell.value     = f"FY {yr}"
        cell.fill      = F(C["hdr_year"])
        cell.font      = FK(C["txt_white"], bold=True, size=9)
        cell.alignment = AL("center")
        cell.border    = Border(left=_side_med(), right=_side_med())

    # ── Quarter spans ─────────────────────────────────────────────────────────
    q_spans = []
    cur_qk, q_start = None, None
    for p in periods:
        qk = p["qtr_label"]
        if qk != cur_qk:
            if cur_qk is not None:
                q_spans.append((cur_qk, q_start, DCOL+p["idx"]-1))
            cur_qk, q_start = qk, DCOL+p["idx"]
    if cur_qk is not None:
        q_spans.append((cur_qk, q_start, DCOL+len(periods)-1))

    for qlbl, c1, c2 in q_spans:
        if c1 == c2:
            cell = ws.cell(R_QT, c1)
        else:
            ws.merge_cells(start_row=R_QT,start_column=c1,end_row=R_QT,end_column=c2)
            cell = ws.cell(R_QT, c1)
        cell.value     = qlbl
        cell.fill      = F(C["hdr_qtr"])
        cell.font      = FK(C["txt_white"], bold=False, size=8)
        cell.alignment = AL("center")
        cell.border    = Border(left=_side(), right=_side())

    # ── Month labels ──────────────────────────────────────────────────────────
    for p in periods:
        c = DCOL + p["idx"]
        cell = ws.cell(R_MN, c)
        cell.value     = p["label"]
        cell.fill      = F(C["hdr_month"])
        cell.font      = FK(C["txt_white"], bold=True, size=8)
        cell.alignment = AL("center")
        cell.border    = BORDER_THIN

    # ── Tag row (Actual / BP) ─────────────────────────────────────────────────
    for p in periods:
        c   = DCOL + p["idx"]
        tag = p["tag"]
        bg  = C["hdr_tag_act"] if p["is_actual"] else C["hdr_tag_bp"]
        cell = ws.cell(R_TG, c)
        cell.value     = tag
        cell.fill      = F(bg)
        cell.font      = FK(C["txt_white"], bold=False, size=7, italic=True)
        cell.alignment = AL("center")

    # ── Label / unit cells for header rows ───────────────────────────────────
    for r, lbl in [(R_YR,"Period"), (R_QT,"Quarter"), (R_MN,"Month"), (R_TG,"")]:
        cell = ws.cell(r, LCOL)
        cell.value     = lbl
        cell.fill      = F(C["sec_1"])
        cell.font      = FK(C["txt_grey"], bold=False, size=8, italic=True)
        cell.alignment = AL("right")
        ws.cell(r, UCOL).fill = F(C["sec_1"])

    # ── Annual columns headers ────────────────────────────────────────────────
    if annual_cols:
        for ac in annual_cols:
            col = ac["col"]
            # Year row
            cell = ws.cell(R_YR, col)
            cell.value     = ac["label"]
            cell.fill      = F("152844")
            cell.font      = FK(C["txt_white"], bold=True, size=9)
            cell.alignment = AL("center")
            cell.border    = Border(left=_side_med(),right=_side_med())
            # Quarter
            ws.cell(R_QT, col).fill = F("152844")
            ws.cell(R_QT, col).font = FK(C["txt_grey"],size=8)
            ws.cell(R_QT, col).alignment = AL("center")
            ws.cell(R_QT, col).value = "Annual"
            # Month
            cell = ws.cell(R_MN, col)
            cell.value     = f"FY {ac['year']}"
            cell.fill      = F("1A3A5C")
            cell.font      = FK(C["txt_white"], bold=True, size=9)
            cell.alignment = AL("center")
            cell.border    = BORDER_THIN
            # Tag
            cell = ws.cell(R_TG, col)
            cell.value     = "Full Year"
            cell.fill      = F("1A3A5C")
            cell.font      = FK(C["txt_white"], size=7, italic=True)
            cell.alignment = AL("center")

# ══════════════════════════════════════════════════════════════════════════════
# ROW WRITERS
# ══════════════════════════════════════════════════════════════════════════════

class RowCounter:
    def __init__(self, start): self._r = start
    def next(self, n=1):
        r = self._r; self._r += n; return r
    def skip(self, n=1): self._r += n
    def peek(self): return self._r

def _fill_label_area(ws, row, n_months, annual_cols, bg):
    for c in [LCOL, UCOL]:
        ws.cell(row, c).fill = F(bg)
    for i in range(n_months):
        ws.cell(row, DCOL+i).fill = F(bg)
    for ac in (annual_cols or []):
        ws.cell(row, ac["col"]).fill = F(bg)

def section_row(ws, row, label, n_months, annual_cols, level=1, height=15):
    ws.row_dimensions[row].height = height
    bg   = C["sec_1"] if level==1 else C["sec_2"] if level==2 else C["sec_3"]
    _fill_label_area(ws, row, n_months, annual_cols, bg)
    cell = ws.cell(row, LCOL)
    cell.value     = f"  {label}"
    cell.fill      = F(bg)
    cell.font      = FK(C["txt_white"], bold=True, size=9)
    cell.alignment = AL("left")

def data_row(ws, row, label, monthly_vals, unit="",
             bold=False, bg="white", fmt=FMT["int"],
             height=13, indent=0, input_style=False, link_style=False,
             annual_vals=None, annual_cols=None):
    ws.row_dimensions[row].height = height
    lbg = bg if bg != "white" else C["white"]
    # label
    lc = ws.cell(row, LCOL)
    lc.value     = ("    "*indent) + label
    lc.fill      = F(lbg)
    lc.alignment = AL("left")
    txt_c = (C["txt_input"] if input_style else
             C["txt_link"]  if link_style  else
             C["txt_navy"]  if bold        else C["txt_black"])
    lc.font = FK(txt_c, bold=bold, size=9)
    # unit
    uc = ws.cell(row, UCOL)
    uc.value     = unit
    uc.fill      = F(lbg)
    uc.font      = FK(C["txt_grey"], size=8)
    uc.alignment = AL("center")
    # data
    dat_c = (C["txt_input"] if input_style else
             C["txt_link"]  if link_style  else
             C["txt_navy"]  if bold        else C["txt_black"])
    for i, v in enumerate(monthly_vals):
        c = ws.cell(row, DCOL+i)
        c.value     = v
        c.fill      = F(lbg)
        c.font      = FK(dat_c, bold=bold, size=9)
        c.alignment = AL("right")
        c.number_format = fmt
        c.border    = Border(bottom=_side("hair","E2E8F0"))
    # annual
    if annual_vals and annual_cols:
        for ac, av in zip(annual_cols, annual_vals):
            c = ws.cell(row, ac["col"])
            c.value     = av
            c.fill      = F("EFF6FF")
            c.font      = FK(dat_c, bold=bold, size=9)
            c.alignment = AL("right")
            c.number_format = fmt
            c.border    = BORDER_THIN

def subtotal_row(ws, row, label, monthly_vals, unit="",
                 fmt=FMT["int"], height=14,
                 annual_vals=None, annual_cols=None):
    ws.row_dimensions[row].height = height
    lc = ws.cell(row, LCOL)
    lc.value = "  " + label
    lc.fill  = F(C["subtot"])
    lc.font  = FK(C["txt_navy"], bold=True, size=9)
    lc.alignment = AL("left")
    lc.border    = BORDER_TOP
    ws.cell(row, UCOL).fill  = F(C["subtot"])
    ws.cell(row, UCOL).border= BORDER_TOP
    ws.cell(row, UCOL).value = unit
    ws.cell(row, UCOL).font  = FK(C["txt_grey"], size=8)
    ws.cell(row, UCOL).alignment = AL("center")
    for i, v in enumerate(monthly_vals):
        c = ws.cell(row, DCOL+i)
        c.value = v; c.fill = F(C["subtot"])
        c.font  = FK(C["txt_navy"], bold=True, size=9)
        c.alignment = AL("right"); c.number_format = fmt
        c.border = BORDER_TOP
    if annual_vals and annual_cols:
        for ac, av in zip(annual_cols, annual_vals):
            c = ws.cell(row, ac["col"])
            c.value = av; c.fill = F(C["total"])
            c.font  = FK(C["txt_navy"], bold=True, size=9)
            c.alignment = AL("right"); c.number_format = fmt
            c.border = BORDER_TB

def total_row(ws, row, label, monthly_vals, unit="",
              fmt=FMT["int"], height=15,
              annual_vals=None, annual_cols=None, bg=None):
    ws.row_dimensions[row].height = height
    bg_ = bg or C["total"]
    lc = ws.cell(row, LCOL)
    lc.value = label; lc.fill = F(bg_)
    lc.font  = FK(C["txt_navy"], bold=True, size=9)
    lc.alignment = AL("left"); lc.border = BORDER_TB
    ws.cell(row, UCOL).fill  = F(bg_)
    ws.cell(row, UCOL).border= BORDER_TB
    ws.cell(row, UCOL).value = unit
    ws.cell(row, UCOL).font  = FK(C["txt_grey"], size=8)
    ws.cell(row, UCOL).alignment = AL("center")
    for i, v in enumerate(monthly_vals):
        c = ws.cell(row, DCOL+i)
        c.value = v; c.fill = F(bg_)
        c.font  = FK(C["txt_navy"], bold=True, size=9)
        c.alignment = AL("right"); c.number_format = fmt
        c.border = BORDER_TB
    if annual_vals and annual_cols:
        for ac, av in zip(annual_cols, annual_vals):
            c = ws.cell(row, ac["col"])
            c.value = av; c.fill = F(C["blue_mid"] if bg_ == C["total"] else bg_)
            c.font  = FK(C["txt_navy"], bold=True, size=9)
            c.alignment = AL("right"); c.number_format = fmt
            c.border = BORDER_TB
    return row

def kpi_row(ws, row, label, monthly_vals, unit="",
            fmt=FMT["int"], height=16,
            annual_vals=None, annual_cols=None):
    """Full-width KPI highlight — navy bg, white text."""
    ws.row_dimensions[row].height = height
    for c_idx in [LCOL, UCOL] + [DCOL+i for i in range(len(monthly_vals))] + ([ac["col"] for ac in (annual_cols or [])]):
        ws.cell(row, c_idx).fill = F(C["ebitda_bg"])
    lc = ws.cell(row, LCOL)
    lc.value = label; lc.fill = F(C["ebitda_bg"])
    lc.font  = FK(C["txt_white"], bold=True, size=10)
    lc.alignment = AL("left")
    uc = ws.cell(row, UCOL)
    uc.value = unit; uc.fill = F(C["ebitda_bg"])
    uc.font  = FK("AAAAAA", size=8); uc.alignment = AL("center")
    for i, v in enumerate(monthly_vals):
        c = ws.cell(row, DCOL+i)
        c.value = v; c.fill = F(C["ebitda_bg"])
        c.font  = FK(C["txt_white"], bold=True, size=10)
        c.alignment = AL("right"); c.number_format = fmt
    if annual_vals and annual_cols:
        for ac, av in zip(annual_cols, annual_vals):
            c = ws.cell(row, ac["col"])
            c.value = av; c.fill = F("1A3A5C")
            c.font  = FK(C["txt_white"], bold=True, size=10)
            c.alignment = AL("right"); c.number_format = fmt

def spacer_row(ws, row, n_months, annual_cols, height=5):
    ws.row_dimensions[row].height = height
    for c in [LCOL, UCOL] + [DCOL+i for i in range(n_months)] + [ac["col"] for ac in (annual_cols or [])]:
        ws.cell(row, c).fill = F(C["white"])

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN MODEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════
C["blue_mid"] = "A8C8E8"

def build_model(config, output_path):
    # ── Defensive sanitiser — convert None/str to correct types ──────────────
    def _f(key, default):
        v = config.get(key, default)
        try: return float(v) if v is not None else default
        except: return default
    def _i(key, default):
        v = config.get(key, default)
        try: return int(v) if v is not None else default
        except: return default
    def _s(key, default):
        v = config.get(key, default)
        return str(v) if v is not None else default

    config["company_name"]   = _s("company_name", "Company")
    config["currency"]       = _s("currency", "USD")
    config["business_type"]  = _s("business_type", "industrial")
    config["scenarios"]      = _s("scenarios", "base")
    config["n_years"]        = _i("n_years", 5)
    config["start_year"]     = _i("start_year", 2025)
    config["actuals_months"] = _i("actuals_months", 0)
    config["opening_cash"]   = _f("opening_cash", 5000)
    config["base_revenue"]   = _f("base_revenue", 50000)
    config["revenue_growth"] = _f("revenue_growth", 0.05)
    config["gross_margin"]   = _f("gross_margin", 0.35)
    config["ebitda_margin"]  = _f("ebitda_margin", 0.20)
    config["dso"]            = _f("dso", 60)
    config["dio"]            = _f("dio", 20)
    config["dio_rm"]         = _f("dio_rm", 30)
    config["dpo"]            = _f("dpo", 30)
    config["tax_rate"]       = _f("tax_rate", 0.25)
    config["inflation"]      = _f("inflation", 0.025)
    config["price_per_mt"]   = _f("price_per_mt", 2000)
    config["capacity_mt"]    = _f("capacity_mt", 72000)
    cpx = config.get("capex") or {}
    config["capex"] = {
        "opening_ppe":  float(cpx.get("opening_ppe")  or 85000),
        "maint_capex":  float(cpx.get("maint_capex")  or 2000),
        "expan_capex":  float(cpx.get("expan_capex")  or 5000),
        "useful_life":  int(cpx.get("useful_life")    or 20),
    }
    # Debt: ensure total_debt has a fallback from tranches
    dbt = config.get("debt") or {}
    if not dbt.get("total_debt"):
        tranches = dbt.get("tranches") or []
        dbt["total_debt"] = sum(float(t.get("amount",0) or 0) for t in tranches if isinstance(t,dict))
    if not dbt.get("interest_rate"):
        dbt["interest_rate"] = 0.065
    config["debt"] = dbt
    # ─────────────────────────────────────────────────────────────────────────

    wb = Workbook()
    wb.remove(wb.active)

    company    = config.get("company_name", "Company")
    ccy        = config.get("currency", "USD")
    unit_lbl   = f"{ccy}k"
    n_years    = config.get("n_years", 5)
    start_year = config.get("start_year", 2025)
    n_m        = n_years * 12
    act_m      = config.get("actuals_months", 0)

    periods    = gen_periods(start_year, n_m, act_m)
    ag         = annual_groups(periods)
    years      = sorted(ag.keys())
    n_ac       = len(years)

    ACOL  = DCOL + n_m + 2   # annual columns start
    annual_cols = [{"year": y, "label": f"FY {y}", "col": ACOL+i} for i,y in enumerate(years)]
    LAST_COL = ACOL + n_ac

    def av(vals, avg=False):
        return [ann_avg(vals,ag,y) if avg else ann_sum(vals,ag,y) for y in years]

    # ── Pull assumptions ──────────────────────────────────────────────────────
    base_rev    = config.get("base_revenue", 50000)
    rev_growth  = config.get("revenue_growth", 0.05)
    gm_pct      = config.get("gross_margin", 0.35)
    ebitda_pct  = config.get("ebitda_margin", 0.20)
    dso         = config.get("dso", 60)
    dio_fg      = config.get("dio", 20)
    dio_rm      = config.get("dio_rm", 30)
    dpo         = config.get("dpo", 30)
    tax_rate    = config.get("tax_rate", 0.25)
    inflation   = config.get("inflation", 0.025)
    debt_cfg    = config.get("debt", {})
    total_debt  = debt_cfg.get("total_debt", 0)
    int_rate    = debt_cfg.get("interest_rate", 0.065)
    cpx_cfg     = config.get("capex", {})
    maint_yr    = cpx_cfg.get("maint_capex", 2000)
    expan_yr    = cpx_cfg.get("expan_capex", 5000)
    open_ppe    = cpx_cfg.get("opening_ppe", 85000)
    ul_years    = cpx_cfg.get("useful_life", 20)
    ul_months   = ul_years * 12
    biz         = config.get("business_type", "industrial")
    open_cash   = config.get("opening_cash", 5000)
    price_mt    = config.get("price_per_mt", 2000)
    capacity_mt = config.get("capacity_mt", 6000)

    # ── Monthly series ────────────────────────────────────────────────────────
    def monthly_rev_series():
        out = []
        for p in periods:
            yr_off = p["year"] - start_year
            out.append(base_rev * ((1+rev_growth)**yr_off) / 12)
        return out

    net_rev   = monthly_rev_series()
    gross_rev = [r/0.96 for r in net_rev]
    freight   = [r - nr for r,nr in zip(gross_rev,net_rev)]

    # Direct materials by type
    dm_total_pct = 1 - gm_pct - 0.07
    dm           = [-r*dm_total_pct for r in net_rev]
    dm_sbc       = [d*0.52 for d in dm]
    dm_silica    = [d*0.20 for d in dm]
    dm_chem      = [d*0.13 for d in dm]
    dm_pkg       = [d*0.10 for d in dm]
    dm_other     = [d*0.05 for d in dm]
    variance     = [0.0]*n_m

    util_costs   = [-r*0.08 for r in net_rev]
    pack_costs   = [-r*0.02 for r in net_rev]
    var_costs    = [d+u+p for d,u,p in zip(dm,util_costs,pack_costs)]
    contrib_m    = [r+vc for r,vc in zip(net_rev,var_costs)]

    sal_prod     = [-r*0.05 for r in net_rev]
    sal_mnt      = [-r*0.015 for r in net_rev]
    sal_mgmt     = [-r*0.02 for r in net_rev]
    sal_other    = [-r*0.005 for r in net_rev]
    sal_total    = [sp+sm+smg+so for sp,sm,smg,so in zip(sal_prod,sal_mnt,sal_mgmt,sal_other)]
    spares       = [-r*0.012 for r in net_rev]
    maint_oh     = [-r*0.014 for r in net_rev]
    qc           = [-r*0.006 for r in net_rev]
    hsse         = [-r*0.004 for r in net_rev]
    mfg_oh_oth   = [-r*0.004 for r in net_rev]
    mfg_oh_tot   = [sp+mo+q+h+oo for sp,mo,q,h,oo in zip(spares,maint_oh,qc,hsse,mfg_oh_oth)]
    sell_dist    = [-r*0.02 for r in net_rev]
    admin        = [-r*0.018 for r in net_rev]
    corp_oh      = [-r*0.01 for r in net_rev]
    mgmt_fees    = [-r*0.005 for r in net_rev]
    fixed_total  = [st+mo+sd+ad+co for st,mo,sd,ad,co in zip(sal_total,mfg_oh_tot,sell_dist,admin,corp_oh)]
    gross_profit = [r*gm_pct for r in net_rev]
    adj_ebitda   = [r*ebitda_pct for r in net_rev]
    restruct     = [0.0]*n_m
    ebitda_rep   = [ae+mf+re for ae,mf,re in zip(adj_ebitda,mgmt_fees,restruct)]

    # Capex & depreciation
    maint_m  = [maint_yr/12]*n_m
    expan_m  = [expan_yr/12]*n_m
    cpx_tot  = [a+b for a,b in zip(maint_m,expan_m)]
    ppe_open_m, ppe_close_m, dep_m, accum_dep = [], [], [], []
    ppe = open_ppe; accum = 0.0
    for i in range(n_m):
        ppe_open_m.append(ppe)
        dep = ppe / ul_months
        accum += dep; dep_m.append(dep); accum_dep.append(accum)
        ppe += cpx_tot[i]; ppe_close_m.append(ppe)
    nbv_m = [pc - ad for pc,ad in zip(ppe_close_m,accum_dep)]
    ebit  = [e - d for e,d in zip(ebitda_rep,dep_m)]

    # Debt
    # ═══ DEBT ENGINE INTEGRATION ═══════════════════════════════
    _use_engine = False
    if _DEBT_ENGINE_AVAILABLE and debt_cfg.get('tranches') and debt_cfg['tranches'] and isinstance(debt_cfg['tranches'][0], dict):
        try:
            _br = {**_BASE_RATES, **config.get('base_rates',{})}
            _tranches_obj = parse_tranches(debt_cfg['tranches'])
            _engine = DebtEngine(_tranches_obj, n_m, _br)
            _etot = _engine.totals
            debt_open  = _etot['opening']
            debt_close = _etot['closing']
            debt_repay = _etot['repayment']
            interest   = _etot['interest_total']
            _use_engine = True
        except Exception as _e:
            print(f'Debt engine error: {_e}')
    # ═══════════════════════════════════════════════════════════════
    debt_open, debt_close, debt_repay, interest = [], [], [], []
    od = total_debt
    for i in range(n_m):
        debt_open.append(od)
        rep = od / max(n_m-i, 1) * 0.5 if od > 0 else 0  # partial amort
        rep = min(rep, od)
        intr = od * int_rate / 12
        debt_repay.append(rep); interest.append(intr)
        od = max(od - rep, 0); debt_close.append(od)

    pbt        = [e - i for e,i in zip(ebit,interest)]
    tax_charge = [max(p*tax_rate,0) for p in pbt]
    net_income = [p-t for p,t in zip(pbt,tax_charge)]

    # NWC
    ar       = [r*dso/365 for r in net_rev]
    inv_fg   = [abs(d)*dio_fg/365 for d in dm]
    inv_rm   = [abs(d)*dio_rm/365 for d in dm]
    inv_tot  = [fg+rm for fg,rm in zip(inv_fg,inv_rm)]
    ap       = [abs(d)*dpo/365 for d in dm]
    oth_ca   = [r*0.01 for r in net_rev]
    oth_cl   = [r*0.01 for r in net_rev]
    emp_prov = [abs(s)*0.08 for s in sal_total]
    nwc      = [a+it+oca-p-ocl-ep for a,it,oca,p,ocl,ep in zip(ar,inv_tot,oth_ca,ap,oth_cl,emp_prov)]
    chg_nwc  = [nwc[0]] + [nwc[i]-nwc[i-1] for i in range(1,n_m)]

    # Cash flow
    op_cf    = [ae - cn - tc for ae,cn,tc in zip(adj_ebitda,chg_nwc,tax_charge)]
    inv_cf   = [-c for c in cpx_tot]
    int_cash = [-i*0.80 for i in interest]
    fin_cf   = [ic-r for ic,r in zip(int_cash,debt_repay)]
    net_cf   = [o+i+f for o,i,f in zip(op_cf,inv_cf,fin_cf)]
    cash_cl  = []; prev = open_cash
    for nc in net_cf: prev += nc; cash_cl.append(prev)
    cash_op  = [open_cash] + cash_cl[:-1]
    fcf      = [o+i for o,i in zip(op_cf,inv_cf)]

    # Balance sheet derived
    net_debt = [dc-cc for dc,cc in zip(debt_close,cash_cl)]
    accum_re = []; c_re=0.0
    for ni in net_income: c_re+=ni; accum_re.append(c_re)
    ta_vals  = [ppe+a+it+oca+cc for ppe,a,it,oca,cc in zip(nbv_m,ar,inv_tot,oth_ca,cash_cl)]
    tl_vals  = [dc+p+ep+ocl for dc,p,ep,ocl in zip(debt_close,ap,emp_prov,oth_cl)]
    eq_vals  = accum_re
    tle_vals = [tl+eq for tl,eq in zip(tl_vals,eq_vals)]

    # ══════════════════════════════════════════════════════════════════════════
    # 1.  ASSUMPTIONS
    # ══════════════════════════════════════════════════════════════════════════
    ws = setup_ws(wb, "ASSUMPTIONS", "1F3D7A")
    set_col_widths(ws, n_m, annual_cols)
    write_title(ws, 1, company, "Assumptions & Drivers", unit_lbl, LAST_COL)
    write_period_headers(ws, 3, periods, annual_cols)
    R = RowCounter(3 + HDR_ROWS + 1)

    def adr(label, vals, unit="", input_=True, fmt=FMT["int"], indent=0):
        r = R.next()
        data_row(ws, r, label, vals, unit=unit, fmt=fmt, indent=indent,
                 input_style=input_, bg=C["input_bg"] if input_ else C["white"],
                 annual_vals=av(vals, avg=(fmt in [FMT["pct1"],FMT["mult"]])),
                 annual_cols=annual_cols)
        return r

    section_row(ws, R.next(), "REVENUE DRIVERS", n_m, annual_cols)
    adr("Base Revenue (annual, "+unit_lbl+")", [base_rev]*n_m, unit=unit_lbl)
    adr("Revenue Growth Rate (annual)", [rev_growth]*n_m, unit="%", fmt=FMT["pct1"])
    adr("Gross Margin %", [gm_pct]*n_m, unit="%", fmt=FMT["pct1"])
    adr("Adj. EBITDA Margin %", [ebitda_pct]*n_m, unit="%", fmt=FMT["pct1"])
    spacer_row(ws, R.next(), n_m, annual_cols)

    section_row(ws, R.next(), "WORKING CAPITAL DRIVERS", n_m, annual_cols, level=2)
    adr("Days Sales Outstanding  (DSO)", [dso]*n_m, unit="days", fmt=FMT["int"])
    adr("Days Inventory Outstanding  (DIO) — Finished Goods", [dio_fg]*n_m, unit="days")
    adr("Days Inventory Outstanding  (DIO) — Raw Materials", [dio_rm]*n_m, unit="days")
    adr("Days Payables Outstanding  (DPO)", [dpo]*n_m, unit="days")
    spacer_row(ws, R.next(), n_m, annual_cols)

    section_row(ws, R.next(), "MACRO & TAX", n_m, annual_cols, level=2)
    adr("Tax Rate", [tax_rate]*n_m, unit="%", fmt=FMT["pct1"])
    adr("Inflation (annual)", [inflation]*n_m, unit="%", fmt=FMT["pct1"])
    spacer_row(ws, R.next(), n_m, annual_cols)

    section_row(ws, R.next(), "CAPEX & DEBT", n_m, annual_cols, level=2)
    adr("Maintenance Capex (monthly)", maint_m, unit=unit_lbl)
    adr("Expansion Capex (monthly)", expan_m, unit=unit_lbl)
    adr("Opening PP&E (gross)", [open_ppe]+[None]*(n_m-1), unit=unit_lbl)
    adr("Useful Life (years)", [ul_years]*n_m, unit="yrs")
    adr("Total Gross Debt (opening)", [total_debt]+[None]*(n_m-1), unit=unit_lbl)
    adr("Avg. Interest Rate (annual)", [int_rate]*n_m, unit="%", fmt=FMT["pct1"])

    # ══════════════════════════════════════════════════════════════════════════
    # 2.  P&L
    # ══════════════════════════════════════════════════════════════════════════
    ws_pl = setup_ws(wb, "P&L", "0D1B3E")
    set_col_widths(ws_pl, n_m, annual_cols)
    write_title(ws_pl, 1, company, "Income Statement", unit_lbl, LAST_COL)
    write_period_headers(ws_pl, 3, periods, annual_cols)
    R = RowCounter(3 + HDR_ROWS + 1)

    def pr(label, vals, bold=False, bg="white", fmt=FMT["int"], indent=0,
           input_=False, link_=False, unit=""):
        r = R.next()
        data_row(ws_pl, r, label, vals, unit=unit, bold=bold, bg=bg, fmt=fmt, indent=indent,
                 input_style=input_, link_style=link_,
                 annual_vals=av(vals), annual_cols=annual_cols)
        return r
    def pst(label, vals, fmt=FMT["int"], unit=""):
        r = R.next()
        subtotal_row(ws_pl, r, label, vals, unit=unit, fmt=fmt,
                     annual_vals=av(vals), annual_cols=annual_cols)
        return r
    def ptt(label, vals, fmt=FMT["int"], unit="", bg=None):
        r = R.next()
        total_row(ws_pl, r, label, vals, unit=unit, fmt=fmt, bg=bg,
                  annual_vals=av(vals), annual_cols=annual_cols)
        return r
    def pkpi(label, vals, fmt=FMT["int"], unit=""):
        r = R.next()
        kpi_row(ws_pl, r, label, vals, unit=unit, fmt=fmt,
                annual_vals=av(vals), annual_cols=annual_cols)
        return r
    def psp(): spacer_row(ws_pl, R.next(), n_m, annual_cols)

    # Revenue
    section_row(ws_pl, R.next(), "REVENUE", n_m, annual_cols)
    pr("Gross Revenue", gross_rev, indent=1, unit=unit_lbl)
    pr("Less: Freight & Forwarding", [-f for f in freight], indent=2, unit=unit_lbl)
    pr("Less: Other Deductions / Rebates", [0.0]*n_m, indent=2, unit=unit_lbl, input_=True, bg=C["input_bg"])
    nr_r = ptt("Net Realisation", net_rev, unit=unit_lbl)
    psp()

    # Variable costs
    section_row(ws_pl, R.next(), "VARIABLE COSTS", n_m, annual_cols, level=2)
    pr("  SBC (Styrene-Butadiene Compound)", dm_sbc, indent=2, unit=unit_lbl)
    pr("  Silica & Fillers", dm_silica, indent=2, unit=unit_lbl)
    pr("  Chemicals & Additives", dm_chem, indent=2, unit=unit_lbl)
    pr("  Packaging Materials", dm_pkg, indent=2, unit=unit_lbl)
    pr("  Other Direct Materials", dm_other, indent=2, unit=unit_lbl)
    pr("  Variances & Movements in Stock", variance, indent=2, unit=unit_lbl, input_=True, bg=C["input_bg"])
    dm_r = pst("Total Direct Materials", dm, unit=unit_lbl)
    psp()
    pr("  Utilities (Gas, Electricity, Water)", util_costs, indent=2, unit=unit_lbl)
    pr("  Packing Costs", pack_costs, indent=2, unit=unit_lbl)
    vc_r = pst("Total Variable Costs", var_costs, unit=unit_lbl)
    psp()
    cm_r = ptt("Contribution Margin", contrib_m, unit=unit_lbl, bg=C["subtot"])
    pr("Contribution Margin %",
       [c/r if r else 0 for c,r in zip(contrib_m,net_rev)],
       fmt=FMT["pct1"], unit="%", bg=C["grey_row"])
    psp()

    # Fixed costs
    section_row(ws_pl, R.next(), "FIXED COSTS", n_m, annual_cols, level=2)
    pr("  Production — Salary & Wages", sal_prod, indent=2, unit=unit_lbl)
    pr("  Maintenance — Salary & Wages", sal_mnt, indent=2, unit=unit_lbl)
    pr("  Management — Salary & Wages", sal_mgmt, indent=2, unit=unit_lbl)
    pr("  Other Staff Costs (Benefits, Bonus, Pension)", sal_other, indent=2, unit=unit_lbl)
    sw_r = pst("Total Salary & Wages", sal_total, unit=unit_lbl)
    psp()
    pr("  Spares & Consumables", spares, indent=2, unit=unit_lbl)
    pr("  Maintenance & Repairs", maint_oh, indent=2, unit=unit_lbl)
    pr("  Quality Control & Laboratory", qc, indent=2, unit=unit_lbl)
    pr("  HSSE & Regulatory Compliance", hsse, indent=2, unit=unit_lbl)
    pr("  Other Manufacturing Overhead", mfg_oh_oth, indent=2, unit=unit_lbl)
    mfgoh_r = pst("Total Manufacturing Overhead", mfg_oh_tot, unit=unit_lbl)
    psp()
    pr("  Selling & Distribution", sell_dist, indent=2, unit=unit_lbl)
    pr("  General & Administrative", admin, indent=2, unit=unit_lbl)
    pr("  Corporate Overhead Allocation", corp_oh, indent=2, unit=unit_lbl)
    fc_r = pst("Total Fixed Costs", fixed_total, unit=unit_lbl)
    psp()

    # EBITDA
    adj_ebitda_r = pkpi(f"ADJUSTED EBITDA  ({unit_lbl})", adj_ebitda, unit=unit_lbl)
    pr("Adj. EBITDA Margin %",
       [e/r if r else 0 for e,r in zip(adj_ebitda,net_rev)],
       fmt=FMT["pct1"], unit="%", bg=C["grey_row"])
    psp()
    pr("Management & Advisory Fees", mgmt_fees, indent=1, unit=unit_lbl)
    pr("Restructuring & Exceptional Costs", restruct, indent=1, unit=unit_lbl, input_=True, bg=C["input_bg"])
    ebitda_r = ptt("EBITDA (Reported)", ebitda_rep, unit=unit_lbl)
    psp()

    # Below EBITDA
    section_row(ws_pl, R.next(), "BELOW EBITDA", n_m, annual_cols, level=2)
    dep_r = pr("Depreciation & Amortisation", [-d for d in dep_m], indent=1, unit=unit_lbl, link_=True)
    pr("Exchange Gains / (Losses)", [0.0]*n_m, indent=1, unit=unit_lbl, input_=True, bg=C["input_bg"])
    pr("Other Income / (Expenses)", [0.0]*n_m, indent=1, unit=unit_lbl, input_=True, bg=C["input_bg"])
    pr("Gain / (Loss) on Disposal of Assets", [0.0]*n_m, indent=1, unit=unit_lbl, input_=True, bg=C["input_bg"])
    ebit_r = ptt("EBIT", ebit, unit=unit_lbl)
    psp()

    section_row(ws_pl, R.next(), "FINANCE COSTS", n_m, annual_cols, level=2)
    pr("Interest — Cash Pay (80%)", [-i*0.80 for i in interest], indent=1, unit=unit_lbl, link_=True)
    pr("Interest — PIK / Accrued (20%)", [-i*0.20 for i in interest], indent=1, unit=unit_lbl, link_=True)
    pr("Commitment Fee (undrawn)", [0.0]*n_m, indent=1, unit=unit_lbl, input_=True, bg=C["input_bg"])
    pr("Amortisation of Financing Fees", [0.0]*n_m, indent=1, unit=unit_lbl, input_=True, bg=C["input_bg"])
    ni_r = pst("Net Finance Costs", [-i for i in interest], unit=unit_lbl)
    psp()

    pbt_r = ptt("PROFIT BEFORE TAX  (PBT)", pbt, unit=unit_lbl)
    psp()
    section_row(ws_pl, R.next(), "TAXATION", n_m, annual_cols, level=2)
    pr("Current Tax Charge", [-t for t in tax_charge], indent=1, unit=unit_lbl)
    pr("Deferred Tax", [0.0]*n_m, indent=1, unit=unit_lbl, input_=True, bg=C["input_bg"])
    pst("Total Tax Charge", [-t for t in tax_charge], unit=unit_lbl)
    psp()
    ni_row = pkpi(f"NET INCOME  ({unit_lbl})", net_income, unit=unit_lbl)
    pr("Net Margin %",
       [n/r if r else 0 for n,r in zip(net_income,net_rev)],
       fmt=FMT["pct1"], unit="%", bg=C["grey_row"])

    # ══════════════════════════════════════════════════════════════════════════
    # 3.  NWC
    # ══════════════════════════════════════════════════════════════════════════
    ws_nwc = setup_ws(wb, "NWC", "7B3F00")
    set_col_widths(ws_nwc, n_m, annual_cols)
    write_title(ws_nwc, 1, company, "Net Working Capital", unit_lbl, LAST_COL)
    write_period_headers(ws_nwc, 3, periods, annual_cols)
    R = RowCounter(3+HDR_ROWS+1)

    def nwr(label, vals, bold=False, bg="white", fmt=FMT["int"], indent=0, input_=False, unit=""):
        r = R.next()
        data_row(ws_nwc, r, label, vals, unit=unit, bold=bold, bg=bg, fmt=fmt, indent=indent,
                 input_style=input_, annual_vals=av(vals), annual_cols=annual_cols)
        return r
    def nwst(label, vals, fmt=FMT["int"], unit=""):
        r = R.next(); subtotal_row(ws_nwc, r, label, vals, unit=unit, fmt=fmt, annual_vals=av(vals), annual_cols=annual_cols); return r
    def nwtt(label, vals, bg=None, unit=""):
        r = R.next(); total_row(ws_nwc, r, label, vals, bg=bg, unit=unit, annual_vals=av(vals), annual_cols=annual_cols); return r
    def nwsp(): spacer_row(ws_nwc, R.next(), n_m, annual_cols)

    section_row(ws_nwc, R.next(), "NWC DRIVERS (from Assumptions)", n_m, annual_cols)
    nwr("Days Sales Outstanding (DSO)", [dso]*n_m, unit="days", input_=True, bg=C["input_bg"])
    nwr("Days Inventory Outstanding (DIO) — Finished Goods", [dio_fg]*n_m, unit="days", input_=True, bg=C["input_bg"])
    nwr("Days Inventory Outstanding (DIO) — Raw Materials", [dio_rm]*n_m, unit="days", input_=True, bg=C["input_bg"])
    nwr("Days Payables Outstanding (DPO)", [dpo]*n_m, unit="days", input_=True, bg=C["input_bg"])
    nwsp()

    section_row(ws_nwc, R.next(), "CURRENT ASSETS", n_m, annual_cols, level=2)
    nwr("Trade Receivables — Gross", ar, indent=1, unit=unit_lbl)
    nwr("  Advances from Customers", [0.0]*n_m, indent=2, unit=unit_lbl, input_=True, bg=C["input_bg"])
    nwr("Inventories — Finished Goods", inv_fg, indent=1, unit=unit_lbl)
    nwr("Inventories — Raw Materials", inv_rm, indent=1, unit=unit_lbl)
    nwr("Inventories — Semi-Finished Goods", [0.0]*n_m, indent=2, unit=unit_lbl, input_=True, bg=C["input_bg"])
    nwr("Inventories — Goods in Transit", [0.0]*n_m, indent=2, unit=unit_lbl, input_=True, bg=C["input_bg"])
    nwr("Other Receivables", oth_ca, indent=1, unit=unit_lbl)
    nwr("Prepayments & Accrued Income", [r*0.005 for r in net_rev], indent=2, unit=unit_lbl)
    ca_tot = [a+it+oca for a,it,oca in zip(ar,inv_tot,oth_ca)]
    nwst("Total Current Assets", ca_tot, unit=unit_lbl)
    nwsp()

    section_row(ws_nwc, R.next(), "CURRENT LIABILITIES", n_m, annual_cols, level=2)
    nwr("Trade Payables — Gross", ap, indent=1, unit=unit_lbl)
    nwr("  Advances to Suppliers", [0.0]*n_m, indent=2, unit=unit_lbl, input_=True, bg=C["input_bg"])
    nwr("Employee & Salary Provisions", emp_prov, indent=1, unit=unit_lbl)
    nwr("Other Payables & Accruals", oth_cl, indent=1, unit=unit_lbl)
    cl_tot = [p+ep+ocl for p,ep,ocl in zip(ap,emp_prov,oth_cl)]
    nwst("Total Current Liabilities", cl_tot, unit=unit_lbl)
    nwsp()

    nwc_r = R.next()
    kpi_row(ws_nwc, nwc_r, f"NET WORKING CAPITAL  ({unit_lbl})", nwc, unit=unit_lbl,
            annual_vals=av(nwc), annual_cols=annual_cols)
    nwsp()
    section_row(ws_nwc, R.next(), "CASH FLOW IMPACT", n_m, annual_cols, level=2)
    nwr("Change in NWC (increase = outflow)", chg_nwc, unit=unit_lbl)
    chg_ar  = [ar[0]] + [ar[i]-ar[i-1] for i in range(1,n_m)]
    chg_inv = [inv_tot[0]] + [inv_tot[i]-inv_tot[i-1] for i in range(1,n_m)]
    chg_ap  = [ap[0]] + [ap[i]-ap[i-1] for i in range(1,n_m)]
    nwr("  of which: Change in Receivables", chg_ar, indent=2, unit=unit_lbl)
    nwr("  of which: Change in Inventories", chg_inv, indent=2, unit=unit_lbl)
    nwr("  of which: Change in Payables", [-c for c in chg_ap], indent=2, unit=unit_lbl)

    # ══════════════════════════════════════════════════════════════════════════
    # 4.  CAPEX
    # ══════════════════════════════════════════════════════════════════════════
    ws_cpx = setup_ws(wb, "CAPEX", "2C4A6E")
    set_col_widths(ws_cpx, n_m, annual_cols)
    write_title(ws_cpx, 1, company, "CAPEX & Depreciation Schedule", unit_lbl, LAST_COL)
    write_period_headers(ws_cpx, 3, periods, annual_cols)
    R = RowCounter(3+HDR_ROWS+1)

    def cpr(label, vals, bold=False, bg="white", fmt=FMT["int"], indent=0, input_=False):
        r = R.next(); data_row(ws_cpx, r, label, vals, unit=unit_lbl, bold=bold, bg=bg, fmt=fmt, indent=indent, input_style=input_, annual_vals=av(vals), annual_cols=annual_cols); return r
    def cpst(label, vals): r = R.next(); subtotal_row(ws_cpx, r, label, vals, unit=unit_lbl, annual_vals=av(vals), annual_cols=annual_cols); return r
    def cpkpi(label, vals): r = R.next(); kpi_row(ws_cpx, r, label, vals, unit=unit_lbl, annual_vals=av(vals), annual_cols=annual_cols); return r
    def cpsp(): spacer_row(ws_cpx, R.next(), n_m, annual_cols)

    section_row(ws_cpx, R.next(), "CAPITAL EXPENDITURE", n_m, annual_cols)
    cpr("Maintenance Capex", maint_m, indent=1, input_=True, bg=C["input_bg"])
    cpr("Expansion / Growth Capex", expan_m, indent=1, input_=True, bg=C["input_bg"])
    cpr("IT & Digital Infrastructure", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    cpr("Other / Strategic Capex", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    cpx_r = cpst("Total Capex", cpx_tot)
    cpx_r_vals = cpx_tot
    cpsp()

    section_row(ws_cpx, R.next(), "PP&E ROLL-FORWARD (GROSS)", n_m, annual_cols, level=2)
    cpr("Opening PP&E (gross)", ppe_open_m, indent=1, bg=C["input_bg"], input_=True)
    cpr("Additions in Period", cpx_tot, indent=1)
    cpr("Disposals", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    cpst("Closing PP&E (gross)", ppe_close_m)  # placeholder
    cpsp()

    section_row(ws_cpx, R.next(), "ACCUMULATED DEPRECIATION", n_m, annual_cols, level=2)
    cpr("Annual Depreciation Charge", dep_m, indent=1)
    cpr("Opening Accumulated Depreciation", [0.0]+accum_dep[:-1], indent=1, bg=C["input_bg"], input_=True)
    cpst("Closing Accumulated Depreciation", accum_dep)
    cpsp()

    nbv_r = cpkpi(f"NET BOOK VALUE — PP&E  ({unit_lbl})", nbv_m)

    # ══════════════════════════════════════════════════════════════════════════
    # 5.  DEBT SCHEDULE
    # ══════════════════════════════════════════════════════════════════════════
    # Use enhanced debt sheet if engine available
    if _DEBT_ENGINE_AVAILABLE and _use_engine:
        config["_ebitda_m"] = adj_ebitda
        config["_cash_m"]   = cash_cl
        build_debt_sheet(wb, _engine, config, periods, annual_cols, LCOL, UCOL, DCOL, n_m)
    ws_ds = setup_ws(wb, "DEBT DETAIL", "8B2020")
    set_col_widths(ws_ds, n_m, annual_cols)
    write_title(ws_ds, 1, company, "Debt Schedule & Covenant Tracking", unit_lbl, LAST_COL)
    write_period_headers(ws_ds, 3, periods, annual_cols)
    R = RowCounter(3+HDR_ROWS+1)

    def dsr(label, vals, bold=False, bg="white", fmt=FMT["int"], indent=0, input_=False):
        r = R.next(); data_row(ws_ds, r, label, vals, unit=unit_lbl, bold=bold, bg=bg, fmt=fmt, indent=indent, input_style=input_, annual_vals=av(vals), annual_cols=annual_cols); return r
    def dsst(label, vals, fmt=FMT["int"]): r = R.next(); subtotal_row(ws_ds, r, label, vals, unit=unit_lbl, fmt=fmt, annual_vals=av(vals), annual_cols=annual_cols); return r
    def dstt(label, vals, fmt=FMT["int"]): r = R.next(); total_row(ws_ds, r, label, vals, unit=unit_lbl, fmt=fmt, annual_vals=av(vals), annual_cols=annual_cols); return r
    def dskpi(label, vals, fmt=FMT["int"]): r = R.next(); kpi_row(ws_ds, r, label, vals, unit=unit_lbl, fmt=fmt, annual_vals=av(vals,avg=(fmt!=FMT["int"])), annual_cols=annual_cols); return r
    def dssp(): spacer_row(ws_ds, R.next(), n_m, annual_cols)

    active = debt_cfg.get("tranches", ["senior_a"])
    tranche_names = {
        "senior_a": "Senior Term Loan A  —  Amortising",
        "senior_b": "Senior Term Loan B  —  Bullet",
        "revolving": "Revolving Credit Facility",
        "mezz":     "Mezzanine / Second Lien",
        "murabaha": "Murabaha Facility  (Islamic)",
        "leasing":  "Finance Lease / IFRS 16",
    }
    tranche_debt = total_debt / max(len(active), 1)

    for t in active:
        if isinstance(t, dict):
            t_key = t.get('key', t.get('type', 'tranche'))
            t_name = t.get('name', t.get('type', tranche_names.get(t_key, str(t_key))))
            t_amount = float(t.get('amount', tranche_debt) or tranche_debt)
            t_rate = float(t.get('rate', int_rate * 100) or int_rate * 100) / 100
            t_pik = float(t.get('pik_rate', 0) or 0) / 100
            t_tenor_m = int(float(t.get('tenor', 7)) * 12)
            t_amort = t.get('amortization', 'linear')
            t_freq = t.get('frequency', 'quarterly')
            t_drawn = float(t.get('drawn_pct', 1.0) or 1.0)
        else:
            t_key = t
            t_name = tranche_names.get(t_key, str(t_key))
            t_amount = tranche_debt
            t_rate = int_rate
            t_pik = 0.0
            t_tenor_m = n_m
            t_amort = 'linear'
            t_freq = 'quarterly'
            t_drawn = 1.0
        section_row(ws_ds, R.next(), t_name, n_m, annual_cols, level=2)
        # Per-tranche amortization
        if t_amort == 'bullet':
            t_op  = [t_amount * t_drawn] * n_m
            t_rep = [0.0] * (n_m - 1) + [t_amount * t_drawn]
        elif t_amort == 'linear':
            t_op  = [max(t_amount * t_drawn - t_amount * t_drawn / max(t_tenor_m, n_m) * i, 0) for i in range(n_m)]
            t_rep = [t_amount * t_drawn / max(t_tenor_m, n_m)] * n_m
        else:
            t_op  = [max(t_amount * t_drawn - t_amount * t_drawn / n_m * i, 0) for i in range(n_m)]
            t_rep = [t_amount * t_drawn / n_m] * n_m
        t_int_cash = [op * t_rate / 12 for op in t_op]
        t_int_pik  = [op * t_pik  / 12 for op in t_op]
        t_cl  = [max(op - rep, 0) for op, rep in zip(t_op, t_rep)]
        dsr("Opening Balance", t_op, indent=1, bg=C["input_bg"], input_=True)
        dsr("Repayments  (scheduled)", [-r for r in t_rep], indent=2)
        dsr("Interest — Cash Pay", [-i for i in t_int_cash], indent=2)
        dsr("Interest — PIK / Accrued", [-i for i in t_int_pik], indent=2)
        dsst("Closing Balance", t_cl)
        dssp()

    section_row(ws_ds, R.next(), "CONSOLIDATED DEBT SUMMARY", n_m, annual_cols)
    dsr("Opening Gross Debt", debt_open, indent=1, bg=C["input_bg"], input_=True)
    dsr("Total Repayments", [-r for r in debt_repay], indent=1)
    dsr("PIK Additions", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    dsst("Closing Gross Debt", debt_close)
    dssp()

    section_row(ws_ds, R.next(), "INTEREST & FINANCING COSTS", n_m, annual_cols, level=2)
    dsr("Interest — Cash Pay (80%)", [-i*0.80 for i in interest], indent=1)
    dsr("Interest — PIK / Accrued (20%)", [-i*0.20 for i in interest], indent=1)
    dsr("Commitment Fee (undrawn RCF)", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    dsr("Financing Fee Amortisation", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    dsst("Total Finance Costs", [-i for i in interest])
    dssp()

    section_row(ws_ds, R.next(), "COVENANT TRACKING", n_m, annual_cols, level=2)
    ltm_ae = [sum(adj_ebitda[max(i-11,0):i+1])*12/max(min(i+1,12),1) for i in range(n_m)]
    lev_m  = [nd/max(ltm,0.001) for nd,ltm in zip(net_debt,ltm_ae)]
    icr_m  = [ae*12/max(intr*12,0.001) for ae,intr in zip(adj_ebitda,interest)]
    dsr("Net Debt", net_debt, indent=1)
    dsr("LTM Adj. EBITDA (annualised)", ltm_ae, indent=1)
    dsr("Net Leverage (x)",  lev_m,  indent=1, fmt=FMT["mult"])
    dsr("Interest Cover (x)", icr_m, indent=1, fmt=FMT["mult"])
    dsr("Max Leverage Covenant", [6.0]*n_m, indent=2, input_=True, bg=C["input_bg"], fmt=FMT["mult"])
    dsr("Min ICR Covenant",      [2.5]*n_m, indent=2, input_=True, bg=C["input_bg"], fmt=FMT["mult"])

    # ══════════════════════════════════════════════════════════════════════════
    # 6.  BALANCE SHEET
    # ══════════════════════════════════════════════════════════════════════════
    ws_bs = setup_ws(wb, "BALANCE SHEET", "1A5C2A")
    set_col_widths(ws_bs, n_m, annual_cols)
    write_title(ws_bs, 1, company, "Balance Sheet", unit_lbl, LAST_COL)
    write_period_headers(ws_bs, 3, periods, annual_cols)
    R = RowCounter(3+HDR_ROWS+1)

    def bsr(label, vals, bold=False, bg="white", indent=0, link_=False):
        r = R.next(); data_row(ws_bs, r, label, vals, unit=unit_lbl, bold=bold, bg=bg, indent=indent, link_style=link_, annual_vals=av(vals), annual_cols=annual_cols); return r
    def bsst(label, vals): r = R.next(); subtotal_row(ws_bs, r, label, vals, unit=unit_lbl, annual_vals=av(vals), annual_cols=annual_cols); return r
    def bstt(label, vals, bg=None): r = R.next(); total_row(ws_bs, r, label, vals, unit=unit_lbl, bg=bg, annual_vals=av(vals), annual_cols=annual_cols); return r
    def bskpi(label, vals): r = R.next(); kpi_row(ws_bs, r, label, vals, unit=unit_lbl, annual_vals=av(vals), annual_cols=annual_cols); return r
    def bssp(): spacer_row(ws_bs, R.next(), n_m, annual_cols)

    section_row(ws_bs, R.next(), "ASSETS", n_m, annual_cols)
    section_row(ws_bs, R.next(), "Non-Current Assets", n_m, annual_cols, level=2)
    bsr("Property, Plant & Equipment  (net)", nbv_m, indent=1, link_=True)
    bsr("Goodwill & Intangibles", [0.0]*n_m, indent=1)
    bsr("Loans Receivable (non-current)", [0.0]*n_m, indent=1)
    bsr("Investments in Subsidiaries", [0.0]*n_m, indent=1)
    nca = nbv_m
    bsst("Total Non-Current Assets", nca)
    bssp()

    section_row(ws_bs, R.next(), "Current Assets", n_m, annual_cols, level=2)
    bsr("Trade Receivables", ar, indent=1, link_=True)
    bsr("Inventories — Finished Goods", inv_fg, indent=2, link_=True)
    bsr("Inventories — Raw Materials", inv_rm, indent=2, link_=True)
    bsr("Other Receivables & Prepayments", oth_ca, indent=1)
    bsr("Cash & Cash Equivalents", cash_cl, indent=1, bold=True, bg=C["check_ok"], link_=True)
    ca = [a+it+oca+cc for a,it,oca,cc in zip(ar,inv_tot,oth_ca,cash_cl)]
    bsst("Total Current Assets", ca)
    bssp()

    ta_vals_bs = [n+c for n,c in zip(nca,ca)]
    bskpi("TOTAL ASSETS", ta_vals_bs)
    bssp(); bssp()

    section_row(ws_bs, R.next(), "LIABILITIES & EQUITY", n_m, annual_cols)
    section_row(ws_bs, R.next(), "Non-Current Liabilities", n_m, annual_cols, level=2)
    bsr("Financial Debt (non-current)", debt_close, indent=1, link_=True)
    bsr("Accrued Interest", [0.0]*n_m, indent=1)
    bsr("Deferred Tax Liabilities", [0.0]*n_m, indent=1)
    bsr("Government Grants", [0.0]*n_m, indent=1)
    ncl = debt_close
    bsst("Total Non-Current Liabilities", ncl)
    bssp()

    section_row(ws_bs, R.next(), "Current Liabilities", n_m, annual_cols, level=2)
    bsr("Trade Payables", ap, indent=1, link_=True)
    bsr("Current Portion of Debt", [debt_open[i]-debt_close[i] for i in range(n_m)], indent=1, link_=True)
    bsr("Employee & Salary Provisions", emp_prov, indent=1, link_=True)
    bsr("Other Payables & Accruals", oth_cl, indent=1)
    cl = [p+ep+ocl for p,ep,ocl in zip(ap,emp_prov,oth_cl)]
    bsst("Total Current Liabilities", cl)
    bssp()

    tl = [n+c for n,c in zip(ncl,cl)]
    bstt("TOTAL LIABILITIES", tl)
    bssp()

    section_row(ws_bs, R.next(), "Equity", n_m, annual_cols, level=2)
    bsr("Share Capital & Share Premium", [0.0]*n_m, indent=1, bg=C["input_bg"])
    bsr("Retained Earnings (current period)", net_income, indent=1, link_=True)
    bsr("Accumulated Retained Earnings", accum_re, indent=1)
    eq = accum_re
    bsst("TOTAL EQUITY", eq)
    bssp()

    tle = [tl_+eq_ for tl_,eq_ in zip(tl,eq)]
    bskpi("TOTAL LIABILITIES & EQUITY", tle_vals)
    bssp()

    section_row(ws_bs, R.next(), "BALANCE SHEET CHECK", n_m, annual_cols, level=3)
    check = [abs(ta-tle_) for ta,tle_ in zip(ta_vals_bs,tle)]
    bsr("Assets − (Liabilities + Equity)  [should be 0]", check)

    # ══════════════════════════════════════════════════════════════════════════
    # 7.  CASH FLOW
    # ══════════════════════════════════════════════════════════════════════════
    ws_cf = setup_ws(wb, "CASH FLOW", "4A235A")
    set_col_widths(ws_cf, n_m, annual_cols)
    write_title(ws_cf, 1, company, "Cash Flow Statement", unit_lbl, LAST_COL)
    write_period_headers(ws_cf, 3, periods, annual_cols)
    R = RowCounter(3+HDR_ROWS+1)

    def cfr(label, vals, bold=False, bg="white", indent=0, link_=False, input_=False):
        r = R.next(); data_row(ws_cf, r, label, vals, unit=unit_lbl, bold=bold, bg=bg, indent=indent, link_style=link_, input_style=input_, annual_vals=av(vals), annual_cols=annual_cols); return r
    def cfst(label, vals): r = R.next(); subtotal_row(ws_cf, r, label, vals, unit=unit_lbl, annual_vals=av(vals), annual_cols=annual_cols); return r
    def cfkpi(label, vals): r = R.next(); kpi_row(ws_cf, r, label, vals, unit=unit_lbl, annual_vals=av(vals), annual_cols=annual_cols); return r
    def cfsp(): spacer_row(ws_cf, R.next(), n_m, annual_cols)

    section_row(ws_cf, R.next(), "A.  OPERATING CASH FLOW", n_m, annual_cols)
    cfr("Adjusted EBITDA", adj_ebitda, indent=1, link_=True)
    cfsp()
    section_row(ws_cf, R.next(), "  Change in Trade Working Capital", n_m, annual_cols, level=2)
    cfr("Change in Trade Receivables", [-c for c in chg_ar], indent=2, link_=True)
    cfr("Change in Inventories", [-c for c in chg_inv], indent=2, link_=True)
    chg_ap2 = [ap[0]]+[ap[i]-ap[i-1] for i in range(1,n_m)]
    cfr("Change in Trade Payables", chg_ap2, indent=2, link_=True)
    cfr("Change in Employee Provisions", [emp_prov[0]]+[emp_prov[i]-emp_prov[i-1] for i in range(1,n_m)], indent=2)
    cfr("Change in Other Working Capital", [oth_ca[0]-oth_cl[0]]+[(oth_ca[i]-oth_cl[i])-(oth_ca[i-1]-oth_cl[i-1]) for i in range(1,n_m)], indent=2)
    cfst("Total Change in Working Capital", [-c for c in chg_nwc])
    cfsp()
    cfr("Tax Paid (current)", [-t for t in tax_charge], indent=1, link_=True)
    cfr("Other Operating Items", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    cfkpi("CASH FLOW FROM OPERATIONS", op_cf)
    cfsp(); cfsp()

    section_row(ws_cf, R.next(), "B.  INVESTING CASH FLOW", n_m, annual_cols)
    cfr("Capital Expenditure", [-c for c in cpx_tot], indent=1, link_=True)
    cfr("Proceeds from Disposal of Assets", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    cfr("Acquisitions / Investments", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    cfr("Change in Loan Receivables", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    cfst("CASH FLOW FROM INVESTING", inv_cf)
    cfsp(); cfsp()

    cfkpi("FREE CASH FLOW  (pre-debt service)", fcf)
    cfsp(); cfsp()

    section_row(ws_cf, R.next(), "C.  FINANCING CASH FLOW", n_m, annual_cols)
    cfr("Interest Paid — Cash", int_cash, indent=1, link_=True)
    cfr("Debt Repayments", [-r for r in debt_repay], indent=1, link_=True)
    cfr("New Drawdowns / Refinancing", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    cfr("Equity Injections", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    cfr("Dividends / Equity Distributions", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    cfr("Government Grants Received", [0.0]*n_m, indent=1, input_=True, bg=C["input_bg"])
    cfst("CASH FLOW FROM FINANCING", fin_cf)
    cfsp(); cfsp()

    section_row(ws_cf, R.next(), "CLOSING CASH POSITION", n_m, annual_cols, level=2)
    cfr("Net Movement in Cash", net_cf, bold=True)
    cfr("Opening Cash Balance", cash_op, indent=1, bg=C["input_bg"], input_=True)
    cfkpi("CLOSING CASH BALANCE", cash_cl)

    # ══════════════════════════════════════════════════════════════════════════
    # 8.  KPIs
    # ══════════════════════════════════════════════════════════════════════════
    ws_kpi = setup_ws(wb, "KPIs & RATIOS", "2E7D32")
    set_col_widths(ws_kpi, n_m, annual_cols)
    write_title(ws_kpi, 1, company, "KPIs & Financial Ratios", unit_lbl, LAST_COL)
    write_period_headers(ws_kpi, 3, periods, annual_cols)
    R = RowCounter(3+HDR_ROWS+1)

    def kr(label, vals, fmt=FMT["int"], indent=0, unit="", avg=False):
        r = R.next()
        data_row(ws_kpi, r, label, vals, unit=unit, fmt=fmt, indent=indent,
                 annual_vals=av(vals,avg=avg), annual_cols=annual_cols)
        return r
    def ksp(): spacer_row(ws_kpi, R.next(), n_m, annual_cols)

    gm_m  = [gp/r if r else 0 for gp,r in zip(gross_profit,net_rev)]
    em_m  = [e/r if r else 0 for e,r in zip(adj_ebitda,net_rev)]
    eim_m = [e/r if r else 0 for e,r in zip(ebit,net_rev)]
    nm_m  = [n/r if r else 0 for n,r in zip(net_income,net_rev)]

    section_row(ws_kpi, R.next(), "1.  PROFITABILITY METRICS", n_m, annual_cols)
    kr("Net Revenue", net_rev, unit=unit_lbl)
    kr("YoY Revenue Growth", [0.0]+[(net_rev[i]-net_rev[i-1])/max(net_rev[i-1],1) for i in range(1,n_m)], fmt=FMT["pct1"], unit="%", avg=True)
    kr("Gross Profit", gross_profit, unit=unit_lbl)
    kr("Gross Margin %", gm_m, fmt=FMT["pct1"], unit="%", avg=True)
    kr("Adj. EBITDA", adj_ebitda, unit=unit_lbl)
    kr("Adj. EBITDA Margin %", em_m, fmt=FMT["pct1"], unit="%", avg=True)
    kr("EBIT", ebit, unit=unit_lbl)
    kr("EBIT Margin %", eim_m, fmt=FMT["pct1"], unit="%", avg=True)
    kr("Net Income", net_income, unit=unit_lbl)
    kr("Net Margin %", nm_m, fmt=FMT["pct1"], unit="%", avg=True)
    ksp()

    section_row(ws_kpi, R.next(), "2.  LEVERAGE & CREDIT", n_m, annual_cols, level=2)
    kr("Gross Debt", debt_close, unit=unit_lbl)
    kr("Cash & Cash Equivalents", cash_cl, unit=unit_lbl)
    kr("Net Debt", net_debt, unit=unit_lbl)
    ltm_ebitda = [sum(adj_ebitda[max(i-11,0):i+1])*12/max(min(i+1,12),1) for i in range(n_m)]
    kr("LTM Adj. EBITDA (annualised)", ltm_ebitda, unit=unit_lbl)
    kr("Net Leverage  (ND / LTM EBITDA)", [nd/max(ltm,0.001) for nd,ltm in zip(net_debt,ltm_ebitda)], fmt=FMT["mult"], unit="x", avg=True)
    kr("Interest Cover  (EBITDA / Interest)", [ae*12/max(i*12,0.001) for ae,i in zip(adj_ebitda,interest)], fmt=FMT["mult"], unit="x", avg=True)
    kr("DSCR  (Op.CF / Debt Service)", [oc/max(abs(fc),0.001) for oc,fc in zip(op_cf,fin_cf)], fmt=FMT["mult"], unit="x", avg=True)
    kr("Debt / Equity", [dc/max(abs(eq_),0.001) for dc,eq_ in zip(debt_close,eq)], fmt=FMT["mult"], unit="x", avg=True)
    ksp()

    section_row(ws_kpi, R.next(), "3.  EFFICIENCY & RETURNS", n_m, annual_cols, level=2)
    kr("Free Cash Flow", fcf, unit=unit_lbl)
    kr("FCF Conversion  (FCF / Adj EBITDA)", [f/max(abs(ae),0.001) for f,ae in zip(fcf,adj_ebitda)], fmt=FMT["pct1"], unit="%", avg=True)
    kr("Capex Intensity  (Capex / Revenue)", [c/max(r,0.001) for c,r in zip(cpx_tot,net_rev)], fmt=FMT["pct1"], unit="%", avg=True)
    kr("Working Capital Days (DSO+DIO-DPO)", [dso+dio_fg+dio_rm-dpo]*n_m, unit="days", avg=True)
    ksp()

    if biz == "industrial":
        section_row(ws_kpi, R.next(), "4.  INDUSTRIAL / OPERATIONAL KPIs", n_m, annual_cols, level=2)
        vol_m = [r/price_mt for r in net_rev]
        cap_m = [capacity_mt/12]*n_m
        util_m= [v/max(c,1) for v,c in zip(vol_m,cap_m)]
        spr_m = [r/max(v,1)-abs(d)/max(v,1) for r,v,d in zip(net_rev,vol_m,dm)]
        em_mt = [ae/max(v,1) for ae,v in zip(adj_ebitda,vol_m)]
        kr("Sales Volume", vol_m, fmt=FMT["dec1"], unit="MT/month")
        kr("Installed Capacity", cap_m, fmt=FMT["int"], unit="MT/month")
        kr("Capacity Utilisation", util_m, fmt=FMT["pct1"], unit="%", avg=True)
        kr("Avg Sales Price  ($/MT)", [price_mt]*n_m, fmt=FMT["usdmt"], unit="$/MT", avg=True)
        kr("Net Spread  ($/MT)", spr_m, fmt=FMT["usdmt"], unit="$/MT", avg=True)
        kr("Adj. EBITDA per MT", em_mt, fmt=FMT["usdmt"], unit="$/MT", avg=True)
        kr("Utility Cost  ($/MT)", [abs(u)/max(v,1) for u,v in zip(util_costs,vol_m)], fmt=FMT["usdmt"], unit="$/MT", avg=True)
        kr("Labour Cost  ($/MT)", [abs(s)/max(v,1) for s,v in zip(sal_total,vol_m)], fmt=FMT["usdmt"], unit="$/MT", avg=True)

    # ══════════════════════════════════════════════════════════════════════════
    # 9.  OUTPUT — print-ready annual summary
    # ══════════════════════════════════════════════════════════════════════════
    ws_out = wb.create_sheet("OUTPUT")
    ws_out.sheet_view.showGridLines = False
    ws_out.sheet_properties.tabColor = "0D1B3E"
    ws_out.sheet_view.zoomScale = 85
    ws_out.page_setup.paperSize   = 9
    ws_out.page_setup.orientation = "landscape"
    ws_out.page_setup.fitToPage   = True
    ws_out.page_setup.fitToWidth  = 1
    ws_out.page_margins.left=0.5; ws_out.page_margins.right=0.5
    ws_out.page_margins.top=0.7;  ws_out.page_margins.bottom=0.7
    ws_out.page_margins.header=0.3; ws_out.page_margins.footer=0.3
    ws_out.oddHeader.center.text = f"&\"Calibri,Bold\"&11{company}  —  Business Plan  |  Confidential"
    ws_out.oddFooter.left.text   = "&\"Calibri\"&8Prepared by MG Advisory  ·  Strictly Confidential & Not for Distribution"
    ws_out.oddFooter.right.text  = "&\"Calibri\"&8&D  ·  Page &P"

    OL = 2; OU = 3; ODC = 5
    n_oc = len(annual_cols)
    for c in [OL, OU] + list(range(ODC, ODC+n_oc+2)):
        ws_out.column_dimensions[col_l(c)].width = 2 if c == OL-1 else 44 if c==OL else 8 if c==OU else 13
    ws_out.column_dimensions["A"].width = 1.5

    # Title block
    ws_out.row_dimensions[1].height = 40
    ws_out.row_dimensions[2].height = 5
    for c in range(OL, ODC+n_oc+1):
        ws_out.cell(1,c).fill = F(C["hdr_year"])
    ws_out.cell(1,OL).value = f"  {company}"
    ws_out.cell(1,OL).font  = FK(C["txt_white"],bold=True,size=16)
    ws_out.cell(1,OL).alignment = AL("left")
    ws_out.cell(1,OU).value = "Business Plan Summary"
    ws_out.cell(1,OU).font  = FK("AAAAAA",size=9)
    ws_out.cell(1,OU).alignment = AL("left")

    # Column headers
    ws_out.row_dimensions[3].height = 18
    ws_out.cell(3,OL).value=""; ws_out.cell(3,OU).value="Unit"
    ws_out.cell(3,OU).font=FK(C["txt_grey"],size=8); ws_out.cell(3,OU).alignment=AL("center")
    for i,ac in enumerate(annual_cols):
        c_ = ws_out.cell(3, ODC+i)
        c_.value=ac["label"]; c_.fill=F(C["hdr_year"])
        c_.font=FK(C["txt_white"],bold=True,size=10); c_.alignment=AL("center")
        c_.border=Border(left=_side_med(),right=_side_med(),top=_side_med(),bottom=_side_med())

    R2 = RowCounter(4)
    def or_(label, vals, fmt=FMT["int"], bold=False, bg=None, indent=0, unit=""):
        r = R2.next(); ws_out.row_dimensions[r].height=13
        bg_ = bg or C["white"]
        c_=ws_out.cell(r,OL); c_.value=("   "*indent)+label
        c_.fill=F(bg_); c_.font=FK(C["txt_navy"] if bold else C["txt_black"],bold=bold,size=9); c_.alignment=AL("left")
        c_=ws_out.cell(r,OU); c_.value=unit; c_.fill=F(bg_); c_.font=FK(C["txt_grey"],size=8); c_.alignment=AL("center")
        for i,v in enumerate(vals):
            c_=ws_out.cell(r,ODC+i); c_.value=v; c_.fill=F(bg_)
            c_.font=FK(C["txt_navy"] if bold else C["txt_black"],bold=bold,size=9)
            c_.alignment=AL("right"); c_.number_format=fmt; c_.border=BORDER_THIN
        return r
    def ost(label, vals, fmt=FMT["int"], unit=""):
        r = R2.next(); ws_out.row_dimensions[r].height=14
        c_=ws_out.cell(r,OL); c_.value=label; c_.fill=F(C["subtot"])
        c_.font=FK(C["txt_navy"],bold=True,size=9); c_.border=BORDER_TOP; c_.alignment=AL("left")
        c_=ws_out.cell(r,OU); c_.value=unit; c_.fill=F(C["subtot"]); c_.font=FK(C["txt_grey"],size=8); c_.border=BORDER_TOP; c_.alignment=AL("center")
        for i,v in enumerate(vals):
            c_=ws_out.cell(r,ODC+i); c_.value=v; c_.fill=F(C["subtot"])
            c_.font=FK(C["txt_navy"],bold=True,size=9); c_.border=BORDER_TB
            c_.alignment=AL("right"); c_.number_format=fmt
        return r
    def okpi(label, vals, fmt=FMT["int"], unit=""):
        r = R2.next(); ws_out.row_dimensions[r].height=18
        for c_i in [OL,OU]+[ODC+i for i in range(n_oc)]: ws_out.cell(r,c_i).fill=F(C["ebitda_bg"])
        c_=ws_out.cell(r,OL); c_.value=label; c_.font=FK(C["txt_white"],bold=True,size=11); c_.alignment=AL("left")
        c_=ws_out.cell(r,OU); c_.value=unit; c_.font=FK("888888",size=8); c_.alignment=AL("center")
        for i,v in enumerate(vals):
            c_=ws_out.cell(r,ODC+i); c_.value=v
            c_.font=FK(C["txt_white"],bold=True,size=11); c_.alignment=AL("right"); c_.number_format=fmt
        return r
    def osec(label):
        r = R2.next(); ws_out.row_dimensions[r].height=14
        for c_i in range(OL,ODC+n_oc+1): ws_out.cell(r,c_i).fill=F(C["sec_2"])
        c_=ws_out.cell(r,OL); c_.value=f"  {label}"; c_.fill=F(C["sec_2"])
        c_.font=FK(C["txt_white"],bold=True,size=9); c_.alignment=AL("left")
    def osp(h=5):
        r = R2.next(); ws_out.row_dimensions[r].height=h
        for c_i in range(OL,ODC+n_oc+1): ws_out.cell(r,c_i).fill=F(C["white"])

    ann_rev  = [ann_sum(net_rev,ag,y)  for y in years]
    ann_gp   = [ann_sum(gross_profit,ag,y) for y in years]
    ann_ae   = [ann_sum(adj_ebitda,ag,y) for y in years]
    ann_ebit = [ann_sum(ebit,ag,y)     for y in years]
    ann_ni   = [ann_sum(net_income,ag,y) for y in years]
    ann_fcf  = [ann_sum(fcf,ag,y)      for y in years]
    ann_opcf = [ann_sum(op_cf,ag,y)    for y in years]
    ann_cpx  = [ann_sum(cpx_tot,ag,y)  for y in years]
    ann_tax  = [ann_sum(tax_charge,ag,y) for y in years]
    ann_int  = [ann_sum(interest,ag,y) for y in years]
    ann_dep  = [ann_sum(dep_m,ag,y)    for y in years]
    rev_g    = [0.0]+[(ann_rev[i]-ann_rev[i-1])/max(ann_rev[i-1],1) for i in range(1,n_ac)]

    # Use last month of each year for balance sheet items
    last_idx = [ag[y][-1] for y in years]
    ann_cash = [cash_cl[i] for i in last_idx]
    ann_debt = [debt_close[i] for i in last_idx]
    ann_nd   = [net_debt[i]  for i in last_idx]
    ann_ta   = [ta_vals_bs[i]  for i in last_idx]
    ann_eq   = [eq[i]        for i in last_idx]
    ann_lev  = [nd/max(ae,0.001) for nd,ae in zip(ann_nd,ann_ae)]
    ann_icr  = [ae/max(i,0.001) for ae,i in zip(ann_ae,ann_int)]

    osec("INCOME STATEMENT SUMMARY")
    or_("Net Revenue",   ann_rev,  bold=True, unit=unit_lbl)
    or_("  YoY Growth",  rev_g,    fmt=FMT["pct1"], indent=1, unit="%")
    or_("Gross Profit",  ann_gp,   unit=unit_lbl)
    or_("  Gross Margin %", [gp/r if r else 0 for gp,r in zip(ann_gp,ann_rev)], fmt=FMT["pct1"], indent=1, unit="%")
    osp()
    or_("  Variable Costs", [ann_sum(v,ag,y) for v,y in zip([var_costs]*n_ac,years)], indent=1, unit=unit_lbl)
    or_("  Fixed Costs",    [ann_sum(v,ag,y) for v,y in zip([fixed_total]*n_ac,years)], indent=1, unit=unit_lbl)
    okpi(f"ADJ. EBITDA  ({unit_lbl})", ann_ae, unit=unit_lbl)
    or_("  Adj. EBITDA Margin %", [ae/r if r else 0 for ae,r in zip(ann_ae,ann_rev)], fmt=FMT["pct1"], indent=1, unit="%")
    osp()
    or_("D&A", [-d for d in ann_dep], indent=1, unit=unit_lbl)
    or_("Net Finance Costs", [-i for i in ann_int], indent=1, unit=unit_lbl)
    or_("Tax", [-t for t in ann_tax], indent=1, unit=unit_lbl)
    osp()
    okpi(f"NET INCOME  ({unit_lbl})", ann_ni, unit=unit_lbl)
    or_("  Net Margin %", [ni/r if r else 0 for ni,r in zip(ann_ni,ann_rev)], fmt=FMT["pct1"], indent=1, unit="%")
    osp(); osp()

    osec("CASH FLOW SUMMARY")
    okpi(f"FREE CASH FLOW  ({unit_lbl})", ann_fcf, unit=unit_lbl)
    or_("  Operating CF", ann_opcf, indent=1, unit=unit_lbl)
    or_("  Capex", [-c for c in ann_cpx], indent=1, unit=unit_lbl)
    or_("  Financing CF", [ann_sum(fin_cf,ag,y) for y in years], indent=1, unit=unit_lbl)
    ost("Closing Cash Balance", ann_cash, unit=unit_lbl)
    osp(); osp()

    osec("LEVERAGE & CREDIT METRICS")
    or_("Gross Debt",      ann_debt, unit=unit_lbl)
    or_("Net Debt",        ann_nd,   bold=True, unit=unit_lbl)
    or_("Adj. EBITDA (annual)", ann_ae, unit=unit_lbl)
    or_("Net Leverage  (x)", ann_lev, fmt=FMT["mult"], unit="x")
    or_("Interest Cover  (x)", ann_icr, fmt=FMT["mult"], unit="x")
    osp(); osp()

    osec("BALANCE SHEET HIGHLIGHTS  (year-end)")
    or_("Total Assets",    ann_ta, bold=True, unit=unit_lbl)
    or_("  Cash",          ann_cash, indent=1, unit=unit_lbl)
    or_("  PP&E (net)",    [nbv_m[i] for i in last_idx], indent=1, unit=unit_lbl)
    or_("Total Equity",    ann_eq, bold=True, unit=unit_lbl)
    or_("Equity / Assets", [eq_/max(ta,0.001) for eq_,ta in zip(ann_eq,ann_ta)], fmt=FMT["pct1"], unit="%")

    # Print area
    ws_out.print_area = f"A1:{col_l(ODC+n_oc)}{R2.peek()+2}"

    # ══════════════════════════════════════════════════════════════════════════
    # 10.  SCENARIOS
    # ══════════════════════════════════════════════════════════════════════════
    if config.get("scenarios") == "all":
        for sc_name, rm, em in [("LOW CASE",0.85,0.75),("BEST CASE",1.15,1.25)]:
            ws_sc = wb.create_sheet(sc_name)
            ws_sc.sheet_view.showGridLines = False
            ws_sc.sheet_properties.tabColor = "8B2020" if "LOW" in sc_name else "1A5C2A"
            ws_sc.sheet_view.zoomScale = 85
            ws_sc.page_setup.paperSize=9; ws_sc.page_setup.orientation="landscape"
            ws_sc.page_setup.fitToPage=True; ws_sc.page_setup.fitToWidth=1
            ws_sc.column_dimensions["A"].width=1.5
            ws_sc.column_dimensions[col_l(OL)].width=44
            ws_sc.column_dimensions[col_l(OU)].width=8
            for i in range(n_oc): ws_sc.column_dimensions[col_l(ODC+i)].width=13

            sc_rev = [r*rm for r in ann_rev]
            sc_ae  = [a*em for a in ann_ae]
            sc_ni  = [a*em*0.65 for a in ann_ae]
            sc_nd  = [nd*(2-rm) for nd in ann_nd]
            sc_fcf = [f*em for f in ann_fcf]

            ws_sc.row_dimensions[1].height=32
            for c in range(OL,ODC+n_oc+1): ws_sc.cell(1,c).fill=F(C["hdr_year"])
            ws_sc.cell(1,OL).value=f"  {company}  —  {sc_name}"
            ws_sc.cell(1,OL).font=FK(C["txt_white"],bold=True,size=13); ws_sc.cell(1,OL).alignment=AL("left")
            ws_sc.row_dimensions[2].height=5
            ws_sc.row_dimensions[3].height=18
            ws_sc.cell(3,OL).value="Key Metrics"; ws_sc.cell(3,OU).value="Unit"
            ws_sc.cell(3,OU).font=FK(C["txt_grey"],size=8); ws_sc.cell(3,OU).alignment=AL("center")
            for i,ac in enumerate(annual_cols):
                c_=ws_sc.cell(3,ODC+i); c_.value=ac["label"]
                c_.fill=F(C["hdr_year"]); c_.font=FK(C["txt_white"],bold=True,size=10); c_.alignment=AL("center")

            R3=RowCounter(4)
            def scrow(label, vals, fmt=FMT["int"], bold=False, bg=None, unit=""):
                r=R3.next(); ws_sc.row_dimensions[r].height=14
                bg_=bg or C["white"]
                c_=ws_sc.cell(r,OL); c_.value=label; c_.fill=F(bg_)
                c_.font=FK(C["txt_navy"] if bold else C["txt_black"],bold=bold,size=9); c_.alignment=AL("left")
                c_=ws_sc.cell(r,OU); c_.value=unit; c_.fill=F(bg_); c_.font=FK(C["txt_grey"],size=8); c_.alignment=AL("center")
                for i,v in enumerate(vals):
                    c_=ws_sc.cell(r,ODC+i); c_.value=v; c_.fill=F(bg_)
                    c_.font=FK(C["txt_navy"] if bold else C["txt_black"],bold=bold,size=9)
                    c_.alignment=AL("right"); c_.number_format=fmt; c_.border=BORDER_THIN
            def sckpi(label, vals, fmt=FMT["int"], unit=""):
                r=R3.next(); ws_sc.row_dimensions[r].height=18
                for c_i in [OL,OU]+[ODC+i for i in range(n_oc)]: ws_sc.cell(r,c_i).fill=F(C["ebitda_bg"])
                c_=ws_sc.cell(r,OL); c_.value=label; c_.font=FK(C["txt_white"],bold=True,size=11); c_.alignment=AL("left")
                c_=ws_sc.cell(r,OU); c_.value=unit; c_.font=FK("888888",size=8); c_.alignment=AL("center")
                for i,v in enumerate(vals):
                    c_=ws_sc.cell(r,ODC+i); c_.value=v; c_.font=FK(C["txt_white"],bold=True,size=11)
                    c_.alignment=AL("right"); c_.number_format=fmt

            sckpi("Net Revenue", sc_rev, unit=unit_lbl)
            scrow("  vs Base Case", [sc-b for sc,b in zip(sc_rev,ann_rev)], fmt=FMT["int"], unit=unit_lbl)
            scrow("  vs Base Case %", [(sc-b)/max(b,1) for sc,b in zip(sc_rev,ann_rev)], fmt=FMT["pct1"], unit="%")
            R3.skip()
            sckpi("Adj. EBITDA", sc_ae, unit=unit_lbl)
            scrow("  EBITDA Margin %", [ae/r if r else 0 for ae,r in zip(sc_ae,sc_rev)], fmt=FMT["pct1"], unit="%")
            scrow("  vs Base Case", [sc-b for sc,b in zip(sc_ae,ann_ae)], fmt=FMT["int"], unit=unit_lbl)
            R3.skip()
            sckpi("Free Cash Flow", sc_fcf, unit=unit_lbl)
            sckpi("Net Income", sc_ni, unit=unit_lbl)
            R3.skip()
            scrow("Net Debt", sc_nd, bold=True, unit=unit_lbl)
            scrow("Net Leverage (x)", [nd/max(ae,0.001) for nd,ae in zip(sc_nd,sc_ae)], fmt=FMT["mult"], unit="x")

    wb.save(output_path)
    return output_path
