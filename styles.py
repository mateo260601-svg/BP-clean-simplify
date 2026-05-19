from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colours ──────────────────────────────────────────────────────────────────
NAVY        = "0D1B3E"
NAVY_LIGHT  = "1A2F5A"
BLUE_HEADER = "1F3D7A"
BLUE_LIGHT  = "D6E4F0"
BLUE_MID    = "A8C8E8"
GREY_LIGHT  = "F5F5F5"
GREY_MID    = "D9D9D9"
GREY_DARK   = "BFBFBF"
WHITE       = "FFFFFF"
YELLOW_ASS  = "FFFACD"
GREEN_POS   = "E8F5E9"
RED_NEG     = "FFEBEE"
ORANGE_WARN = "FFF3E0"

# ── Fills ─────────────────────────────────────────────────────────────────────
fill_navy        = PatternFill("solid", fgColor=NAVY)
fill_navy_light  = PatternFill("solid", fgColor=NAVY_LIGHT)
fill_blue_header = PatternFill("solid", fgColor=BLUE_HEADER)
fill_blue_light  = PatternFill("solid", fgColor=BLUE_LIGHT)
fill_blue_mid    = PatternFill("solid", fgColor=BLUE_MID)
fill_grey_light  = PatternFill("solid", fgColor=GREY_LIGHT)
fill_grey_mid    = PatternFill("solid", fgColor=GREY_MID)
fill_white       = PatternFill("solid", fgColor=WHITE)
fill_yellow      = PatternFill("solid", fgColor=YELLOW_ASS)
fill_green       = PatternFill("solid", fgColor=GREEN_POS)
fill_red         = PatternFill("solid", fgColor=RED_NEG)
fill_orange      = PatternFill("solid", fgColor=ORANGE_WARN)

# ── Fonts ─────────────────────────────────────────────────────────────────────
font_title       = Font(name="Calibri", size=14, bold=True,  color=WHITE)
font_header      = Font(name="Calibri", size=10, bold=True,  color=WHITE)
font_header_navy = Font(name="Calibri", size=10, bold=True,  color=NAVY)
font_section     = Font(name="Calibri", size=10, bold=True,  color=NAVY)
font_bold        = Font(name="Calibri", size=10, bold=True,  color="000000")
font_normal      = Font(name="Calibri", size=10, bold=False, color="000000")
font_small       = Font(name="Calibri", size=9,  bold=False, color="595959")
font_italic      = Font(name="Calibri", size=10, italic=True, color="595959")
font_input       = Font(name="Calibri", size=10, bold=False, color="00008B")
font_total       = Font(name="Calibri", size=10, bold=True,  color=NAVY)
font_negative    = Font(name="Calibri", size=10, bold=False, color="CC0000")
font_positive    = Font(name="Calibri", size=10, bold=False, color="006400")

# ── Alignments ───────────────────────────────────────────────────────────────
align_left    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
align_center  = Alignment(horizontal="center", vertical="center", wrap_text=False)
align_right   = Alignment(horizontal="right",  vertical="center", wrap_text=False)
align_wrap    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Borders ───────────────────────────────────────────────────────────────────
_thin   = Side(style="thin",   color="BFBFBF")
_medium = Side(style="medium", color=NAVY)
_thick  = Side(style="thick",  color=NAVY)
_none   = Side(style=None)

border_all      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
border_bottom   = Border(bottom=_thin)
border_top      = Border(top=_thin)
border_outer    = Border(left=_medium, right=_medium, top=_medium, bottom=_medium)
border_top_bold = Border(top=_medium)
border_none     = Border(left=_none, right=_none, top=_none, bottom=_none)

# ── Number formats ────────────────────────────────────────────────────────────
FMT_INT        = "#,##0"
FMT_INT_NEG    = "#,##0;(#,##0)"
FMT_DEC1       = "#,##0.0"
FMT_DEC2       = "#,##0.00"
FMT_PCT0       = "0%"
FMT_PCT1       = "0.0%"
FMT_PCT2       = "0.00%"
FMT_MULT       = "0.0x"
FMT_EUR        = u'#,##0\u20ac'
FMT_USD        = "$#,##0"
FMT_GBP        = u'\xa3#,##0'
FMT_DATE       = "DD/MM/YYYY"
FMT_YEAR       = "YYYY"
FMT_TEXT       = "@"

# ── Helpers ───────────────────────────────────────────────────────────────────

def apply_style(cell, fill=None, font=None, alignment=None,
                border=None, number_format=None):
    if fill:            cell.fill            = fill
    if font:            cell.font            = font
    if alignment:       cell.alignment       = alignment
    if border:          cell.border          = border
    if number_format:   cell.number_format   = number_format


def style_header_row(ws, row, col_start, col_end,
                     fill=None, font=None, height=18):
    _fill = fill or fill_blue_header
    _font = font or font_header
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = _fill
        cell.font      = _font
        cell.alignment = align_center
        cell.border    = border_all
    if height:
        ws.row_dimensions[row].height = height


def style_section_row(ws, row, col_start, col_end, label=None, height=16):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = fill_navy
        cell.font      = font_title if col == col_start else font_header
        cell.alignment = align_left if col == col_start else align_center
        cell.border    = border_all
    if label and col_start:
        ws.cell(row=row, column=col_start).value = label
    if height:
        ws.row_dimensions[row].height = height


def style_total_row(ws, row, col_start, col_end, height=15):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = fill_blue_light
        cell.font   = font_total
        cell.border = border_top_bold
    if height:
        ws.row_dimensions[row].height = height


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def freeze(ws, cell_ref="B5"):
    ws.freeze_panes = cell_ref


def hide_gridlines(ws):
    ws.sheet_view.showGridLines = False
