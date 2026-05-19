"""
sheet_debt_v2.py  —  Full debt schedule sheet builder
Uses DebtEngine for per-tranche and consolidated waterfall.
"""
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

FONT_NAME = "Calibri"
def F(hex_): return PatternFill("solid", fgColor=hex_)
def FK(color="000000", bold=False, size=9):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)
def AL(h="left", v="center"): return Alignment(horizontal=h, vertical=v)
def side(style="thin", color="D0D5DD"): return Side(style=style, color=color)
BORDER_TOP  = Border(top=Side(style="medium", color="0D1B3E"))
BORDER_TB   = Border(top=Side(style="medium",color="0D1B3E"), bottom=Side(style="medium",color="0D1B3E"))
BORDER_THIN = Border(left=side(),right=side(),top=side(),bottom=side())

CAT_COLORS = {
    "Senior Secured":    "1A3A6C",
    "Second Lien":       "2D6A9F",
    "Mezzanine":         "7B2D8B",
    "Capital Markets":   "B45309",
    "Equity-Linked":     "6B21A8",
    "Islamic Finance":   "065F46",
    "Structured Finance":"0E7490",
    "Restructuring":     "991B1B",
    "Leasing":           "374151",
}

def build_debt_sheet(wb, engine, config, periods, annual_cols, LCOL, UCOL, DCOL, n_m):
    from debt_engine import DEBT_CATALOGUE
    company  = config.get("company_name", "Company")
    unit_lbl = f"{config.get('currency','USD')}k"
    LAST_COL = DCOL + n_m + len(annual_cols) + 2

    ws = wb.create_sheet("DEBT SCHEDULE")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8B2020"
    ws.sheet_view.zoomScale = 75
    ws.freeze_panes = f"{get_column_letter(DCOL)}7"
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.oddHeader.right.text = "&\"Calibri\"&9Confidential"
    ws.oddFooter.left.text  = "&\"Calibri\"&8MG Advisory  ·  Strictly Confidential"
    ws.oddFooter.right.text = "&\"Calibri\"&8&D  ·  Page &P"

    ws.column_dimensions["A"].width = 1.5
    ws.column_dimensions[get_column_letter(LCOL)].width = 48
    ws.column_dimensions[get_column_letter(UCOL)].width = 8
    ws.column_dimensions[get_column_letter(DCOL-1)].width = 1
    for i in range(n_m):
        ws.column_dimensions[get_column_letter(DCOL+i)].width = 10.5
    for ac in annual_cols:
        ws.column_dimensions[get_column_letter(ac["col"])].width = 13

    # Title
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 4
    for c in range(LCOL, LAST_COL+1):
        ws.cell(1, c).fill = F("0D1B3E")
        ws.cell(2, c).fill = F("3A6BBF")
    cell = ws.cell(1, LCOL)
    cell.value = f"  {company}   ·   Debt Schedule & Covenant Tracking   ·   {unit_lbl}"
    cell.font  = FK("FFFFFF", bold=True, size=11)
    cell.alignment = AL("left")

    # Period headers (4 rows)
    from build_v3 import write_period_headers
    write_period_headers(ws, 3, periods, annual_cols)

    MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    def ag_sum(vals, ag_dict, year):
        return sum(vals[i] for i in ag_dict.get(year, []) if i < len(vals))

    from build_v4 import gen_periods, annual_groups
    ag_dict = annual_groups(periods)
    years = sorted(ag_dict.keys())

    def av(vals):
        return [ag_sum(vals, ag_dict, y) for y in years]

    def sc(row, col, value=None, fill_=None, font_=None, align_=None, fmt=None, border=None, height=None):
        c = ws.cell(row, col)
        if value is not None: c.value = value
        if fill_:  c.fill      = fill_
        if font_:  c.font      = font_
        if align_: c.alignment = align_
        if fmt:    c.number_format = fmt
        if border: c.border    = border
        if height: ws.row_dimensions[row].height = height
        return c

    FMT_INT  = '#,##0_);(#,##0);"-"'
    FMT_PCT1 = '0.0%_);(0.0%);"-"'
    FMT_MULT = '0.0"x"'
    FMT_PCT0 = '0%'

    def write_row(row, label, vals, bold=False, bg="FFFFFF", fmt=FMT_INT, indent=0, link=False, input_=False, unit="", height=13):
        ws.row_dimensions[row].height = height
        txt = ("    "*indent) + label
        tc  = "0000CD" if input_ else "006400" if link else "0D1B3E" if bold else "000000"
        bg_ = bg if input_ else bg
        lc = sc(row, LCOL, txt, F(bg_), FK(tc, bold, 9), AL("left"))
        sc(row, UCOL, unit, F(bg_), FK("718096",size=8), AL("center"))
        for i, v in enumerate(vals):
            c = ws.cell(row, DCOL+i)
            c.value = v; c.fill = F(bg_)
            c.font  = FK(tc, bold, 9); c.alignment = AL("right"); c.number_format = fmt
            c.border = Border(bottom=Side(style="hair", color="E2E8F0"))
        for j, ac in enumerate(annual_cols):
            c = ws.cell(row, ac["col"])
            c.value = av(vals)[j]; c.fill = F("EFF6FF")
            c.font  = FK(tc, bold, 9); c.alignment = AL("right"); c.number_format = fmt
            c.border = BORDER_THIN

    def subtotal_row(row, label, vals, fmt=FMT_INT, unit=""):
        ws.row_dimensions[row].height = 14
        sc(row, LCOL, "  "+label, F("DDEEFF"), FK("0D1B3E",True,9), AL("left"), border=BORDER_TOP)
        sc(row, UCOL, unit, F("DDEEFF"), FK("718096",size=8), AL("center"), border=BORDER_TOP)
        for i, v in enumerate(vals):
            c = ws.cell(row, DCOL+i)
            c.value=v; c.fill=F("DDEEFF"); c.font=FK("0D1B3E",True,9)
            c.alignment=AL("right"); c.number_format=fmt; c.border=BORDER_TOP
        for j, ac in enumerate(annual_cols):
            c = ws.cell(row, ac["col"])
            c.value=av(vals)[j]; c.fill=F("C5DCF5"); c.font=FK("0D1B3E",True,9)
            c.alignment=AL("right"); c.number_format=fmt; c.border=BORDER_TB

    def kpi_row(row, label, vals, fmt=FMT_INT, unit="", height=16):
        ws.row_dimensions[row].height = height
        for c in [LCOL,UCOL]+[DCOL+i for i in range(n_m)]+[ac["col"] for ac in annual_cols]:
            ws.cell(row,c).fill = F("0D1B3E")
        sc(row, LCOL, label, F("0D1B3E"), FK("FFFFFF",True,10), AL("left"))
        sc(row, UCOL, unit, F("0D1B3E"), FK("AAAAAA",size=8), AL("center"))
        for i,v in enumerate(vals):
            c=ws.cell(row,DCOL+i); c.value=v; c.fill=F("0D1B3E")
            c.font=FK("FFFFFF",True,10); c.alignment=AL("right"); c.number_format=fmt
        for j,ac in enumerate(annual_cols):
            c=ws.cell(row,ac["col"]); c.value=av(vals)[j]; c.fill=F("1A3A5C")
            c.font=FK("FFFFFF",True,10); c.alignment=AL("right"); c.number_format=fmt

    def section_hdr(row, label, color="0D1B3E", height=15):
        ws.row_dimensions[row].height = height
        for c in range(LCOL, LAST_COL+1):
            ws.cell(row,c).fill = F(color)
        cell = ws.cell(row, LCOL)
        cell.value = f"  {label}"
        cell.font  = FK("FFFFFF", bold=True, size=9)
        cell.alignment = AL("left")

    def spacer(row, height=5):
        ws.row_dimensions[row].height = height
        for c in range(LCOL, LAST_COL+1):
            ws.cell(row,c).fill = F("FFFFFF")

    row = 3 + 4 + 1  # after title + 4 header rows + spacer

    # ── DEBT SUMMARY ──────────────────────────────────────────────────────────
    summary = engine.summary()
    section_hdr(row, "DEBT CAPITAL STRUCTURE  —  OVERVIEW"); row += 1
    write_row(row, "Total Committed Facilities", [summary["total_debt_committed"]]*n_m,
              bold=True, bg="F0F4FF", unit=unit_lbl); row += 1
    write_row(row, "Total Drawn Debt", [summary["total_debt_drawn"]]*n_m,
              bg="F0F4FF", unit=unit_lbl); row += 1
    write_row(row, "Total Upfront Fees & OID",
              [-(summary["total_upfront_fees"]+summary["total_oid"])]*n_m,
              bg="F0F4FF", unit=unit_lbl); row += 1
    write_row(row, "Net Proceeds",
              [summary["total_debt_drawn"]-summary["total_upfront_fees"]-summary["total_oid"]]*n_m,
              bold=True, bg="EFF6FF", unit=unit_lbl); row += 1
    spacer(row); row += 1

    # ── PER TRANCHE SCHEDULES ──────────────────────────────────────────────────
    for idx, (sch_key, sch) in enumerate(engine.schedules.items()):
        tranche = engine.tranches[idx]
        cat     = DEBT_CATALOGUE.get(tranche.key, {}).get("category", "Senior Secured")
        color   = CAT_COLORS.get(cat, "1A3A6C")

        section_hdr(row, f"TRANCHE {idx+1}  |  {tranche.display_name}  |  {unit_lbl}{tranche.amount:,.0f}  |  {tranche.currency}",
                    color=color); row += 1

        # Parameters info row
        ws.row_dimensions[row].height = 12
        if tranche.cash_rate > 0:
            rate_str = f"Fixed {tranche.cash_rate*100:.2f}%"
        else:
            from debt_engine import BASE_RATES
            base = BASE_RATES.get(tranche.base_rate_type, 0)
            rate_str = f"{tranche.base_rate_type} ({base*100:.2f}%) + {tranche.margin*100:.2f}% = {(base+tranche.margin)*100:.2f}%"
            if tranche.floor > 0:
                rate_str += f"  (floor {tranche.floor*100:.2f}%)"
        pik_str = f"  |  PIK {tranche.pik_rate*100:.2f}%" if tranche.pik_rate > 0 else ""
        oid_str = f"  |  OID {tranche.oid_pct*100:.1f}%" if tranche.oid_pct > 0 else ""
        amort_str = tranche.amort.capitalize()
        info_txt = f"    Cash: {rate_str}{pik_str}{oid_str}  |  Amort: {amort_str}  |  Tenor: {tranche.tenor_months//12}yr  |  Grace: {tranche.grace_months}m"
        cell = ws.cell(row, LCOL)
        cell.value = info_txt
        cell.font  = FK("374151", size=7)
        cell.alignment = AL("left")
        for c in range(LCOL, LAST_COL+1):
            ws.cell(row, c).fill = F("F8FAFF")
        row += 1

        write_row(row, "Opening Balance", sch["opening"], bg="EBF5FF", input_=True, unit=unit_lbl, indent=1); row += 1
        write_row(row, "Drawdown", [sch["opening"][m]-sch["opening"][max(m-1,0)] if m>0 and sch["opening"][m]>sch["opening"][m-1] else (sch["opening"][m] if m==0 and sch["opening"][0]>0 else 0) for m in range(n_m)],
                  unit=unit_lbl, indent=2); row += 1
        write_row(row, "Scheduled Repayment", [-r for r in sch["repayment"]], unit=unit_lbl, indent=2); row += 1
        write_row(row, "PIK Capitalisation", sch["pik_accrual"], unit=unit_lbl, indent=2); row += 1
        subtotal_row(row, "Closing Balance", sch["closing"], unit=unit_lbl); row += 1

        write_row(row, "Interest — Cash Pay", [-i for i in sch["interest_cash"]], unit=unit_lbl, indent=1); row += 1
        write_row(row, "Interest — PIK", [-i for i in sch["interest_pik"]], unit=unit_lbl, indent=1); row += 1
        if any(f > 0 for f in sch["commitment_fee"]):
            write_row(row, "Commitment Fee (undrawn)", [-f for f in sch["commitment_fee"]], unit=unit_lbl, indent=1); row += 1
        if any(o > 0 for o in sch["oid_amort"]):
            write_row(row, "OID Amortisation", [-o for o in sch["oid_amort"]], unit=unit_lbl, indent=1); row += 1
        subtotal_row(row, "Total Finance Cost (incl. PIK)", [-i for i in sch["interest_total"]], unit=unit_lbl); row += 1
        spacer(row); row += 1

    # ── CONSOLIDATED WATERFALL ─────────────────────────────────────────────────
    section_hdr(row, "CONSOLIDATED DEBT WATERFALL", color="0D1B3E"); row += 1
    tot = engine.totals

    write_row(row, "Opening Gross Debt", tot["opening"], bold=True, bg="EBF5FF", input_=True, unit=unit_lbl); row += 1
    write_row(row, "Total Repayments (scheduled)", [-r for r in tot["repayment"]], unit=unit_lbl, indent=1); row += 1
    write_row(row, "PIK Capitalisation (all tranches)", tot.get("pik_accrual",[0]*n_m), unit=unit_lbl, indent=1); row += 1
    subtotal_row(row, "Closing Gross Debt", tot["closing"], unit=unit_lbl); row += 1
    spacer(row); row += 1

    section_hdr(row, "FINANCE COST BREAKDOWN", color="1F3D7A"); row += 1
    write_row(row, "Interest — Cash Pay", [-i for i in tot["interest_cash"]], unit=unit_lbl, indent=1); row += 1
    write_row(row, "Interest — PIK / Accrued", [-i for i in tot["interest_pik"]], unit=unit_lbl, indent=1); row += 1
    write_row(row, "Commitment Fees (undrawn)", [-f for f in tot["commitment_fee"]], unit=unit_lbl, indent=1); row += 1
    write_row(row, "OID Amortisation", [-o for o in tot["oid_amort"]], unit=unit_lbl, indent=1); row += 1
    kpi_row(row, f"TOTAL FINANCE COSTS  ({unit_lbl})", [-i for i in tot["interest_total"]], unit=unit_lbl); row += 1
    spacer(row); row += 1

    # ── COVENANT TRACKING ─────────────────────────────────────────────────────
    section_hdr(row, "COVENANT TRACKING & CREDIT METRICS", color="1F3D7A"); row += 1

    ebitda_m = config.get("_ebitda_m", [1000.0]*n_m)
    cash_m   = config.get("_cash_m",   [5000.0]*n_m)

    net_debt = [tot["closing"][m] - cash_m[m] for m in range(n_m)]
    # LTM EBITDA
    ltm_ebitda = [sum(ebitda_m[max(m-11,0):m+1])*12/max(min(m+1,12),1) for m in range(n_m)]
    lev = [nd/max(le,0.001) for nd,le in zip(net_debt,ltm_ebitda)]
    icr = [ae*12/max(tot["interest_total"][m]*12,0.001) for m,ae in enumerate(ebitda_m)]
    dscr= [ae/max(tot["repayment"][m]+tot["interest_cash"][m],0.001) for m,ae in enumerate(ebitda_m)]

    write_row(row, "Gross Debt (closing)", tot["closing"], unit=unit_lbl, indent=1, link=True); row += 1
    write_row(row, "Cash & Equivalents", cash_m, unit=unit_lbl, indent=1, link=True); row += 1
    subtotal_row(row, "Net Debt", net_debt, unit=unit_lbl); row += 1
    write_row(row, "LTM Adj. EBITDA (annualised)", ltm_ebitda, unit=unit_lbl, indent=1, link=True); row += 1
    write_row(row, "Net Leverage  (ND / LTM EBITDA)", lev, fmt=FMT_MULT, unit="x", indent=1); row += 1
    write_row(row, "  Leverage Covenant  (max)", config.get("covenant_lev_max",[6.0]*n_m) if isinstance(config.get("covenant_lev_max"),list) else [config.get("covenant_lev_max",6.0)]*n_m,
              fmt=FMT_MULT, unit="x", indent=2, input_=True, bg="EBF5FF"); row += 1
    write_row(row, "  Headroom to Leverage Covenant", [config.get("covenant_lev_max",6.0)-l for l in lev],
              fmt=FMT_MULT, unit="x", indent=2); row += 1
    write_row(row, "Interest Cover  (EBITDA / Cash Interest)", icr, fmt=FMT_MULT, unit="x", indent=1); row += 1
    write_row(row, "  ICR Covenant  (min)", [config.get("covenant_icr_min",2.5)]*n_m,
              fmt=FMT_MULT, unit="x", indent=2, input_=True, bg="EBF5FF"); row += 1
    write_row(row, "DSCR  (EBITDA / Debt Service)", dscr, fmt=FMT_MULT, unit="x", indent=1); row += 1
    spacer(row); row += 1

    # ── MATURITY PROFILE ──────────────────────────────────────────────────────
    section_hdr(row, "MATURITY PROFILE BY TRANCHE  (repayments per month)", color="1F3D7A"); row += 1
    for idx, (sch_key, sch) in enumerate(engine.schedules.items()):
        t = engine.tranches[idx]
        write_row(row, t.display_name, [-r for r in sch["repayment"]],
                  unit=unit_lbl, indent=1); row += 1
    subtotal_row(row, "Total Scheduled Repayments", [-r for r in tot["repayment"]], unit=unit_lbl); row += 1

    ws.print_area = f"A1:{get_column_letter(LAST_COL)}{row}"
    return ws
